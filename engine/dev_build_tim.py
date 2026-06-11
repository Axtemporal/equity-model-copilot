"""Builds the TELECOM input template and fills it with TIM's reported actuals.

Data source: the 'Inputs' tab of a local reference workbook, not under version
control (data as published by TIM, quarterly). Extracts ONLY historical quarters
(1Q23 to 3Q25); never estimates. Values in R$ thousand converted to R$ mn. 4Q25 and
1Q26 have already been reported in the real world but are missing from the local
source: known gap.

Usage: python engine/dev_build_tim.py <path to local source workbook>
Outputs: templates/input_template_telecom.xlsx and inputs/TIM_inputs.xlsx
"""

from pathlib import Path
import shutil
import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from build_input_template import (PERIODS, FIRST_DATA_COL, HEADER_ROW, NUM_MN,
                                  NUM_1D, NUM_2D, NUM_FX, TITLE, SUB, BOLD,
                                  sheet_titles, period_header, section, data_row,
                                  build_assumptions)

BLUE = Font(name="Arial", size=10, color="0000FF")
PCT = '0.0%;(0.0%);"-"'
HIST_QUARTERS = [f"{q}Q{y}" for y in (23, 24) for q in (1, 2, 3, 4)] + ["1Q25", "2Q25", "3Q25"]

IS_LINES = [
    ("(=) Net revenue", "R$ mn", "sum", NUM_MN),
    ("memo: Mobile service revenue", "R$ mn", "sum", NUM_MN),
    ("memo: Fixed service revenue", "R$ mn", "sum", NUM_MN),
    ("memo: Product revenue", "R$ mn", "sum", NUM_MN),
    ("(-) Operating expenses (ex-D&A)", "R$ mn", "sum", NUM_MN),
    ("(=) EBITDA (as disclosed)", "R$ mn", "sum", NUM_MN),
    ("(-) Depreciation & amortization", "R$ mn", "sum", NUM_MN),
    ("(=) EBIT", "R$ mn", "sum", NUM_MN),
    ("(+/-) Equity income", "R$ mn", "sum", NUM_MN),
    ("(+/-) Net financial result", "R$ mn", "sum", NUM_MN),
    ("(=) EBT", "R$ mn", "sum", NUM_MN),
    ("(-) Income taxes", "R$ mn", "sum", NUM_MN),
    ("(=) Net income", "R$ mn", "sum", NUM_MN),
]
BS_LINES = [
    ("Cash and equivalents", NUM_MN), ("Short-term investments", NUM_MN),
    ("Trade receivables", NUM_MN), ("Inventories", NUM_MN),
    ("Other current assets", NUM_MN), ("(=) Total current assets", NUM_MN),
    ("PP&E", NUM_MN), ("Intangible assets and goodwill", NUM_MN),
    ("Right-of-use assets", NUM_MN), ("Deferred tax assets", NUM_MN),
    ("Other non-current assets", NUM_MN), ("(=) Total non-current assets", NUM_MN),
    ("(=) Total assets", NUM_MN),
    ("Trade payables", NUM_MN), ("Loans and financing (current)", NUM_MN),
    ("Lease liabilities (current)", NUM_MN), ("Other current liabilities", NUM_MN),
    ("(=) Total current liabilities", NUM_MN),
    ("Loans and financing (non-current)", NUM_MN), ("Lease liabilities (non-current)", NUM_MN),
    ("Provisions and ARO", NUM_MN), ("Deferred tax liabilities", NUM_MN),
    ("Other non-current liabilities", NUM_MN), ("(=) Total liabilities", NUM_MN),
    ("Share capital and reserves", NUM_MN), ("Retained earnings (accumulated)", NUM_MN),
    ("(=) Total equity", NUM_MN),
]
CF_LINES = ["(-) Capex", "(=) Net change in cash",
            "Cash — beginning of period", "Cash — end of period"]
OP_LINES = [
    ("MOBILE", None, None, None),
    ("Total lines", "000", "eop", NUM_MN),
    ("Prepaid lines", "000", "eop", NUM_MN),
    ("Postpaid lines", "000", "eop", NUM_MN),
    ("memo: Net additions", "000", "sum", NUM_MN),
    ("Churn", "%", "avg", PCT),
    ("ARPU total", "R$/mo", "avg", NUM_1D),
    ("ARPU prepaid", "R$/mo", "avg", NUM_1D),
    ("ARPU postpaid", "R$/mo", "avg", NUM_1D),
    ("Market share", "%", "eop", PCT),
    ("FIXED", None, None, None),
    ("TIM Live clients", "000", "eop", NUM_MN),
    ("ARPU TIM Live", "R$/mo", "avg", NUM_1D),
    ("GENERAL", None, None, None),
    ("Capex", "R$ mn", "sum", NUM_MN),
    ("Employees", "#", "eop", NUM_MN),
]


def build_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = "808080"
    ws.column_dimensions["B"].width = 110
    for i, txt in enumerate([
            "TIM — Input workbook (Telecom template)",
            "Type only into blue cells, as reported (R$ mn). Expenses negative.",
            "Annual columns calculate themselves. Gaps stay blank and flow to the gaps report.",
            "Operational data by segment (Mobile/Fixed); fill in whatever the company discloses."], start=2):
        c = ws.cell(row=i, column=2, value=txt)
        c.font = TITLE if i == 2 else SUB

    fin = wb.create_sheet("Input Financials")
    fin.sheet_properties.tabColor = "2E75B6"
    sheet_titles(fin, "Financial history as reported (IFRS). Manual entry in blue cells.")
    fin["B1"] = "TIM — Equity Model Input (Telecom template)"
    fin["B3"] = ("Currency: R$ mn  |  Sign convention: revenues (+), expenses (−)  |  "
                 "Blue = input, black = formula")
    period_header(fin)
    r = HEADER_ROW + 2
    r = section(fin, r, "INCOME STATEMENT (as reported)")
    for label, unit, agg, fmt in IS_LINES:
        r = data_row(fin, r, label, unit, agg, fmt)
    r += 1
    r = section(fin, r, "BALANCE SHEET (as reported)")
    bs_first = r
    for label, fmt in BS_LINES:
        r = data_row(fin, r, label, "R$ mn", "eop", fmt)
    ra, rl, re_ = bs_first + 12, bs_first + 23, bs_first + 26
    from openpyxl.utils import get_column_letter as L
    def chk(row, col):
        c = L(col)
        return f"={c}{ra}-{c}{rl}-{c}{re_}"
    r = data_row(fin, r, "Balance check (assets − liab. − equity) = 0", "R$ mn",
                 "sum", NUM_2D, kind="calc", formula=chk)
    r += 1
    r = section(fin, r, "CASH FLOW (partial: documented gaps)")
    for label in CF_LINES:
        r = data_row(fin, r, label, "R$ mn", "sum", NUM_MN)

    op = wb.create_sheet("Input Operational")
    op.sheet_properties.tabColor = "2E75B6"
    sheet_titles(op, "Operational history by segment. Fill in whatever the company discloses.")
    op["B1"] = "TIM — Equity Model Input (Telecom template)"
    period_header(op)
    r = HEADER_ROW + 2
    for label, unit, agg, fmt in OP_LINES:
        if unit is None:
            r = section(op, r, label)
        else:
            r = data_row(op, r, label, unit, agg, fmt)

    build_assumptions(wb)
    wb.save(path)
    return path


def fill_tim(template, out, source):
    src = load_workbook(source, read_only=True, data_only=True)
    si = src["Inputs"]
    scol = {str(si.cell(row=5, column=c).value): c for c in range(2, 60)
            if si.cell(row=5, column=c).value}
    wb = load_workbook(template)
    fin, op = wb["Input Financials"], wb["Input Operational"]
    rows = {}
    for ws in (fin, op):
        for rr in range(6, ws.max_row + 1):
            v = ws.cell(row=rr, column=2).value
            if isinstance(v, str):
                rows[(ws.title, v.strip())] = rr
    tcol = {str(fin.cell(row=HEADER_ROW, column=c).value): c
            for c in range(FIRST_DATA_COL, fin.max_column + 1)
            if fin.cell(row=HEADER_ROW, column=c).value}

    def sv(row, q):       # value from the source
        v = si.cell(row=row, column=scol[q]).value
        return v if isinstance(v, (int, float)) else None

    def put(ws, label, q, val, flip=False):
        if val is None:
            return
        if flip:
            val = -abs(val)
        ws.cell(row=rows[(ws.title, label)], column=tcol[q], value=round(val, 4)).font = BLUE

    K = 1 / 1000.0
    for q in HIST_QUARTERS:
        # ------------------------- IS (R$ mil -> mn)
        put(fin, "(=) Net revenue", q, (sv(46, q) or 0) * K)
        put(fin, "memo: Mobile service revenue", q, (sv(49, q) or 0) * K)
        put(fin, "memo: Fixed service revenue", q, (sv(57, q) or 0) * K)
        put(fin, "memo: Product revenue", q, (sv(59, q) or 0) * K)
        put(fin, "(-) Operating expenses (ex-D&A)", q, (sv(60, q) or 0) * K)
        put(fin, "(=) EBITDA (as disclosed)", q, (sv(69, q) or 0) * K)
        da = sv(73, q)
        put(fin, "(-) Depreciation & amortization", q, da * K if da else None, flip=True)
        put(fin, "(=) EBIT", q, (sv(77, q) or 0) * K)
        put(fin, "(+/-) Equity income", q, (sv(76, q) or 0) * K)
        put(fin, "(+/-) Net financial result", q, (sv(79, q) or 0) * K)
        put(fin, "(=) EBT", q, (sv(83, q) or 0) * K)
        tax = sv(84, q)
        put(fin, "(-) Income taxes", q, tax * K if tax is not None else None,
            flip=(tax or 0) > 0)
        put(fin, "(=) Net income", q, (sv(85, q) or 0) * K)
        # ------------------------- BS
        KB = 1.0   # BS already in R$ mn at the source
        cash, sti = (sv(111, q) or 0) * KB, (sv(112, q) or 0) * KB
        ar, inv = (sv(113, q) or 0) * KB, (sv(114, q) or 0) * KB
        tca = (sv(110, q) or 0) * KB
        oca = tca - cash - sti - ar - inv
        ta = (sv(109, q) or 0) * KB
        ppe, intan = (sv(135, q) or 0) * KB, (sv(136, q) or 0) * KB
        rou, dta = (sv(131, q) or 0) * KB, (sv(127, q) or 0) * KB
        tnca = ta - tca
        onca = tnca - ppe - intan - rou - dta
        tcl = (sv(138, q) or 0) * KB
        ap, std = (sv(142, q) or 0) * KB, (sv(139, q) or 0) * KB
        stl = (sv(141, q) or 0) * KB
        ocl = tcl - ap - std - stl
        tncl = (sv(150, q) or 0) * KB
        ltd, ltl = (sv(151, q) or 0) * KB, (sv(153, q) or 0) * KB
        prov = ((sv(158, q) or 0) + (sv(160, q) or 0)) * KB
        dtl = (sv(157, q) or 0) * KB
        oncl = tncl - ltd - ltl - prov - dtl
        teq = (sv(163, q) or 0) * KB
        re = ((sv(167, q) or 0) + (sv(168, q) or 0) + (sv(170, q) or 0)) * KB
        cap = teq - re
        for lbl, v in [("Cash and equivalents", cash), ("Short-term investments", sti),
                       ("Trade receivables", ar), ("Inventories", inv),
                       ("Other current assets", oca), ("(=) Total current assets", tca),
                       ("PP&E", ppe), ("Intangible assets and goodwill", intan),
                       ("Right-of-use assets", rou), ("Deferred tax assets", dta),
                       ("Other non-current assets", onca), ("(=) Total non-current assets", tnca),
                       ("(=) Total assets", ta), ("Trade payables", ap),
                       ("Loans and financing (current)", std), ("Lease liabilities (current)", stl),
                       ("Other current liabilities", ocl), ("(=) Total current liabilities", tcl),
                       ("Loans and financing (non-current)", ltd),
                       ("Lease liabilities (non-current)", ltl), ("Provisions and ARO", prov),
                       ("Deferred tax liabilities", dtl), ("Other non-current liabilities", oncl),
                       ("(=) Total liabilities", tcl + tncl),
                       ("Share capital and reserves", cap),
                       ("Retained earnings (accumulated)", re), ("(=) Total equity", teq)]:
            put(fin, lbl, q, v)
        # ------------------------- CF parcial
        capex = sv(26, q)
        put(fin, "(-) Capex", q, capex, flip=True)
        # ------------------------- operacional
        for lbl, row in [("Total lines", 11), ("Prepaid lines", 12), ("Postpaid lines", 13),
                         ("memo: Net additions", 15), ("Churn", 16), ("ARPU total", 17),
                         ("ARPU prepaid", 18), ("ARPU postpaid", 19), ("Market share", 10),
                         ("TIM Live clients", 24), ("ARPU TIM Live", 25), ("Capex", 26),
                         ("Employees", 28)]:
            put(op, lbl, q, sv(row, q))

    # cash: change and BOP/EOP from the BS (4Q22 anchors the first BOP)
    prev = (sv(111, "4Q22") or 0) * 1.0
    for q in HIST_QUARTERS:
        cash = (sv(111, q) or 0) * 1.0
        put(fin, "Cash — beginning of period", q, prev)
        put(fin, "Cash — end of period", q, cash)
        put(fin, "(=) Net change in cash", q, cash - prev)
        prev = cash
    wb.save(out)
    print(f"OK: {out}")


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    tpl = root / "templates" / "input_template_telecom.xlsx"
    build_template(tpl)
    print(f"OK: {tpl}")
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python engine/dev_build_tim.py <path to local source workbook>")
    fill_tim(tpl, root / "inputs" / "TIM_inputs.xlsx", Path(sys.argv[1]))
