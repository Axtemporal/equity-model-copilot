"""Engine v0, TELECOM variant: subscribers × ARPU → revenue; EBITDA margin as driver.

Same architecture as the O&G variant (history linked to the input, Premises with
seeds, roll-forwards, plugs, balance check). Refactoring into a declarative sector
template (YAML) is deferred to v1.

Usage: python engine/build_model_telecom.py <input.xlsx> <model.xlsx>
"""

import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as L
from build_model import (NUM_MN, NUM_1D, PCT, HDR, header_map, label_rows,
                         detect_history, build_timeline, block_cols, prev_period,
                         sheet_shell, section, add_line, seed_premises, prem, q_seq,
                         PROJ_YEARS_ANNUAL, TITLE, SUB, Alignment)

NUM_0D = NUM_MN


def main(input_path, out_path):
    shutil.copy(input_path, out_path)
    wb = load_workbook(out_path)
    fin, op = wb["Input Financials"], wb["Input Operational"]
    fhdr, ohdr = header_map(fin), header_map(op)
    frow, orow = label_rows(fin), label_rows(op)
    hist_q = detect_history(fin)
    tl = build_timeline(hist_q)
    last = hist_q[-1]
    lc_f, lc_o = fhdr[last], ohdr[last]
    g = lambda ws, rr, cc: ws.cell(row=rr, column=cc).value or 0

    def delta_avg(row, n=4):
        vals = [g(op, row, ohdr[q]) for q in hist_q[-(n + 1):]]
        ds = [b - a for a, b in zip(vals, vals[1:]) if a and b]
        return sum(ds) / len(ds) if ds else 0

    rev_l = g(fin, frow["(=) Net revenue"], lc_f)
    seeds = [
        ("adds", "Mobile net additions (per period)", "000", NUM_0D, delta_avg(orow["Total lines"]), 4),
        ("arpu", "ARPU total", "R$/mo", NUM_1D, g(op, orow["ARPU total"], lc_o), 1),
        ("ladds", "TIM Live net additions (per period)", "000", NUM_0D, delta_avg(orow["TIM Live clients"]), 4),
        ("larpu", "ARPU TIM Live", "R$/mo", NUM_1D, g(op, orow["ARPU TIM Live"], lc_o), 1),
        ("fixoth", "Fixed revenue beyond TIM Live (per period)", "R$ mn", NUM_0D,
         g(fin, frow["memo: Fixed service revenue"], lc_f)
         - g(op, orow["TIM Live clients"], lc_o) * g(op, orow["ARPU TIM Live"], lc_o) * 3 / 1000, 4),
        ("prodrev", "Product revenue (per period)", "R$ mn", NUM_0D,
         g(fin, frow["memo: Product revenue"], lc_f), 4),
        ("margin", "EBITDA margin (% of revenue)", "%", PCT,
         g(fin, frow["(=) EBITDA (as disclosed)"], lc_f) / rev_l if rev_l else 0.5, 1),
        ("da", "D&A (per period)", "R$ mn", NUM_0D,
         -g(fin, frow["(-) Depreciation & amortization"], lc_f), 4),
        ("finres", "Net financial result (per period)", "R$ mn", NUM_0D,
         g(fin, frow["(+/-) Net financial result"], lc_f), 4),
        ("tax", "Effective tax rate (% of EBT)", "%", PCT,
         -g(fin, frow["(-) Income taxes"], lc_f) / g(fin, frow["(=) EBT"], lc_f)
         if g(fin, frow["(=) EBT"], lc_f) else 0.30, 1),
        ("capex", "Capex (per period)", "R$ mn", NUM_0D, g(op, orow["Capex"], lc_o), 4),
        ("payout", "Dividend payout (% of NI)", "%", PCT, 0.0, 1),
    ]
    pws, prows, pcols = seed_premises(wb, tl, fin, op, [], seeds)
    P, Praw = prem("Premises", prows, pcols)

    # ------------------------------------------------ Operational Model
    om = sheet_shell(wb, "Operational Model", "Assinantes × ARPU → receita", tl)
    r = HDR + 2
    r = section(om, r, "MOBILE", tl)
    hist_op = lambda lbl: (lambda p: f"='Input Operational'!{L(ohdr[p['label']])}{orow[lbl]}")
    hist_fin = lambda lbl: (lambda p: f"='Input Financials'!{L(fhdr[p['label']])}{frow[lbl]}")
    row_lines = r

    def stock_chain(row, seed_lbl, adds_key, src):
        def f(p, i):
            pv = prev_period(tl, i)
            base = (f"='Input Operational'!{L(ohdr[last])}{orow[seed_lbl]}"
                    if pv["kind"] == "QH" else f"={L(pv['col'])}{row}")
            return base[0] + f"({base[1:]})+{Praw(adds_key, p)}" if False else base + f"+{Praw(adds_key, p)}"
        return f

    lines_f = stock_chain(row_lines, "Total lines", "adds", op)
    r = add_line(om, r, "Total lines (EOP)", "000", NUM_0D, tl,
                 hist=hist_op("Total lines"), qp=lines_f, yb="eop", ys=lines_f)
    row_arpu = r
    r = add_line(om, r, "ARPU total", "R$/mo", NUM_1D, tl, hist=hist_op("ARPU total"),
                 qp=lambda p, i: P("arpu", p), yb="avg", ys=lambda p, i: P("arpu", p))
    row_mob = r
    mob_f = lambda p, i: (f"={L(p['col'])}{row_lines}*{L(p['col'])}{row_arpu}*3/1000"
                          if p["q"] else
                          f"={L(p['col'])}{row_lines}*{L(p['col'])}{row_arpu}*12/1000")
    r = add_line(om, r, "(=) Mobile service revenue", "R$ mn", NUM_MN, tl,
                 hist=hist_fin("memo: Mobile service revenue"), qp=mob_f, yb="sum", ys=mob_f)
    r += 1
    r = section(om, r, "FIXED AND PRODUCTS", tl)
    row_live = r
    live_f = stock_chain(row_live, "TIM Live clients", "ladds", op)
    r = add_line(om, r, "TIM Live clients (EOP)", "000", NUM_0D, tl,
                 hist=hist_op("TIM Live clients"), qp=live_f, yb="eop", ys=live_f)
    row_larpu = r
    r = add_line(om, r, "ARPU TIM Live", "R$/mo", NUM_1D, tl, hist=hist_op("ARPU TIM Live"),
                 qp=lambda p, i: P("larpu", p), yb="avg", ys=lambda p, i: P("larpu", p))
    row_fix = r
    fix_f = lambda p, i: (f"={L(p['col'])}{row_live}*{L(p['col'])}{row_larpu}*3/1000+{Praw('fixoth', p)}"
                          if p["q"] else
                          f"={L(p['col'])}{row_live}*{L(p['col'])}{row_larpu}*12/1000+{Praw('fixoth', p)}")
    r = add_line(om, r, "(=) Fixed service revenue", "R$ mn", NUM_MN, tl,
                 hist=hist_fin("memo: Fixed service revenue"), qp=fix_f, yb="sum", ys=fix_f)
    row_prod = r
    r = add_line(om, r, "(=) Product revenue", "R$ mn", NUM_MN, tl,
                 hist=hist_fin("memo: Product revenue"),
                 qp=lambda p, i: P("prodrev", p), yb="sum", ys=lambda p, i: P("prodrev", p))
    row_rev = r
    rev_f = lambda p, i: f"={L(p['col'])}{row_mob}+{L(p['col'])}{row_fix}+{L(p['col'])}{row_prod}"
    r = add_line(om, r, "(=) Net revenue build", "R$ mn", NUM_MN, tl,
                 hist=hist_fin("(=) Net revenue"), qp=rev_f, yb="sum", ys=rev_f)
    OM = "Operational Model"

    # ------------------------------------------------ IS
    is_ = sheet_shell(wb, "IS", "Income statement integrado", tl)
    ri = HDR + 2
    ri = section(is_, ri, "INCOME STATEMENT", tl)
    R = {}
    def line(lbl, unit, fmt, **kw):
        nonlocal ri
        R[lbl] = ri
        ri = add_line(is_, ri, lbl, unit, fmt, tl, **kw)
    line("(=) Net revenue", "R$ mn", NUM_MN, hist=hist_fin("(=) Net revenue"),
         qp=lambda p, i: f"='{OM}'!{L(p['col'])}{row_rev}", yb="sum",
         ys=lambda p, i: f"='{OM}'!{L(p['col'])}{row_rev}")
    opex_f = lambda p, i: f"=-{L(p['col'])}{R['(=) Net revenue']}*(1-{Praw('margin', p)})"
    line("(-) Operating expenses (ex-D&A)", "R$ mn", NUM_MN,
         hist=hist_fin("(-) Operating expenses (ex-D&A)"), qp=opex_f, yb="sum", ys=opex_f)
    ebitda_f = lambda p, i: f"={L(p['col'])}{R['(=) Net revenue']}+{L(p['col'])}{R['(-) Operating expenses (ex-D&A)']}"
    line("(=) EBITDA", "R$ mn", NUM_MN, hist=hist_fin("(=) EBITDA (as disclosed)"),
         qp=ebitda_f, yb="sum", ys=ebitda_f)
    da_f = lambda p, i: f"=-{Praw('da', p)}"
    line("(-) Depreciation & amortization", "R$ mn", NUM_MN,
         hist=hist_fin("(-) Depreciation & amortization"), qp=da_f, yb="sum", ys=da_f)
    ebit_f = lambda p, i: f"={L(p['col'])}{R['(=) EBITDA']}+{L(p['col'])}{R['(-) Depreciation & amortization']}"
    line("(=) EBIT", "R$ mn", NUM_MN, hist=hist_fin("(=) EBIT"), qp=ebit_f, yb="sum", ys=ebit_f)
    line("(+/-) Net financial result", "R$ mn", NUM_MN,
         hist=hist_fin("(+/-) Net financial result"),
         qp=lambda p, i: P("finres", p), yb="sum", ys=lambda p, i: P("finres", p))
    ebt_f = lambda p, i: f"={L(p['col'])}{R['(=) EBIT']}+{L(p['col'])}{R['(+/-) Net financial result']}"
    line("(=) EBT", "R$ mn", NUM_MN, hist=hist_fin("(=) EBT"), qp=ebt_f, yb="sum", ys=ebt_f)
    tax_f = lambda p, i: f"=-MAX({L(p['col'])}{R['(=) EBT']},0)*{Praw('tax', p)}"
    line("(-) Income taxes", "R$ mn", NUM_MN, hist=hist_fin("(-) Income taxes"),
         qp=tax_f, yb="sum", ys=tax_f)
    ni_f = lambda p, i: f"={L(p['col'])}{R['(=) EBT']}+{L(p['col'])}{R['(-) Income taxes']}"
    line("(=) Net income", "R$ mn", NUM_MN, hist=hist_fin("(=) Net income"),
         qp=ni_f, yb="sum", ys=ni_f)
    mg_f = lambda p, i: f'=IFERROR({L(p["col"])}{R["(=) EBITDA"]}/{L(p["col"])}{R["(=) Net revenue"]},"-")'
    ri = add_line(is_, ri, "EBITDA margin", "%", PCT, tl, hist=lambda p: mg_f(p, 0),
                  qp=mg_f, yb=mg_f, ys=mg_f)
    NI = R["(=) Net income"]

    # ------------------------------------------------ Schedules
    sc = sheet_shell(wb, "Schedules", "Roll-forwards: PP&E+intangibles, RE, cash", tl)
    r = HDR + 2
    SCH = {}
    def corkscrew(r, name, seed_formula, plus, minus):
        r = section(sc, r, name, tl)
        rb, rp, rm, re_ = r, r + 1, r + 2, r + 3
        def bop(p, i):
            pv = prev_period(tl, i)
            if pv is None or pv["kind"] == "QH":
                return seed_formula
            if p["kind"] == "YB":
                return f"={L(block_cols(tl, p)[0])}{rb}"
            return f"={L(pv['col'])}{re_}"
        add_line(sc, rb, "BOP", "R$ mn", NUM_MN, tl, hist=lambda p: None, qp=bop, yb=bop, ys=bop)
        add_line(sc, rp, "(+) additions", "R$ mn", NUM_MN, tl, qp=plus, yb="sum", ys=plus)
        add_line(sc, rm, "(-) reductions", "R$ mn", NUM_MN, tl, qp=minus, yb="sum", ys=minus)
        eop = lambda p, i: f"=SUM({L(p['col'])}{rb}:{L(p['col'])}{rm})"
        eop_yb = lambda p, i: f"={L(block_cols(tl, p)[-1])}{re_}"
        add_line(sc, re_, "(=) EOP", "R$ mn", NUM_MN, tl, qp=eop, yb=eop_yb, ys=eop)
        return re_ + 2, dict(bop=rb, eop=re_)

    seed_fa = (f"='Input Financials'!{L(fhdr[last])}{frow['PP&E']}"
               f"+'Input Financials'!{L(fhdr[last])}{frow['Intangible assets and goodwill']}")
    r, SCH["fa"] = corkscrew(r, "FIXED + INTANGIBLE ASSETS", seed_fa,
                             plus=lambda p, i: P("capex", p)[1:] and f"={Praw('capex', p)}",
                             minus=lambda p, i: f"=-{Praw('da', p)}")
    r, SCH["re"] = corkscrew(r, "RETAINED EARNINGS",
                             f"='Input Financials'!{L(fhdr[last])}{frow['Retained earnings (accumulated)']}",
                             plus=lambda p, i: f"=IS!{L(p['col'])}{NI}",
                             minus=lambda p, i: f"=-IS!{L(p['col'])}{NI}*{Praw('payout', p)}")

    # ------------------------------------------------ CF
    cf = sheet_shell(wb, "CF", "Cash flow: reconciliation", tl)
    rc = HDR + 2
    rc = section(cf, rc, "CASH FLOW", tl)
    C = {}
    def cline(lbl, src, qp, ys=None):
        nonlocal rc
        C[lbl] = rc
        rc = add_line(cf, rc, lbl, "R$ mn", NUM_MN, tl,
                      hist=hist_fin(src) if src else None, qp=qp, yb="sum", ys=ys or qp)
    cline("(=) Net income", "(=) Net income", lambda p, i: f"=IS!{L(p['col'])}{NI}")
    cline("(+) D&A", "(-) Depreciation & amortization",
          lambda p, i: f"=-IS!{L(p['col'])}{R['(-) Depreciation & amortization']}")
    cline("(+/-) Working capital and other", None, lambda p, i: "=0")
    cfo = lambda p, i: f"=SUM({L(p['col'])}{C['(=) Net income']}:{L(p['col'])}{C['(+/-) Working capital and other']})"
    cline("(=) Cash from operations", None, cfo)
    cline("(-) Capex", "(-) Capex", lambda p, i: f"=-{Praw('capex', p)}")
    cline("(-) Dividends", None, lambda p, i: f"=-IS!{L(p['col'])}{NI}*{Praw('payout', p)}")
    dch = lambda p, i: (f"={L(p['col'])}{C['(=) Cash from operations']}+{L(p['col'])}{C['(-) Capex']}"
                        f"+{L(p['col'])}{C['(-) Dividends']}")
    cline("(=) Net change in cash", "(=) Net change in cash", dch)
    r, SCH["cash"] = corkscrew(r, "CASH", f"='Input Financials'!{L(fhdr[last])}{frow['Cash and equivalents']}",
                               plus=lambda p, i: f"=CF!{L(p['col'])}{C['(=) Net change in cash']}",
                               minus=lambda p, i: "=0")
    for lbl, src, key in [("Cash — beginning of period", "Cash — beginning of period", "bop"),
                          ("Cash — end of period", "Cash — end of period", "eop")]:
        rr = rc
        C[lbl] = rr
        def yb_cash(p, i, key=key, rr=rr):
            bc = block_cols(tl, p)
            return f"={L(bc[0] if key == 'bop' else bc[-1])}{rr}"
        rc = add_line(cf, rc, lbl, "R$ mn", NUM_MN, tl, hist=hist_fin(src),
                      qp=lambda p, i, key=key: f"=Schedules!{L(p['col'])}{SCH['cash'][key]}",
                      yb=yb_cash,
                      ys=lambda p, i, key=key: f"=Schedules!{L(p['col'])}{SCH['cash'][key]}")

    # ------------------------------------------------ BS
    bs = sheet_shell(wb, "BS", "Balance sheet integrado", tl)
    rb_ = HDR + 2
    rb_ = section(bs, rb_, "BALANCE SHEET", tl)
    B = {}
    def bline(lbl, qp=None, ys=None):
        nonlocal rb_
        B[lbl] = rb_
        flat = lambda p, i: f"={L(prev_period(tl, i)['col'])}{B[lbl]}"
        rb_ = add_line(bs, rb_, lbl, "R$ mn", NUM_MN, tl, hist=hist_fin(lbl),
                       qp=qp or flat, yb="eop", ys=ys or qp or flat)
    bline("Cash and equivalents", qp=lambda p, i: f"=Schedules!{L(p['col'])}{SCH['cash']['eop']}",
          ys=lambda p, i: f"=Schedules!{L(p['col'])}{SCH['cash']['eop']}")
    for lbl in ["Short-term investments", "Trade receivables", "Inventories", "Other current assets"]:
        bline(lbl)
    tca = lambda p, i: f"=SUM({L(p['col'])}{B['Cash and equivalents']}:{L(p['col'])}{B['Other current assets']})"
    bline("(=) Total current assets", qp=tca, ys=tca)
    fa_f = lambda p, i: f"=Schedules!{L(p['col'])}{SCH['fa']['eop']}"
    bline("PP&E", qp=fa_f, ys=fa_f)
    def zero(p, i):
        return "=0"
    bline("Intangible assets and goodwill", qp=zero, ys=zero)   # consolidated into PP&E in the projection
    for lbl in ["Right-of-use assets", "Deferred tax assets", "Other non-current assets"]:
        bline(lbl)
    tnca = lambda p, i: f"=SUM({L(p['col'])}{B['PP&E']}:{L(p['col'])}{B['Other non-current assets']})"
    bline("(=) Total non-current assets", qp=tnca, ys=tnca)
    ta = lambda p, i: f"={L(p['col'])}{B['(=) Total current assets']}+{L(p['col'])}{B['(=) Total non-current assets']}"
    bline("(=) Total assets", qp=ta, ys=ta)
    for lbl in ["Trade payables", "Loans and financing (current)", "Lease liabilities (current)",
                "Other current liabilities"]:
        bline(lbl)
    tcl = lambda p, i: f"=SUM({L(p['col'])}{B['Trade payables']}:{L(p['col'])}{B['Other current liabilities']})"
    bline("(=) Total current liabilities", qp=tcl, ys=tcl)
    for lbl in ["Loans and financing (non-current)", "Lease liabilities (non-current)",
                "Provisions and ARO", "Deferred tax liabilities", "Other non-current liabilities"]:
        bline(lbl)
    tlb = lambda p, i: (f"={L(p['col'])}{B['(=) Total current liabilities']}"
                        f"+SUM({L(p['col'])}{B['Loans and financing (non-current)']}:"
                        f"{L(p['col'])}{B['Other non-current liabilities']})")
    bline("(=) Total liabilities", qp=tlb, ys=tlb)
    bline("Share capital and reserves")
    re_f = lambda p, i: f"=Schedules!{L(p['col'])}{SCH['re']['eop']}"
    bline("Retained earnings (accumulated)", qp=re_f, ys=re_f)
    teq = lambda p, i: (f"={L(p['col'])}{B['Share capital and reserves']}"
                        f"+{L(p['col'])}{B['Retained earnings (accumulated)']}")
    bline("(=) Total equity", qp=teq, ys=teq)
    chk = lambda p, i: (f"={L(p['col'])}{B['(=) Total assets']}-{L(p['col'])}{B['(=) Total liabilities']}"
                        f"-{L(p['col'])}{B['(=) Total equity']}")
    add_line(bs, rb_, "Balance check (= 0)", "R$ mn", NUM_1D, tl,
             hist=lambda p: chk(p, 0), qp=chk, yb=chk, ys=chk)

    # ------------------------------------------------ Cover e ordem
    cov = wb.create_sheet("Cover", 0)
    cov.sheet_properties.tabColor = "1F3864"
    cov.column_dimensions["B"].width = 100
    for k, txt in enumerate([
            "TIM — modelo integrado (motor v0 telecom)",
            f"History {hist_q[0]} to {hist_q[-1]} (extracted from reported data). Projection: quarterly through "
            f"4Q{str(q_seq(last)[0] + 1)[2:]}, annual through {q_seq(last)[0] + 1 + PROJ_YEARS_ANNUAL}.",
            "Assumptions in the Premises tab (seeds = last historical value; replace in the line-by-line session).",
            "v0 limitations: no revolver, constant WC and debt, intangibles consolidated into PP&E in the projection.",
            "Known input gaps: 4Q25 and 1Q26 already reported by TIM are missing from the local source; partial CF."],
            start=2):
        c = cov.cell(row=k, column=2, value=txt)
        c.font = TITLE if k == 2 else SUB
        c.alignment = Alignment(wrap_text=True, vertical="top")
    order = ["Cover", "IS", "BS", "CF", "Operational Model", "Schedules", "Premises",
             "Assumptions", "Input Financials", "Input Operational", "README"]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + \
                 [ws for ws in wb._sheets if ws.title not in order]
    wb.save(out_path)
    print(f"OK: {out_path} | hist {hist_q[0]}..{hist_q[-1]} | {len(tl)} period columns")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
