"""Pre-build input validation (sector-aware).

Runs on the analyst's INPUT workbook before the engine builds, so a malformed input fails
fast with every problem listed at once, instead of a bare KeyError deep inside the build
(the engine looks up line labels by exact string). Cheap and pure-openpyxl — no recalc.
"""
from __future__ import annotations

import difflib
import re

import openpyxl

from ..canonical_schema import known_sectors, required_labels
from .invariants import ERROR_VALUES, LABEL_COL, find_label_row, period_columns
from .report import CellRef, InvariantResult, Report, Severity

_QUARTER_RE = re.compile(r"^[1-4]Q\d{2}$")
_YEAR_RE = re.compile(r"^\d{4}$")

# Exact labels each sector's engine looks up by string (a typo here is a build-time KeyError).
# Sourced from the canonical schema (universal base + per-sector delta), over the sectors that
# exist on disk — so the validated set matches what the engine hard-reads, with no hardcoded list.
REQUIRED_LABELS = {sector: required_labels(sector) for sector in known_sectors()}


class InputValidationError(Exception):
    """Raised by assert_valid_input when the input workbook fails a hard check."""

    def __init__(self, problems):
        self.problems = list(problems)
        super().__init__("; ".join(self.problems))


def detect_sector(wb) -> str | None:
    """Detect the sector from the operational tab's declared signals (data-driven, agnostic).

    Delegates to template_loader.identify_sector, which matches each sector's signals from the
    knowledge base — so no sector is hardcoded here. None when ambiguous → the caller asks.
    """
    if "Input Operational" not in wb.sheetnames:
        return None
    op = wb["Input Operational"]
    labels = {
        op.cell(r, LABEL_COL).value.strip(): r
        for r in range(1, op.max_row + 1)
        if isinstance(op.cell(r, LABEL_COL).value, str) and op.cell(r, LABEL_COL).value.strip()
    }
    from ..template_loader import identify_sector

    return identify_sector(labels)


def _labels(ws) -> set[str]:
    out = set()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, LABEL_COL).value
        if isinstance(value, str) and value.strip():
            out.add(value.strip())
    return out


def validate_input(source, sector: str | None = None) -> Report:
    """Validate an input workbook (path or openpyxl Workbook). Returns a Report; never raises."""
    path = source if isinstance(source, str) else "<workbook>"
    wb = openpyxl.load_workbook(source) if isinstance(source, str) else source
    results: list[InvariantResult] = []

    missing_sheets = [s for s in ("Input Financials", "Input Operational") if s not in wb.sheetnames]
    if missing_sheets:
        results.append(InvariantResult(
            "required_sheets", Severity.FAIL, False,
            f"missing sheet(s): {', '.join(missing_sheets)} (found: {', '.join(wb.sheetnames)})",
        ))
        return Report(path, "validator", results)

    sector = sector or detect_sector(wb)
    if sector not in REQUIRED_LABELS:
        results.append(InvariantResult(
            "sector", Severity.FAIL, False,
            "could not determine the sector from the operational signals — pass --sector "
            "explicitly, or ensure the input carries the sector's declared signal labels",
        ))
        return Report(path, "validator", results)
    results.append(InvariantResult("sector", Severity.INFO, True, f"sector detected: {sector}"))

    for sheet, needed in REQUIRED_LABELS[sector].items():
        present = _labels(wb[sheet])
        missing = [lbl for lbl in needed if lbl not in present]
        if missing:
            details = []
            for lbl in missing:
                near = difflib.get_close_matches(lbl, present, n=1)
                details.append(f"'{lbl}'" + (f" (closest: '{near[0]}')" if near else ""))
            results.append(InvariantResult(
                f"labels[{sheet}]", Severity.FAIL, False,
                f"{len(missing)} required label(s) missing in '{sheet}': " + "; ".join(details),
            ))
        else:
            results.append(InvariantResult(
                f"labels[{sheet}]", Severity.FAIL, True,
                f"all {len(needed)} required labels present in '{sheet}'",
            ))

    fin = wb["Input Financials"]
    cols = period_columns(fin)
    bad = [lbl for _c, lbl in cols if not (_QUARTER_RE.match(lbl) or _YEAR_RE.match(lbl))]
    if not cols:
        results.append(InvariantResult("period_headers", Severity.FAIL, False,
                                       "no period headers in row 5 of Input Financials"))
    elif bad:
        results.append(InvariantResult("period_headers", Severity.FAIL, False,
                                       f"unparseable period header(s): {bad[:8]}"))
    else:
        results.append(InvariantResult("period_headers", Severity.FAIL, True,
                                       f"{len(cols)} period headers parse"))

    rev = find_label_row(fin, "(=) Net revenue")
    n_hist = 0
    if rev:
        for col, lbl in cols:
            if _QUARTER_RE.match(lbl) and fin.cell(rev, col).value is not None:
                n_hist += 1
    if n_hist == 0:
        results.append(InvariantResult("history_present", Severity.FAIL, False,
                                       "no filled historical quarter on the '(=) Net revenue' line"))
    else:
        results.append(InvariantResult("history_present", Severity.FAIL, True,
                                       f"{n_hist} historical quarter(s) detected on Net revenue"))

    errors = []
    for sheet in ("Input Financials", "Input Operational"):
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in ERROR_VALUES:
                    errors.append(CellRef(sheet, cell.coordinate, actual=cell.value))
    results.append(InvariantResult(
        "no_error_cells_input", Severity.FAIL, not errors,
        "no error cells in input" if not errors else f"{len(errors)} error cell(s) in input",
        errors[:20],
    ))

    return Report(path, "validator", results)


def assert_valid_input(source, sector: str | None = None) -> Report:
    """Validate and raise InputValidationError listing every problem if it fails."""
    report = validate_input(source, sector)
    if not report.ok:
        raise InputValidationError([r.message for r in report.failures()])
    return report
