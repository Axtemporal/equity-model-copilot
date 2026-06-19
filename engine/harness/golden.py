"""Golden regression test.

Builds the synthetic model from a frozen input, recalculates it, and compares a curated set
of computed cells against a stored snapshot. Catches engine regressions that still produce a
(differently) balanced model — the balance check alone would not notice those.

Regenerate the snapshot deliberately after an intended engine change:
    python -m engine.harness.golden --update <input.xlsx> <golden.json>
"""
from __future__ import annotations

import json
import os
import tempfile

from openpyxl.utils import get_column_letter

from .invariants import find_label_row, period_columns
from .recalc import recalc_workbook
from .report import CellRef, InvariantResult, Report, Severity

# (sheet, label-substring) of the lines to freeze; captured across every period column.
GOLDEN_LINES = [
    ("BS", "balance check"),
    ("BS", "(=) total assets"),
    ("BS", "(=) total equity"),
    ("BS", "retained earnings (accumulated)"),
    ("IS", "(=) net revenue"),
    ("IS", "(=) net income"),
    ("CF", "cash — end of period"),
]


def capture(calc) -> dict[str, float]:
    """Extract {Sheet!A1: value} for the curated lines across all period columns."""
    snapshot: dict[str, float] = {}
    for sheet, needle in GOLDEN_LINES:
        if sheet not in calc.sheetnames:
            continue
        ws = calc[sheet]
        row = find_label_row(ws, needle)
        if row is None:
            continue
        for col, _label in period_columns(ws):
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)):
                snapshot[f"{sheet}!{get_column_letter(col)}{row}"] = round(float(value), 6)
    return snapshot


def build_and_capture(input_path: str, *, backend: str = "formulas") -> dict[str, float]:
    from engine import build_model  # the synthetic fixture is Oil & Gas

    out = os.path.join(tempfile.mkdtemp(prefix="emc_golden_"), "model.xlsx")
    build_model.main(input_path, out)
    calc, _out = recalc_workbook(out, backend=backend)
    return capture(calc)


def make_golden(input_path: str, golden_path: str, *, backend: str = "formulas") -> dict[str, float]:
    snapshot = build_and_capture(input_path, backend=backend)
    with open(golden_path, "w", encoding="utf-8") as handle:
        json.dump(snapshot, handle, indent=2, sort_keys=True)
    return snapshot


def check_golden(input_path: str, golden_path: str, *, backend: str = "formulas") -> Report:
    with open(golden_path, encoding="utf-8") as handle:
        golden = json.load(handle)
    current = build_and_capture(input_path, backend=backend)
    failing = []
    for key, golden_value in golden.items():
        sheet, a1 = key.split("!", 1)
        actual = current.get(key)
        if actual is None:
            failing.append(CellRef(sheet, a1, expected=golden_value, actual=None))
        elif abs(actual - golden_value) > max(1e-6, 1e-6 * abs(golden_value)):
            failing.append(CellRef(sheet, a1, expected=golden_value, actual=actual,
                                   abs_error=abs(actual - golden_value)))
    passed = not failing
    msg = (
        f"all {len(golden)} golden cells match"
        if passed
        else f"{len(failing)}/{len(golden)} golden cells differ from the snapshot"
    )
    return Report(input_path, backend, [InvariantResult("golden_match", Severity.FAIL, passed, msg, failing[:30])])


if __name__ == "__main__":
    import argparse
    import sys

    parser = argparse.ArgumentParser(prog="engine.harness.golden")
    parser.add_argument("input", help="input .xlsx to build from")
    parser.add_argument("golden", help="path to the golden .json snapshot")
    parser.add_argument("--update", action="store_true", help="(re)write the snapshot")
    parser.add_argument("--backend", default="formulas")
    args = parser.parse_args()

    if args.update:
        snap = make_golden(args.input, args.golden, backend=args.backend)
        print(f"wrote {len(snap)} cells to {args.golden}")
    else:
        report = check_golden(args.input, args.golden, backend=args.backend)
        print(report.summary())
        for cell in report.results[0].failing_cells:
            print("  -", cell)
        sys.exit(0 if report.ok else 1)
