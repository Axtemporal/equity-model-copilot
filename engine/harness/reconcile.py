"""Post-build reconciliation gate — reported aggregates vs the model-computed values.

Runs AFTER the build and BEFORE the projection-premise session (the "model is sound?" gate in
the Fase 1 design). Compares the INGESTED reported aggregates (disclosed EBITDA; reported net
debt if provided) against what the recalculated model computes over the HISTORICAL columns — a
data-quality WARNING surfaced to the analyst, never an engine-correctness FAIL. The arithmetic
is deterministic here; Claude only explains a flagged gap.

Kept OUT of the always-on INVARIANTS list (which guarantees a clean synth has no warnings): a
reported-vs-computed gap is a property of the *input data*, not the engine, so it is its own gate.
"""
from __future__ import annotations

import openpyxl

from .invariants import classify_period_columns, find_label_row, period_columns
from .recalc import recalc_workbook
from .report import CellRef, InvariantResult, Report, Severity


def compare_series(reported: dict, computed: dict, *, rel_tol: float = 0.01,
                   abs_tol: float = 1e-6) -> list[tuple[str, float, float]]:
    """Periods (present in both, numeric) where |reported - computed| exceeds tolerance.

    Returns [(period, reported, computed)] — the deterministic core, independent of any workbook.
    """
    gaps = []
    for period, rep in reported.items():
        comp = computed.get(period)
        if not isinstance(rep, (int, float)) or not isinstance(comp, (int, float)):
            continue
        if abs(rep - comp) > max(abs_tol, rel_tol * abs(rep)):
            gaps.append((period, float(rep), float(comp)))
    return gaps


def _exact_row(ws, label: str):
    """Row whose column-B label equals `label` (case-insensitive, trimmed), else None."""
    target = label.strip().lower()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 2).value
        if isinstance(value, str) and value.strip().lower() == target:
            return row
    return None


def _hist_series(ws, row, hist_labels: set[str]) -> dict:
    return {lbl: ws.cell(row, col).value
            for col, lbl in period_columns(ws) if lbl in hist_labels}


# (label, input-tab line, model-tab, model line) reconciliations to attempt.
_RECONCILIATIONS = [
    ("EBITDA", "Input Financials", "memo: (=) EBITDA (as disclosed)", "IS", "memo: (=) EBITDA"),
    ("Net debt", "Input Financials", "memo: Net debt (as disclosed)", "Schedules", "(=) Net debt"),
]


def reconcile_reported(model_path: str, *, backend: str = "formulas",
                       rel_tol: float = 0.01) -> Report:
    """Reconcile each reported aggregate against its computed counterpart over history (WARN)."""
    orig = openpyxl.load_workbook(model_path)
    calc, _out = recalc_workbook(model_path, backend=backend)
    results: list[InvariantResult] = []

    hist_labels = set()
    if "BS" in orig.sheetnames:
        kinds = classify_period_columns(orig)
        hist_labels = {lbl for col, lbl in period_columns(orig["BS"]) if kinds.get(col) == "history"}

    for name, in_sheet, in_line, mdl_sheet, mdl_line in _RECONCILIATIONS:
        if in_sheet not in calc.sheetnames or mdl_sheet not in calc.sheetnames:
            continue
        in_row = _exact_row(calc[in_sheet], in_line)
        mdl_row = _exact_row(calc[mdl_sheet], mdl_line)
        if in_row is None or mdl_row is None:
            continue
        reported = _hist_series(calc[in_sheet], in_row, hist_labels)
        computed = _hist_series(calc[mdl_sheet], mdl_row, hist_labels)
        gaps = compare_series(reported, computed, rel_tol=rel_tol)
        if not any(isinstance(v, (int, float)) for v in reported.values()):
            continue  # nothing reported to reconcile against
        cells = [CellRef(mdl_sheet, name, period=p, expected=rep, actual=comp,
                         abs_error=abs(rep - comp)) for p, rep, comp in gaps]
        results.append(InvariantResult(
            f"reconcile_{name.lower().replace(' ', '_')}", Severity.WARN, not gaps,
            (f"reported {name} ties to computed in every historical period" if not gaps
             else f"{name}: reported vs computed differ in {len(gaps)} historical period(s) "
                  f"(> {rel_tol:.0%}) — review input or normalization"),
            cells,
        ))

    if not results:
        results.append(InvariantResult("reconcile", Severity.INFO, True,
                                       "no reported aggregates available to reconcile"))
    return Report(model_path, backend, results)
