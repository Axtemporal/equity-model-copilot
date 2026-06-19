"""Model invariants checked on the recalculated workbook.

Each check has signature `check(orig, calc) -> InvariantResult`, where `orig` is the
formula-only workbook (openpyxl, data_only=False) used to locate formula cells and read
geometry, and `calc` is the recalculated workbook (data_only=True) holding the computed
values.

Severity policy: integrity guarantees that must hold for the engine to be correct are
FAIL (structure, no-unresolved, balance check on projections, aggregation, history
fidelity, error cells); soft guardrails that can have legitimate edge cases are WARN
(sign conventions, history-only balance breaks).

The dynamic schedules of foundation item 4 (working capital, leases, debt, revolver)
extend this file by appending guarded check functions to INVARIANTS — no other module
changes. Planned: wc_change_ties, lease_rollforward_ties, debt_interest_on_opening,
revolver_min_cash, schedule_sanity (PP&E EOP >= 0), kd_linked_to_schedule.
"""
from __future__ import annotations

import re

from openpyxl.utils import get_column_letter

from .report import CellRef, InvariantResult, Severity

HDR = 5            # period header row
FIRST_COL = 4      # first data column (D)
LABEL_COL = 2      # column B holds the line labels
MODEL_TABS = ["IS", "BS", "CF", "Operational Model", "Schedules"]
CORE_SHEETS = ["IS", "BS", "CF", "Operational Model", "Schedules", "Premises",
               "Input Financials", "Input Operational"]
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}
TOL = 1e-6         # absolute tolerance, in reporting-currency units

_QUARTER_RE = re.compile(r"^([1-4])Q(\d{2})$")
_YEAR_RE = re.compile(r"^(\d{4})$")
_INPUT_LINK_RE = re.compile(
    r"^=\s*'?(Input Financials|Input Operational)'?!\$?([A-Z]{1,3})\$?(\d+)\s*$", re.I
)
_SINGLE_REF_RE = re.compile(r"^=\$?[A-Z]{1,3}\$?\d+$")
_SINGLE_REF_PARTS = re.compile(r"^=\$?([A-Z]{1,3})\$?(\d+)$")
_REVENUE_HINTS = (
    "net revenue", "mobile service revenue", "fixed service revenue", "product revenue",
    "revenue build", "sales volume", "realized price", "brent average", "total production",
)


# --------------------------------------------------------------------------- geometry

def period_columns(ws):
    """[(col_index, label)] for every populated period header from column D onward."""
    cols = []
    for col in range(FIRST_COL, ws.max_column + 1):
        label = ws.cell(HDR, col).value
        if label not in (None, ""):
            cols.append((col, str(label)))
    return cols


def find_label_row(ws, needle: str):
    """First row whose column-B label contains `needle` (case-insensitive), else None."""
    needle = needle.lower()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, LABEL_COL).value
        if isinstance(value, str) and needle in value.lower():
            return row
    return None


def classify_period_columns(orig) -> dict[int, str]:
    """Map each period column to 'history' or 'projection'.

    A quarter column is history when its income-statement Net-revenue cell is a direct
    green link into an Input tab; otherwise projection. Annual (block) columns aggregate
    their four quarters, so they inherit the kind of the quarter immediately to their left.
    """
    is_ws = orig["IS"] if "IS" in orig.sheetnames else None
    rev_row = find_label_row(is_ws, "net revenue") if is_ws else None
    columns = period_columns(orig["BS"])
    kinds: dict[int, str] = {}
    for col, _label in columns:
        formula = is_ws.cell(rev_row, col).value if rev_row else None
        if isinstance(formula, str) and (
            "input financ" in formula.lower() or "input oper" in formula.lower()
        ):
            kinds[col] = "history"
        elif isinstance(formula, str) and formula.upper().replace(" ", "").startswith("=SUM("):
            kinds[col] = ""  # annual block — resolved from the left neighbour below
        else:
            kinds[col] = "projection"
    order = [col for col, _ in columns]
    for i, col in enumerate(order):
        if not kinds[col]:
            kinds[col] = (kinds.get(order[i - 1]) if i else "") or "projection"
    return kinds


def yb_blocks(orig):
    """[(annual_col, [q1, q2, q3, q4])] for each annual column closing a 4-quarter block."""
    cols = period_columns(orig["BS"])
    blocks = []
    for i, (col, label) in enumerate(cols):
        year = _YEAR_RE.match(label)
        if not year or i < 4:
            continue
        prev4 = cols[i - 4:i]
        quarters = [_QUARTER_RE.match(lbl) for _c, lbl in prev4]
        if (all(quarters)
                and [q.group(1) for q in quarters] == ["1", "2", "3", "4"]
                and all(q.group(2) == year.group(1)[2:] for q in quarters)):
            blocks.append((col, [c for c, _lbl in prev4]))
    return blocks


def _agg_op(formula: str):
    """Classify an annual cell's aggregation operation from its formula, else None."""
    s = formula.strip().upper().replace(" ", "")
    if s.startswith("=SUM("):
        return "sum"
    if "AVERAGE(" in s:
        return "avg"
    if _SINGLE_REF_RE.match(s):
        return "eop"
    return None


# ------------------------------------------------------------------------- invariants

def inv16_structure(orig, calc) -> InvariantResult:
    """INV-16: the workbook has the expected shape (sheets, period headers, balance-check row)."""
    problems = []
    for sheet in CORE_SHEETS:
        if sheet not in orig.sheetnames:
            problems.append(f"missing sheet '{sheet}'")
    if "BS" in orig.sheetnames:
        if find_label_row(orig["BS"], "balance check") is None:
            problems.append("BS has no 'Balance check' row")
        if not period_columns(orig["BS"]):
            problems.append("no period headers in row 5 of BS")
    passed = not problems
    msg = "structure ok" if passed else "; ".join(problems)
    return InvariantResult("structure", Severity.FAIL, passed, msg)


def inv0_no_unresolved(orig, calc) -> InvariantResult:
    """INV-0: every formula cell on the model tabs resolved to a value (proves recalc ran)."""
    failing = []
    for tab in MODEL_TABS:
        if tab not in orig.sheetnames or tab not in calc.sheetnames:
            continue
        ws_orig, ws_calc = orig[tab], calc[tab]
        for row in ws_orig.iter_rows():
            for cell in row:
                if cell.data_type == "f":  # a formula in the source workbook
                    if ws_calc.cell(cell.row, cell.column).value is None:
                        failing.append(CellRef(tab, cell.coordinate, actual=None))
    passed = not failing
    msg = (
        "all formula cells resolved to a value"
        if passed
        else f"{len(failing)} formula cell(s) blank after recalc (recalc did not run or a backend gap)"
    )
    return InvariantResult("no_unresolved_formulas", Severity.FAIL, passed, msg, failing[:50])


def inv2_balance_check(orig, calc) -> InvariantResult:
    """INV-2: the 'Balance check (= 0)' row is ~0 in every period column.

    A break in a PROJECTION column is an engine integration failure (FAIL). A break only in
    HISTORICAL columns means the reported input balance sheet does not foot (rounding or a
    transcription gap) and is a WARN — the engine links history verbatim, it does not compute it.
    """
    if "BS" not in calc.sheetnames:
        return InvariantResult("balance_check_zero", Severity.FAIL, False, "no BS sheet")
    ws = calc["BS"]
    row = find_label_row(ws, "balance check")
    if row is None:
        return InvariantResult(
            "balance_check_zero", Severity.FAIL, False, "no 'Balance check' row found in BS"
        )
    kinds = classify_period_columns(orig)
    hist_fail, proj_fail = [], []
    for col, label in period_columns(orig["BS"]):
        value = ws.cell(row, col).value
        if isinstance(value, (int, float)) and abs(value) > TOL:
            ref = CellRef("BS", f"{get_column_letter(col)}{row}", label, 0.0, value, abs(value))
            (proj_fail if kinds.get(col) == "projection" else hist_fail).append(ref)

    if proj_fail:
        worst = max(c.abs_error for c in proj_fail)
        msg = (
            f"balance check != 0 in {len(proj_fail)} projection column(s) "
            f"(engine integration); worst abs_err={worst:.4g}"
        )
        return InvariantResult("balance_check_zero", Severity.FAIL, False, msg, proj_fail + hist_fail)
    if hist_fail:
        worst = max(c.abs_error for c in hist_fail)
        msg = (
            f"balance check != 0 only in {len(hist_fail)} historical column(s): the reported "
            f"input balance sheet does not foot (rounding/gap); projections tie. "
            f"worst abs_err={worst:.4g}"
        )
        return InvariantResult("balance_check_zero", Severity.WARN, False, msg, hist_fail)
    return InvariantResult(
        "balance_check_zero", Severity.FAIL, True, "balance check is 0 in every column"
    )


def inv7_aggregation(orig, calc) -> InvariantResult:
    """INV-7/8/9: each annual block column aggregates its own 4 quarters correctly.

    The expected 4 quarters come from GEOMETRY (the block's columns), while the operation
    (SUM / AVERAGE / end-of-period) comes from the annual cell's own formula — so a wrong
    column range in the engine's aggregation is caught even though recalc honoured it.
    """
    blocks = yb_blocks(orig)
    col_label = {c: l for c, l in period_columns(orig["BS"])}
    failing = []
    for tab in MODEL_TABS:
        if tab not in orig.sheetnames or tab not in calc.sheetnames:
            continue
        ws_orig, ws_calc = orig[tab], calc[tab]
        for annual_col, qcols in blocks:
            for row in range(HDR + 1, ws_orig.max_row + 1):
                formula = ws_orig.cell(row, annual_col).value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                op = _agg_op(formula)
                if op is None:
                    continue
                annual_v = ws_calc.cell(row, annual_col).value
                if op in ("sum", "avg"):
                    q_vals = [ws_calc.cell(row, qc).value for qc in qcols]
                    if not isinstance(annual_v, (int, float)):
                        continue
                    if any(not isinstance(v, (int, float)) for v in q_vals):
                        continue
                    expected = sum(q_vals) if op == "sum" else sum(q_vals) / len(q_vals)
                    if abs(annual_v - expected) > max(TOL, TOL * abs(expected)):
                        failing.append(CellRef(
                            tab, f"{get_column_letter(annual_col)}{row}", col_label.get(annual_col),
                            expected, annual_v, abs(annual_v - expected),
                        ))
                else:  # single-cell ref (stock BOP/EOP) must point at one of the block's quarters
                    parts = _SINGLE_REF_PARTS.match(formula.strip().upper().replace(" ", ""))
                    block_letters = {get_column_letter(qc) for qc in qcols}
                    if parts and parts.group(1) not in block_letters:
                        failing.append(CellRef(
                            tab, f"{get_column_letter(annual_col)}{row}", col_label.get(annual_col),
                            actual=formula,
                        ))
    passed = not failing
    msg = (
        "annual columns aggregate their quarters correctly"
        if passed
        else f"{len(failing)} annual cell(s) do not match the aggregation of their 4 quarters"
    )
    return InvariantResult("annual_aggregation", Severity.FAIL, passed, msg, failing[:50])


def inv10_history_fidelity(orig, calc) -> InvariantResult:
    """INV-10: every green history link resolves to the value in its input cell.

    Catches an off-by-one in the engine's header/label lookups: a history cell that points
    at the wrong input row/column would differ from the value it claims to mirror.
    """
    failing = []
    for tab in ["IS", "BS", "CF", "Operational Model"]:
        if tab not in orig.sheetnames or tab not in calc.sheetnames:
            continue
        ws_orig, ws_calc = orig[tab], calc[tab]
        for row in ws_orig.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                match = _INPUT_LINK_RE.match(str(cell.value))
                if not match:
                    continue
                target_sheet = match.group(1)
                target_a1 = f"{match.group(2)}{match.group(3)}"
                if target_sheet not in calc.sheetnames:
                    continue
                model_v = ws_calc.cell(cell.row, cell.column).value
                input_v = calc[target_sheet][target_a1].value
                if not isinstance(model_v, (int, float)) or not isinstance(input_v, (int, float)):
                    continue
                if abs(model_v - input_v) > max(TOL, TOL * abs(input_v)):
                    failing.append(CellRef(
                        tab, cell.coordinate, expected=input_v, actual=model_v,
                        abs_error=abs(model_v - input_v),
                    ))
    passed = not failing
    msg = (
        "history links resolve to their input values"
        if passed
        else f"{len(failing)} history cell(s) differ from the input cell they link to"
    )
    return InvariantResult("history_fidelity", Severity.FAIL, passed, msg, failing[:50])


def inv12_no_error_cells(orig, calc) -> InvariantResult:
    """INV-12: no Excel error value anywhere in the recalculated workbook.

    The engine's intentional IFERROR(..., "-") sentinel is a plain '-' string and is
    therefore not flagged.
    """
    failing = []
    for tab in calc.sheetnames:
        ws = calc[tab]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in ERROR_VALUES:
                    failing.append(CellRef(tab, cell.coordinate, actual=cell.value))
    passed = not failing
    msg = "no error cells" if passed else f"{len(failing)} error cell(s) found"
    return InvariantResult("no_error_cells", Severity.FAIL, passed, msg, failing[:50])


def inv14_sign_conventions(orig, calc) -> InvariantResult:
    """INV-14: '(-)'-prefixed expense lines are <= 0 and top revenue/volume lines are >= 0.

    A soft guardrail (WARN): a wrong sign usually also breaks the balance check, but flagging
    it points straight at the offending driver. '(=)' subtotals and '(+/-)' lines are exempt.
    """
    failing = []
    for tab in ["IS", "CF", "Operational Model"]:
        if tab not in orig.sheetnames or tab not in calc.sheetnames:
            continue
        ws_orig, ws_calc = orig[tab], calc[tab]
        for row in range(HDR + 1, ws_orig.max_row + 1):
            label = ws_orig.cell(row, LABEL_COL).value
            if not isinstance(label, str):
                continue
            stripped = label.strip()
            low = stripped.lower()
            if stripped.startswith("(-)"):
                want = "neg"
            elif stripped.startswith("(+/-)") or stripped.startswith("(=)"):
                continue
            elif any(hint in low for hint in _REVENUE_HINTS):
                want = "pos"
            else:
                continue
            for col, plabel in period_columns(ws_orig):
                value = ws_calc.cell(row, col).value
                if not isinstance(value, (int, float)):
                    continue
                if want == "neg" and value > TOL:
                    failing.append(CellRef(tab, f"{get_column_letter(col)}{row}", plabel, actual=value))
                elif want == "pos" and value < -TOL:
                    failing.append(CellRef(tab, f"{get_column_letter(col)}{row}", plabel, actual=value))
    passed = not failing
    msg = (
        "sign conventions hold"
        if passed
        else f"{len(failing)} cell(s) break the (-)/revenue sign convention"
    )
    return InvariantResult("sign_conventions", Severity.WARN, passed, msg, failing[:30])


# Registry. Foundation item 4 appends its schedule tie-out checks here.
INVARIANTS = [
    inv16_structure,
    inv0_no_unresolved,
    inv2_balance_check,
    inv7_aggregation,
    inv10_history_fidelity,
    inv12_no_error_cells,
    inv14_sign_conventions,
]
