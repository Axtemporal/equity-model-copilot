"""Generates the blank Oil & Gas input workbook (pilot: Prio).

Column layout: 1Q16, 2Q16, 3Q16, 4Q16, 2016, 1Q17, ... 2025, 1Q26.
The year closes the block of four quarters. Quarters are input (blue);
annual columns are formulas (black): SUM for flows, 4Q link for stocks,
simple AVERAGE for rates/prices (weighted average will come from the engine).

Usage: python engine/build_input_template.py
Outputs: templates/input_template_oil_and_gas.xlsx and inputs/PRIO_inputs.xlsx
"""

from pathlib import Path
import shutil

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------- config
FIRST_YEAR, LAST_FULL_YEAR = 2016, 2025
EXTRA_QUARTERS = ["1Q26"]          # reported quarters of the current year
FIRST_DATA_COL = 4                 # column D
HEADER_ROW = 5
N_ASSETS = 8

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")            # hardcode/input
BLACK = Font(name=FONT, size=10, color="000000")           # formula
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True)
SUB = Font(name=FONT, size=9, italic=True, color="595959")
WHITE_BOLD = Font(name=FONT, size=10, bold=True, color="FFFFFF")

NAVY = PatternFill("solid", start_color="1F3864")
GRAY_Y = PatternFill("solid", start_color="EDEDED")        # annual column closes the block
NA_FILL = PatternFill("solid", start_color="BFBFBF")       # not-applicable cell
THIN_BOT = Border(bottom=Side(style="thin", color="9DC3E6"))

NUM_MN = '#,##0;(#,##0);"-"'
NUM_1D = '#,##0.0;(#,##0.0);"-"'
NUM_2D = '#,##0.00;(#,##0.00);"-"'
NUM_FX = "0.0000"

# ----------------------------------------------------------------------------- periods
def build_periods():
    """[('1Q16','Q'), ..., ('4Q16','Q'), ('2016','Y'), ..., ('1Q26','Q')]"""
    periods = []
    for year in range(FIRST_YEAR, LAST_FULL_YEAR + 1):
        yy = str(year)[2:]
        periods += [(f"{q}Q{yy}", "Q") for q in range(1, 5)]
        periods.append((str(year), "Y"))
    periods += [(q, "Q") for q in EXTRA_QUARTERS]
    return periods

PERIODS = build_periods()


def annual_formula(row, ycol, agg):
    """Annual-column formula from the 4 quarters to the left."""
    q1, q4 = get_column_letter(ycol - 4), get_column_letter(ycol - 1)
    rng = f"{q1}{row}:{q4}{row}"
    if agg == "sum":
        return f"=SUM({rng})"
    if agg == "avg":
        return f'=IFERROR(AVERAGE({rng}),"-")'
    if agg == "eop":
        return f"={q4}{row}"
    raise ValueError(agg)


# ----------------------------------------------------------------------------- helpers
def sheet_titles(ws, purpose):
    ws["B1"], ws["B1"].font = "PRIO — Equity Model Input (Oil & Gas template)", TITLE
    ws["B2"], ws["B2"].font = purpose, SUB
    ws["B3"] = ("Currency: USD mn unless noted (confirm reporting currency)  |  "
                "Sign convention: revenues (+), expenses (−)  |  Blue = input, black = formula")
    ws["B3"].font = SUB
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = f"{get_column_letter(FIRST_DATA_COL)}{HEADER_ROW + 1}"


def period_header(ws):
    for i, (label, kind) in enumerate(PERIODS):
        c = ws.cell(row=HEADER_ROW, column=FIRST_DATA_COL + i, value=label)
        c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BOT
        if kind == "Y":
            c.fill = GRAY_Y
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 9.5
    ws.cell(row=HEADER_ROW, column=2, value="Line item").font = BOLD
    ws.cell(row=HEADER_ROW, column=3, value="Unit").font = BOLD


def section(ws, row, text):
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = WHITE_BOLD
    for col in range(2, FIRST_DATA_COL + len(PERIODS)):
        ws.cell(row=row, column=col).fill = NAVY
    return row + 1


def data_row(ws, row, label, unit, agg, numfmt=NUM_MN, kind="input",
             formula=None, annual_only=False, comment=None):
    """kind: 'input' (typeable blue quarters) or 'calc' (black formula everywhere)."""
    ws.cell(row=row, column=2, value=label).font = BLACK
    ws.cell(row=row, column=3, value=unit).font = SUB
    for i, (_, pkind) in enumerate(PERIODS):
        col = FIRST_DATA_COL + i
        c = ws.cell(row=row, column=col)
        c.number_format = numfmt
        if pkind == "Y":
            c.fill = GRAY_Y
        if kind == "calc":
            c.value = formula(row, col) if formula else None
            c.font = BLACK
        elif pkind == "Q":
            if annual_only:
                c.fill = NA_FILL                      # annual series: quarter not applicable
            else:
                c.font = BLUE                         # data-entry cell
        else:                                         # annual column of an input row
            c.value = data_row_annual(row, col, agg, annual_only)
            c.font = BLUE if annual_only else BLACK
    if comment:
        ws.cell(row=row, column=2).comment = Comment(comment, "engine")
    return row + 1


def data_row_annual(row, col, agg, annual_only):
    if annual_only:
        return None                                   # typed directly into the year (blue)
    return annual_formula(row, col, agg)


# ----------------------------------------------------------------------------- sheets
def build_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = "808080"
    ws.column_dimensions["B"].width = 110
    lines = [
        ("PRIO — Input workbook (Oil & Gas template)", TITLE),
        ("Generated by engine/build_input_template.py. Version 0.1, 10/06/2026.", SUB),
        ("", None),
        ("HOW TO FILL IN", BOLD),
        ("1. Type only in the blue cells, copying the numbers as reported (release, ITR/DFP).", None),
        ("2. Quarter columns are input; the year column closes the block and calculates itself "
         "(sum for flows, 4Q for balances, simple average for prices and rates).", None),
        ("3. Older years without quarterly breakdown: overwrite the annual cell with the reported value "
         "(it becomes blue/hardcode, an accepted and documented practice).", None),
        ("4. Expenses and costs are entered with a negative sign. Revenues positive (sign convention 1).", None),
        ("5. Input Operational tab: name the assets in the ASSET SETUP block; the labels of the "
         "per-asset blocks update automatically. Unused assets stay blank.", None),
        ("6. Data the company does not disclose: leave blank. The AI will map the gaps and propose "
         "an estimation method for your approval (never enter an estimate as if it were reported).", None),
        ("", None),
        ("COLOR LEGEND", BOLD),
        ("Blue = input/hardcode  |  Black = formula  |  Green = link between tabs  |  "
         "Light gray = annual column  |  Dark gray = not applicable", None),
        ("", None),
        ("COMPLIANCE", BOLD),
        ("Public sources only. Every estimate will be labeled with method and source in the Assumptions tab. "
         "This material does not constitute an investment recommendation.", None),
    ]
    for i, (text, font) in enumerate(lines, start=2):
        c = ws.cell(row=i, column=2, value=text)
        if font:
            c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="top")


def build_financials(wb):
    ws = wb.create_sheet("Input Financials")
    ws.sheet_properties.tabColor = "2E75B6"
    sheet_titles(ws, "Financial history as reported (IFRS). Manual entry in the blue cells.")
    period_header(ws)
    r = HEADER_ROW + 2

    r = section(ws, r, "INCOME STATEMENT (as reported)")
    for label, unit, agg, fmt in [
        ("(=) Net revenue", "USD mn", "sum", NUM_MN),
        ("(-) Cost of goods sold", "USD mn", "sum", NUM_MN),
        ("(=) Gross profit", "USD mn", "sum", NUM_MN),
        ("(-) General & administrative expenses", "USD mn", "sum", NUM_MN),
        ("(-) Exploration expenses", "USD mn", "sum", NUM_MN),
        ("(+/-) Other operating income (expenses)", "USD mn", "sum", NUM_MN),
        ("(=) EBIT", "USD mn", "sum", NUM_MN),
        ("(+/-) Financial result", "USD mn", "sum", NUM_MN),
        ("(+/-) Equity income and other", "USD mn", "sum", NUM_MN),
        ("(=) EBT", "USD mn", "sum", NUM_MN),
        ("(-) Income taxes", "USD mn", "sum", NUM_MN),
        ("(=) Net income", "USD mn", "sum", NUM_MN),
        ("(-) Minority interest", "USD mn", "sum", NUM_MN),
        ("(=) Net income to shareholders", "USD mn", "sum", NUM_MN),
        ("memo: (+) Depreciation, depletion & amortization", "USD mn", "sum", NUM_MN),
        ("memo: (=) EBITDA (as disclosed)", "USD mn", "sum", NUM_MN),
        ("memo: EPS basic", "USD/sh", "sum", NUM_2D),
        ("memo: Shares outstanding (EOP)", "mn", "eop", NUM_1D),
    ]:
        r = data_row(ws, r, label, unit, agg, fmt)
    r += 1

    r = section(ws, r, "BALANCE SHEET (as reported)")
    bs_rows = [
        ("Cash and equivalents", NUM_MN), ("Short-term investments", NUM_MN),
        ("Trade receivables", NUM_MN), ("Inventories", NUM_MN),
        ("Other current assets", NUM_MN), ("(=) Total current assets", NUM_MN),
        ("PP&E", NUM_MN), ("Intangible assets and goodwill", NUM_MN),
        ("Right-of-use assets", NUM_MN), ("Deferred tax assets", NUM_MN),
        ("Other non-current assets", NUM_MN), ("(=) Total non-current assets", NUM_MN),
        ("(=) Total assets", NUM_MN),
        ("Trade payables", NUM_MN), ("Loans and financing (current)", NUM_MN),
        ("Lease liabilities (current)", NUM_MN), ("Other current liabilities", NUM_MN),
        ("(=) Total current liabilities", NUM_MN),
        ("Loans and financing (non-current)", NUM_MN), ("Lease liabilities (non-current)", NUM_MN),
        ("Provisions (incl. asset retirement obligation)", NUM_MN),
        ("Deferred tax liabilities", NUM_MN), ("Other non-current liabilities", NUM_MN),
        ("(=) Total liabilities", NUM_MN),
        ("Share capital and reserves", NUM_MN), ("Retained earnings (accumulated)", NUM_MN),
        ("(=) Equity attributable to shareholders", NUM_MN), ("Minority interest (equity)", NUM_MN),
        ("(=) Total equity", NUM_MN),
    ]
    bs_first = r
    for label, fmt in bs_rows:
        r = data_row(ws, r, label, "USD mn", "eop", fmt)
    row_assets = bs_first + 12          # (=) Total assets
    row_liab = bs_first + 23            # (=) Total liabilities
    row_eq = bs_first + 28              # (=) Total equity

    def check_formula(row, col):
        L = get_column_letter(col)
        return f"={L}{row_assets}-{L}{row_liab}-{L}{row_eq}"
    r = data_row(ws, r, "Balance check (assets − liab. − equity) = 0", "USD mn",
                 "sum", NUM_MN, kind="calc", formula=check_formula,
                 comment="Must be zero in every populated column.")
    r += 1

    r = section(ws, r, "CASH FLOW STATEMENT (as reported)")
    for label in [
        "(=) Net income (CF basis)", "(+) D&A and non-cash items",
        "(+/-) Working capital changes", "(+/-) Other operating",
        "(=) Cash from operations",
        "(-) Capex", "(+/-) Acquisitions and divestments", "(+/-) Other investing",
        "(=) Cash from investing",
        "(+) Debt raised", "(-) Debt repaid", "(-) Lease payments",
        "(-) Dividends and buybacks", "(+/-) Other financing",
        "(=) Cash from financing",
        "(+/-) FX effect on cash", "(=) Net change in cash",
        "Cash — beginning of period", "Cash — end of period",
    ]:
        r = data_row(ws, r, label, "USD mn", "sum", NUM_MN)


def build_operational(wb):
    ws = wb.create_sheet("Input Operational")
    ws.sheet_properties.tabColor = "2E75B6"
    sheet_titles(ws, "Operational history by asset. Adaptable structure: fill in only what the company discloses.")
    period_header(ws)
    r = HEADER_ROW + 2

    r = section(ws, r, "ASSET SETUP — name the assets/fields (blue). Unused: leave blank")
    asset_name_rows = []
    for i in range(1, N_ASSETS + 1):
        ws.cell(row=r, column=2, value=f"Asset {i} name").font = BLACK
        c = ws.cell(row=r, column=3)
        c.font, c.value = BLUE, f"Asset {i}"
        asset_name_rows.append(r)
        r += 1
    r += 1

    def per_asset_block(r, title, unit, agg, fmt, total_label, total_agg="sum"):
        r = section(ws, r, title)
        first = r
        for arow in asset_name_rows:
            ws.cell(row=r, column=2, value=f"=$C${arow}").font = BLACK
            r = data_row(ws, r, ws.cell(row=r, column=2).value, unit, agg, fmt)
            ws.cell(row=r - 1, column=2, value=f"=$C${arow}").font = BLACK
        def tot(row, col):
            L = get_column_letter(col)
            return f"=SUM({L}{first}:{L}{first + N_ASSETS - 1})"
        r = data_row(ws, r, total_label, unit, total_agg, fmt, kind="calc", formula=tot)
        return r + 1

    r = per_asset_block(r, "PRODUCTION BY ASSET", "kboe/d", "avg", NUM_1D,
                        "(=) Total production", total_agg="avg")
    r = per_asset_block(r, "LIFTING COST BY ASSET", "USD/boe", "avg", NUM_1D,
                        "(=) Sum (engine will compute weighted average)", total_agg="avg")
    r = per_asset_block(r, "CAPEX BY ASSET", "USD mn", "sum", NUM_MN, "(=) Total capex")

    r = section(ws, r, "CONSOLIDATED OPERATIONAL AND PRICES")
    price_rows = {}
    for key, (label, unit, agg, fmt) in {
        "sales": ("Sales volume", "kboe", "sum", NUM_MN),
        "realized": ("Realized price", "USD/boe", "avg", NUM_1D),
        "brent": ("Brent average", "USD/bbl", "avg", NUM_1D),
    }.items():
        price_rows[key] = r
        r = data_row(ws, r, label, unit, agg, fmt)
    def premium(row, col):
        L = get_column_letter(col)
        return f'=IFERROR({L}{price_rows["realized"]}-{L}{price_rows["brent"]},"-")'
    r = data_row(ws, r, "(=) Realized premium (discount) to Brent", "USD/bbl",
                 "avg", NUM_1D, kind="calc", formula=premium)
    for label, unit, agg, fmt in [
        ("Royalties and government take", "USD mn", "sum", NUM_MN),
        ("BRL/USD — average", "BRL", "avg", NUM_FX),
        ("BRL/USD — end of period", "BRL", "eop", NUM_FX),
    ]:
        r = data_row(ws, r, label, unit, agg, fmt)
    r += 1

    r = section(ws, r, "RESERVES (annual only, end of period)")
    for label in ["Reserves 1P", "Reserves 2P", "Reserves 3P"]:
        r = data_row(ws, r, label, "mmboe", "eop", NUM_1D, annual_only=True)


def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = "BF8F00"
    ws["B1"], ws["B1"].font = "Assumptions log — assumptions proposed by the AI and decided by the user", TITLE
    ws["B2"] = ("Nothing enters the model without approval. Every row records method, source with date and rationale. "
                "An estimate is never presented as reported data.")
    ws["B2"].font = SUB
    headers = ["ID", "Date", "Tab", "Line item", "Period(s)", "Proposed value / formula",
               "Method", "Source (link)", "Source date", "Rationale", "Status", "Decided by"]
    widths = [6, 11, 16, 34, 12, 24, 22, 34, 12, 50, 12, 12]
    for i, (h, w) in enumerate(zip(headers, widths), start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font, c.fill = WHITE_BOLD, NAVY
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B5"


def main():
    root = Path(__file__).resolve().parent.parent
    out_template = root / "templates" / "input_template_oil_and_gas.xlsx"
    out_prio = root / "inputs" / "PRIO_inputs.xlsx"
    out_template.parent.mkdir(exist_ok=True)
    out_prio.parent.mkdir(exist_ok=True)

    wb = Workbook()
    build_readme(wb)
    build_financials(wb)
    build_operational(wb)
    build_assumptions(wb)
    wb.save(out_template)
    shutil.copy(out_template, out_prio)
    print(f"OK: {out_template}")
    print(f"OK: {out_prio}")


if __name__ == "__main__":
    main()
