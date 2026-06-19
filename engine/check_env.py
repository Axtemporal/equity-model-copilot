"""Environment preflight for the Equity Model Copilot engine.

Run `python -m engine.check_env` after `pip install -r requirements.txt`, before the
first build. It confirms the things the engine needs that are easy to get wrong —
above all the **recalc backend**, which otherwise only fails deep inside the harness
(Phase 3.2), long after setup.

Design note (decision of 2026-06-18): the interface is Claude Code driving this engine
as a toolkit — there is no GUI and no Python->Claude bridge. So this check does NOT
probe Claude liveness: being in Claude Code already is the live Claude. It checks the
deterministic, brittle parts: Python version, the core imports, and — the real point —
that the recalc backend actually COMPUTES a formula, not merely that it imports.

All third-party imports happen *inside* the check functions, so this module runs (and
reports what is missing) even when the environment is broken.
"""
from __future__ import annotations

import contextlib
import importlib
import os
import sys
import tempfile
from typing import NamedTuple


class Check(NamedTuple):
    label: str
    ok: bool
    optional: bool
    detail: str
    hint: str = ""


_REQ_HINT = "pip install -r requirements.txt"
_LO_HINT = "winget install -e --id TheDocumentFoundation.LibreOffice"


def _version(dist: str) -> str:
    try:
        from importlib.metadata import version

        return version(dist)
    except Exception:  # noqa: BLE001
        return "versão desconhecida"


def _check_python() -> Check:
    v = sys.version_info
    detail = f"{v.major}.{v.minor}.{v.micro}"
    return Check("Python ≥ 3.10", v >= (3, 10), False, detail,
                 "Instale Python 3.10+ (Anaconda recomendado)")


def _check_import(label: str, module: str, *, dist: str | None = None,
                  optional: bool = False, hint: str = _REQ_HINT) -> Check:
    try:
        importlib.import_module(module)
    except Exception as exc:  # noqa: BLE001
        return Check(label, False, optional, f"não importável ({exc.__class__.__name__})", hint)
    return Check(label, True, optional, _version(dist or module))


def _check_recalc() -> Check:
    """The real point: prove the recalc backend resolves CROSS-SHEET links, not just imports.

    Builds a 2-sheet workbook with a cross-sheet reference ('Model'!A1 = 'Input'!A1 * 2)
    and recalculates it through the engine's own backend. Cross-sheet link resolution is
    the harness's hardest dependency — inv10_history_fidelity and inv2_balance_check read
    values across the Input tabs by formula — so a green here means the harness's recalc
    path works, not merely that same-sheet arithmetic does. (A same-sheet-only smoke would
    pass while a backend that mishandles cross-sheet refs silently breaks the harness.)
    """
    label = "recálculo (backend 'formulas')"
    try:
        import openpyxl
    except Exception:  # noqa: BLE001
        return Check(label, False, False, "openpyxl ausente (pré-requisito)", _REQ_HINT)
    try:
        from .harness.recalc import RecalcError, recalc
    except Exception as exc:  # noqa: BLE001
        return Check(label, False, False, f"harness.recalc não importável ({exc.__class__.__name__})", _REQ_HINT)

    # Pre-create both temp paths so the finally can always clean them — recalc() would
    # otherwise mkstemp its own output and leak it when the backend fails (the very case).
    fd, in_path = tempfile.mkstemp(prefix="emc_envcheck_", suffix=".xlsx")
    os.close(fd)
    fd, out_path = tempfile.mkstemp(prefix="emc_envcheck_out_", suffix=".xlsx")
    os.close(fd)
    try:
        wb = openpyxl.Workbook()
        src = wb.active
        src.title = "Input"
        src["A1"] = 21
        dst = wb.create_sheet("Model")
        dst["A1"] = "='Input'!$A$1*2"
        wb.save(in_path)
        # The 'formulas' backend prints a tqdm progress bar; mute it around this call
        # only (exceptions still propagate, so real failures are not swallowed).
        with open(os.devnull, "w") as devnull, \
                contextlib.redirect_stdout(devnull), contextlib.redirect_stderr(devnull):
            recalc(in_path, backend="formulas", out_path=out_path)
        rwb = openpyxl.load_workbook(out_path, data_only=True)
        # the backend uppercases sheet names on write, so match case-insensitively
        by_upper = {name.upper(): name for name in rwb.sheetnames}
        model = rwb[by_upper["MODEL"]] if "MODEL" in by_upper else rwb.active
        val = model["A1"].value
        if val is None:
            return Check(label, False, False, "recalculou mas não gravou valor (cache vazio)", _REQ_HINT)
        if abs(float(val) - 42.0) < 1e-6:
            return Check(label, True, False, "computou link cross-sheet ('Input'!A1*2 = 42) - caminho do harness OK")
        return Check(label, False, False, f"retornou {val!r}, esperado 42", _REQ_HINT)
    except RecalcError as exc:
        return Check(label, False, False, f"falhou: {exc}", _REQ_HINT)
    except Exception as exc:  # noqa: BLE001
        return Check(label, False, False, f"erro inesperado: {exc!r}", _REQ_HINT)
    finally:
        for path in (in_path, out_path):
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass


def _check_libreoffice() -> Check:
    label = "LibreOffice (cross-check opcional)"
    try:
        from .harness.recalc import _find_soffice
    except Exception:  # noqa: BLE001
        return Check(label, False, True, "harness.recalc não importável", "")
    path = _find_soffice()
    if path:
        return Check(label, True, True, path)
    return Check(label, False, True,
                 "não encontrado — backend de alta fidelidade indisponível", _LO_HINT)


def run_checks() -> list[Check]:
    """All checks, in display order. Pure: no printing, returns the results."""
    return [
        _check_python(),
        _check_import("openpyxl", "openpyxl"),
        _check_import("PyYAML", "yaml", dist="PyYAML"),
        _check_recalc(),
        _check_libreoffice(),
        _check_import("pytest (para a suite de testes)", "pytest",
                      optional=True, hint="pip install pytest"),
    ]


# Glyphs the report uses that a non-UTF console (Windows cp1252) cannot encode.
_ASCII_MAP = {"✓": "OK", "○": "--", "✗": "XX", "≥": ">=", "→": "->", "—": "-", "●": "*"}


def _supports_glyphs(stream) -> bool:
    enc = getattr(stream, "encoding", None) or "ascii"
    try:
        "".join(_ASCII_MAP).encode(enc)
        return True
    except Exception:  # noqa: BLE001
        return False


def main(argv: list[str] | None = None) -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # ✓/○ glyphs + accents on Windows
    except Exception:  # noqa: BLE001
        pass

    glyphs_ok = _supports_glyphs(sys.stdout)
    enc = getattr(sys.stdout, "encoding", None) or "ascii"

    def out(text: str) -> None:
        if not glyphs_ok:
            text = text.translate({ord(k): v for k, v in _ASCII_MAP.items()})
        try:
            print(text)
        except UnicodeEncodeError:  # last-resort backstop: never crash the diagnostic
            print(text.encode(enc, errors="replace").decode(enc, errors="replace"))

    checks = run_checks()
    out("Equity Model Copilot — verificação de ambiente\n")
    for c in checks:
        mark = "✓" if c.ok else ("○" if c.optional else "✗")
        out(f"  {mark} {c.label}: {c.detail}")
        if not c.ok and c.hint:
            out(f"      → {c.hint}")
    out("")

    critical = [c for c in checks if not c.ok and not c.optional]
    if critical:
        n = len(critical)
        out(f"Ambiente incompleto: {n} item(ns) crítico(s) faltando.   (exit 1)")
        return 1
    if any(not c.ok and c.optional for c in checks):
        out("Ambiente pronto para build (itens opcionais ausentes não bloqueiam).   (exit 0)")
    else:
        out("Ambiente pronto para build.   (exit 0)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
