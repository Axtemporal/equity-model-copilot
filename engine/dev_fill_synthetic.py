"""Fills the input template with coherent SYNTHETIC data, for engine testing.

Not real data from any company. 13 quarters (1Q23 to 1Q26), 3 fictitious fields.
Coherence is by construction: cash is the residual consistent with the CF, PP&E and
RE follow roll-forwards, working capital is constant. The balance sheet closes in
every period.

Usage: python engine/dev_fill_synthetic.py <template.xlsx> <output.xlsx>
"""

import sys
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=10, color="0000FF")
FILLED_QUARTERS = [f"{q}Q{y}" for y in (23, 24, 25) for q in (1, 2, 3, 4)] + ["1Q26"]
ASSETS = ["Field Alpha", "Field Bravo", "Field Charlie"]


def col_of(ws, label):
    for c in range(4, ws.max_column + 1):
        if ws.cell(row=5, column=c).value == label:
            return c
    raise KeyError(label)


def row_of(ws, label, col=2):
    for r in range(6, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip() == label:
            return r
    raise KeyError(label)


def put(ws, row, col, value):
    c = ws.cell(row=row, column=col)
    c.value = round(value, 4) if isinstance(value, float) else value
    c.font = BLUE


def main(template_path, out_path):
    wb = load_workbook(template_path)

    # label as synthetic on every titled tab
    for name in ("Input Financials", "Input Operational"):
        ws = wb[name]
        ws["B1"] = "SYNTH — FICTITIOUS DATA FOR ENGINE TESTING (not PRIO)"

    fin = wb["Input Financials"]
    op = wb["Input Operational"]
    qcols_f = {q: col_of(fin, q) for q in FILLED_QUARTERS}
    qcols_o = {q: col_of(op, q) for q in FILLED_QUARTERS}

    # ----------------------------------------------------------- operational
    # production by field (kboe/d) with slight growth; stable prices
    base_prod = {"Field Alpha": 45.0, "Field Bravo": 30.0, "Field Charlie": 15.0}
    lifting = {"Field Alpha": 7.0, "Field Bravo": 9.0, "Field Charlie": 12.0}
    setup_first = row_of(op, "Asset 1 name")
    for i, a in enumerate(ASSETS):
        put(op, setup_first + i, 3, a)

    prod_rows = {}
    sec_prod = row_of(op, "PRODUCTION BY ASSET")          # section row
    sec_lift = row_of(op, "LIFTING COST BY ASSET")
    sec_capex = row_of(op, "CAPEX BY ASSET")

    state = {}
    for qi, q in enumerate(FILLED_QUARTERS):
        growth = 1.0 + 0.01 * qi
        days = 91.25
        brent = 80.0 + (2.0 if qi % 4 == 1 else -1.0 if qi % 4 == 3 else 0.0)
        realized = brent - 3.0
        total_prod = 0.0
        for i, a in enumerate(ASSETS):
            p = base_prod[a] * growth
            put(op, sec_prod + 1 + i, qcols_o[q], p)
            put(op, sec_lift + 1 + i, qcols_o[q], lifting[a])
            put(op, sec_capex + 1 + i, qcols_o[q], 30.0 if i == 0 else 20.0 if i == 1 else 10.0)
            total_prod += p
        volume = total_prod * days                          # kboe in the quarter
        lift_avg = sum(base_prod[a] * growth * lifting[a] for a in ASSETS) / total_prod

        put(op, row_of(op, "Sales volume"), qcols_o[q], volume)
        put(op, row_of(op, "Realized price"), qcols_o[q], realized)
        put(op, row_of(op, "Brent average"), qcols_o[q], brent)
        put(op, row_of(op, "Royalties and government take"), qcols_o[q], volume * realized / 1000 * 0.10)
        put(op, row_of(op, "BRL/USD — average"), qcols_o[q], 5.10)
        put(op, row_of(op, "BRL/USD — end of period"), qcols_o[q], 5.15)
        state[q] = dict(volume=volume, realized=realized, lift_avg=lift_avg, capex=60.0)

    # reserves: annual only (2023, 2024, 2025)
    for year, v1, v2, v3 in [("2023", 180, 320, 420), ("2024", 200, 340, 450), ("2025", 220, 360, 470)]:
        ycol = col_of(op, year)
        put(op, row_of(op, "Reserves 1P"), ycol, v1)
        put(op, row_of(op, "Reserves 2P"), ycol, v2)
        put(op, row_of(op, "Reserves 3P"), ycol, v3)

    # ----------------------------------------------------------- financials
    r = {lbl: row_of(fin, lbl) for lbl in [
        "(=) Net revenue", "(-) Cost of goods sold", "(=) Gross profit",
        "(-) General & administrative expenses", "(-) Exploration expenses",
        "(+/-) Other operating income (expenses)", "(=) EBIT",
        "(+/-) Financial result", "(+/-) Equity income and other", "(=) EBT",
        "(-) Income taxes", "(=) Net income", "(-) Minority interest",
        "(=) Net income to shareholders",
        "memo: (+) Depreciation, depletion & amortization",
        "memo: (=) EBITDA (as disclosed)", "memo: EPS basic",
        "memo: Shares outstanding (EOP)",
    ]}
    b = {lbl: row_of(fin, lbl) for lbl in [
        "Cash and equivalents", "Short-term investments", "Trade receivables",
        "Inventories", "Other current assets", "(=) Total current assets",
        "PP&E", "Intangible assets and goodwill", "Right-of-use assets",
        "Deferred tax assets", "Other non-current assets",
        "(=) Total non-current assets", "(=) Total assets",
        "Trade payables", "Loans and financing (current)",
        "Lease liabilities (current)", "Other current liabilities",
        "(=) Total current liabilities", "Loans and financing (non-current)",
        "Lease liabilities (non-current)",
        "Provisions (incl. asset retirement obligation)",
        "Deferred tax liabilities", "Other non-current liabilities",
        "(=) Total liabilities", "Share capital and reserves",
        "Retained earnings (accumulated)", "(=) Equity attributable to shareholders",
        "Minority interest (equity)", "(=) Total equity",
    ]}
    c = {lbl: row_of(fin, lbl) for lbl in [
        "(=) Net income (CF basis)", "(+) D&A and non-cash items",
        "(+/-) Working capital changes", "(+/-) Other operating",
        "(=) Cash from operations", "(-) Capex",
        "(+/-) Acquisitions and divestments", "(+/-) Other investing",
        "(=) Cash from investing", "(+) Debt raised", "(-) Debt repaid",
        "(-) Lease payments", "(-) Dividends and buybacks", "(+/-) Other financing",
        "(=) Cash from financing", "(+/-) FX effect on cash",
        "(=) Net change in cash", "Cash — beginning of period",
        "Cash — end of period",
    ]}

    # balance sheet constants (working capital deliberately flat)
    K = dict(sti=50.0, ar=180.0, inv=90.0, oca=40.0, intan=120.0, rou=60.0,
             dta=30.0, onca=45.0, ap=110.0, stdebt=100.0, stlease=15.0,
             ocl=55.0, ltdebt=700.0, ltlease=45.0, prov=250.0, dtl=80.0,
             oncl=35.0, capital=1000.0, mi=0.0)
    ppe, re, cash = 1500.0, 400.0, None
    depl_rate = 9.0          # depletion in USD/boe
    shares = 500.0

    for q in FILLED_QUARTERS:
        st, fc = state[q], qcols_f[q]
        rev = st["volume"] * st["realized"] / 1000.0
        royal = rev * 0.10
        depl = st["volume"] * depl_rate / 1000.0
        cogs = -(st["volume"] * st["lift_avg"] / 1000.0 + royal + depl)
        ga, expl, other = -rev * 0.04, -8.0, 0.0
        ebit = rev + cogs + ga + expl + other
        finres, eq_inc = -12.0, 0.0
        ebt = ebit + finres + eq_inc
        tax = -max(ebt, 0.0) * 0.25
        ni = ebt + tax

        put(fin, r["(=) Net revenue"], fc, rev)
        put(fin, r["(-) Cost of goods sold"], fc, cogs)
        put(fin, r["(=) Gross profit"], fc, rev + cogs)
        put(fin, r["(-) General & administrative expenses"], fc, ga)
        put(fin, r["(-) Exploration expenses"], fc, expl)
        put(fin, r["(+/-) Other operating income (expenses)"], fc, other)
        put(fin, r["(=) EBIT"], fc, ebit)
        put(fin, r["(+/-) Financial result"], fc, finres)
        put(fin, r["(+/-) Equity income and other"], fc, eq_inc)
        put(fin, r["(=) EBT"], fc, ebt)
        put(fin, r["(-) Income taxes"], fc, tax)
        put(fin, r["(=) Net income"], fc, ni)
        put(fin, r["(-) Minority interest"], fc, 0.0)
        put(fin, r["(=) Net income to shareholders"], fc, ni)
        put(fin, r["memo: (+) Depreciation, depletion & amortization"], fc, depl)
        put(fin, r["memo: (=) EBITDA (as disclosed)"], fc, ebit + depl)
        put(fin, r["memo: EPS basic"], fc, ni / shares)
        put(fin, r["memo: Shares outstanding (EOP)"], fc, shares)

        capex = st["capex"]
        cfo = ni + depl
        prev_cash = cash
        ppe = ppe + capex - depl
        re = re + ni
        liab = (K["ap"] + K["stdebt"] + K["stlease"] + K["ocl"] + K["ltdebt"]
                + K["ltlease"] + K["prov"] + K["dtl"] + K["oncl"])
        equity = K["capital"] + re + K["mi"]
        non_cash_assets = (K["sti"] + K["ar"] + K["inv"] + K["oca"] + ppe
                           + K["intan"] + K["rou"] + K["dta"] + K["onca"])
        cash = liab + equity - non_cash_assets          # residual: balance sheet closes

        put(fin, b["Cash and equivalents"], fc, cash)
        for lbl, key in [("Short-term investments", "sti"), ("Trade receivables", "ar"),
                         ("Inventories", "inv"), ("Other current assets", "oca"),
                         ("Intangible assets and goodwill", "intan"),
                         ("Right-of-use assets", "rou"), ("Deferred tax assets", "dta"),
                         ("Other non-current assets", "onca"), ("Trade payables", "ap"),
                         ("Loans and financing (current)", "stdebt"),
                         ("Lease liabilities (current)", "stlease"),
                         ("Other current liabilities", "ocl"),
                         ("Loans and financing (non-current)", "ltdebt"),
                         ("Lease liabilities (non-current)", "ltlease"),
                         ("Provisions (incl. asset retirement obligation)", "prov"),
                         ("Deferred tax liabilities", "dtl"),
                         ("Other non-current liabilities", "oncl"),
                         ("Share capital and reserves", "capital"),
                         ("Minority interest (equity)", "mi")]:
            put(fin, b[lbl], fc, K[key])
        put(fin, b["PP&E"], fc, ppe)
        put(fin, b["Retained earnings (accumulated)"], fc, re)
        tca = cash + K["sti"] + K["ar"] + K["inv"] + K["oca"]
        tnca = ppe + K["intan"] + K["rou"] + K["dta"] + K["onca"]
        put(fin, b["(=) Total current assets"], fc, tca)
        put(fin, b["(=) Total non-current assets"], fc, tnca)
        put(fin, b["(=) Total assets"], fc, tca + tnca)
        tcl = K["ap"] + K["stdebt"] + K["stlease"] + K["ocl"]
        put(fin, b["(=) Total current liabilities"], fc, tcl)
        put(fin, b["(=) Total liabilities"], fc, liab)
        put(fin, b["(=) Equity attributable to shareholders"], fc, K["capital"] + re)
        put(fin, b["(=) Total equity"], fc, equity)

        delta_cash = cfo - capex
        put(fin, c["(=) Net income (CF basis)"], fc, ni)
        put(fin, c["(+) D&A and non-cash items"], fc, depl)
        put(fin, c["(+/-) Working capital changes"], fc, 0.0)
        put(fin, c["(+/-) Other operating"], fc, 0.0)
        put(fin, c["(=) Cash from operations"], fc, cfo)
        put(fin, c["(-) Capex"], fc, -capex)
        put(fin, c["(+/-) Acquisitions and divestments"], fc, 0.0)
        put(fin, c["(+/-) Other investing"], fc, 0.0)
        put(fin, c["(=) Cash from investing"], fc, -capex)
        for lbl in ["(+) Debt raised", "(-) Debt repaid", "(-) Lease payments",
                    "(-) Dividends and buybacks", "(+/-) Other financing",
                    "(=) Cash from financing", "(+/-) FX effect on cash"]:
            put(fin, c[lbl], fc, 0.0)
        put(fin, c["(=) Net change in cash"], fc, delta_cash)
        bop = cash - delta_cash if prev_cash is None else prev_cash
        put(fin, c["Cash — beginning of period"], fc, bop)
        put(fin, c["Cash — end of period"], fc, cash)

    wb.save(out_path)
    print(f"OK: {out_path}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
