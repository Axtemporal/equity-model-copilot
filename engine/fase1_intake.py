"""Fase 1 intake orchestrator (steps 1.2-1.10) — the deterministic spine Claude Code drives.

Pure-Python pipeline; the AI/conversation handles ONLY the residual + the gap/currency
judgments, consuming this output:
  flatten -> detect + coverage-gate sector -> resolve labels (3-pass) -> classify roles ->
  content sufficiency -> Manifest.

The input may be one messy tab mixing statements, capital-structure facts and reported
aggregates, so flatten operates on a FLAT line list across all sheets, not a fixed layout.
Writing the canonical adjusted-input .xlsx and the AI residual/currency steps plug on top of
the Manifest this produces. Run as: python -m engine.fase1_intake <input.xlsx>
"""
from __future__ import annotations

import openpyxl
from openpyxl.utils import get_column_letter

from . import canonical_schema as cs
from . import label_resolver as lr
from . import period_grid as pg
from . import role_classifier as rc
from . import sector_coverage as scov
from .fase1_manifest import (
    CONTEXTUAL, MAPPED, SLOT_ABSENT, SLOT_MAPPED, LineRecord, Manifest, SlotCoverage,
)
from .harness.invariants import FIRST_COL, HDR, LABEL_COL
from .harness.validator import detect_sector
from .template_loader import identify_company

_SKIP_SHEETS = ("assumptions", "readme", "premises")


def _find_header_row(ws) -> int:
    """The row with the most parseable period headers (falls back to the template's row 5)."""
    best_row, best_n = HDR, 0
    for row in range(1, min(ws.max_row, 40) + 1):
        n = sum(1 for col in range(1, ws.max_column + 1)
                if pg.parse_period(ws.cell(row, col).value) is not None)
        if n > best_n:
            best_row, best_n = row, n
    return best_row


def flatten(wb) -> list[LineRecord]:
    """Collapse the workbook (any layout) into a flat list of labeled line records.

    Period columns are normalized (tolerant parse) and REORDERED into the canonical interleaved
    grid (1Q 2Q 3Q 4Q ANO …) via period_grid, so a messy layout — quarters then a blank gap then
    annuals, PT spellings, etc. — comes out canonical regardless of how it was entered.
    """
    records: list[LineRecord] = []
    for sheet in wb.sheetnames:
        if sheet.strip().lower().startswith(_SKIP_SHEETS):
            continue
        ws = wb[sheet]
        hdr = _find_header_row(ws)
        headers = [
            (col, ws.cell(hdr, col).value,
             any(ws.cell(r, col).value is not None for r in range(hdr + 1, ws.max_row + 1)))
            for col in range(1, ws.max_column + 1)
        ]
        grid = pg.normalize_grid(headers)
        token_to_col: dict[str, int] = {}
        for col, token in grid.col_to_token.items():
            token_to_col.setdefault(token, col)        # first column wins on duplicate headers

        first_pc = min(grid.col_to_token, default=FIRST_COL)
        label_col = LABEL_COL if LABEL_COL < first_pc else 1
        unit_col = label_col + 1 if (label_col + 1) < first_pc else None

        for row in range(hdr + 1, ws.max_row + 1):
            label = ws.cell(row, label_col).value
            if not isinstance(label, str) or not label.strip():
                continue
            values = {}
            for token in grid.order:                   # canonical interleaved order
                value = ws.cell(row, token_to_col[token]).value
                if value is not None:
                    values[token] = value
            unit = ws.cell(row, unit_col).value if unit_col else None
            records.append(LineRecord(
                raw_label=label.strip(),
                sheet=sheet,
                cell=f"{sheet}!{get_column_letter(label_col)}{row}",
                unit=unit.strip() if isinstance(unit, str) else None,
                values=values,
            ))
    return records


def analyze(input_path: str, sector: str | None = None) -> Manifest:
    """Run the deterministic Fase 1 spine and return the Manifest (raises if sector is blocked)."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    fin = wb["Input Financials"] if "Input Financials" in wb.sheetnames else None
    company = identify_company(fin) if fin is not None else "Unknown"

    sector = sector or detect_sector(wb)
    if sector:
        scov.assert_supported(sector)            # coverage gate (raises SectorCoverageError)

    records = flatten(wb)

    if sector:
        matches: dict[str, lr.Match] = {}
        for m in lr.resolve([r.raw_label for r in records], sector):
            matches.setdefault(m.raw, m)         # first occurrence wins (label dedupe parity)
        for rec in records:
            m = matches.get(rec.raw_label)
            if m and m.canonical:
                rec.canonical, rec.method, rec.confidence = m.canonical, m.method, m.score
                rec.role = rc.classify(m.canonical, sector)
                rec.disposition = MAPPED
                rec.provenance = f"{m.method} match"
            else:
                rec.role = rc.CONTEXTUAL
                rec.disposition = CONTEXTUAL

    coverage: list[SlotCoverage] = []
    if sector:
        mapped_canonical = {r.canonical for r in records if r.disposition == MAPPED}
        for sheet, labels in cs.required_labels(sector).items():
            for slot in labels:
                status = SLOT_MAPPED if slot in mapped_canonical else SLOT_ABSENT
                coverage.append(SlotCoverage(slot, sheet, status))

    residual = [r.raw_label for r in records if r.canonical is None]
    return Manifest(company=company, sector=sector, lines=records,
                    coverage=coverage, residual=residual)


# ─── 1.9 — materialize the canonical adjusted input ─────────────────────────────

_HDR_ROW, _LABEL_COL, _UNIT_COL, _FIRST_COL = 5, 2, 3, 4
_PLACED_ROLES = (rc.STATEMENT, rc.CAPITAL_STRUCTURE, rc.REPORTED_CHECK)


def _period_sort_key(token: str):
    period = pg.parse_period(token)
    return period.order_key if period else (9999, 9)


def write_adjusted_input(manifest: Manifest, source, out_path: str) -> str:
    """Write the canonical adjusted-input .xlsx from a Fase 1 manifest (step 1.9).

    Financial sheet: rebuilt with the canonical labels (required base + delta + the mapped
    structural / reported-check lines), values mapped from the resolved records. Sub-lines that
    resolved to the SAME canonical are **summed** into one row (rollup) — the checksum against a
    reported subtotal is a separate safeguard (not yet wired). Operational sheet: carried over
    as-is, because operational organization is deferred to Fase 2/3 (no fixed canonical schema).
    The result balances/builds only if the source operational was already canonical.
    """
    src = openpyxl.load_workbook(source, data_only=True) if isinstance(source, str) else source
    sector = manifest.sector

    # rows to emit: required financial labels (schema order) + mapped structural/reported-checks
    grouped: dict[str, list] = {}
    for rec in manifest.lines:
        if rec.canonical and rec.role in _PLACED_ROLES:
            grouped.setdefault(rec.canonical, []).append(rec)

    row_labels = list(cs.required_labels(sector)[cs.FINANCIALS]) if sector else list(cs.UNIVERSAL_BASE_FINANCIAL)
    for canonical in list(cs.STRUCTURAL) + list(grouped):
        if canonical not in row_labels:
            row_labels.append(canonical)

    # canonical interleaved period columns (union over all placed records)
    tokens = {tok for recs in grouped.values() for rec in recs for tok in rec.values}
    periods = sorted(tokens, key=_period_sort_key)

    wb = openpyxl.Workbook()
    fin = wb.active
    fin.title = cs.FINANCIALS
    fin.cell(1, _LABEL_COL, manifest.company)
    fin.cell(2, _LABEL_COL, "Financial history (canonical — generated by Fase 1 intake)")
    fin.cell(_HDR_ROW, _LABEL_COL, "Line item")
    fin.cell(_HDR_ROW, _UNIT_COL, "Unit")
    for j, token in enumerate(periods):
        fin.cell(_HDR_ROW, _FIRST_COL + j, token)

    for i, label in enumerate(row_labels):
        row = _HDR_ROW + 1 + i
        fin.cell(row, _LABEL_COL, label)
        recs = grouped.get(label, [])
        unit = next((r.unit for r in recs if r.unit), None)
        if unit:
            fin.cell(row, _UNIT_COL, unit)
        for j, token in enumerate(periods):
            total = None
            for rec in recs:                       # rollup: sum sub-lines mapped to this canonical
                value = rec.values.get(token)
                if isinstance(value, (int, float)):
                    total = (total or 0) + value
            if total is not None:
                fin.cell(row, _FIRST_COL + j, total)

    # carry the operational sheet verbatim (organization is Fase 2/3)
    if cs.OPERATIONAL in src.sheetnames:
        src_op = src[cs.OPERATIONAL]
        op = wb.create_sheet(cs.OPERATIONAL)
        for src_row in src_op.iter_rows():
            for cell in src_row:
                if cell.value is not None:
                    op.cell(cell.row, cell.column, cell.value)

    wb.save(out_path)
    return out_path


def build_adjusted_input(source_path: str, out_path: str, sector: str | None = None):
    """Convenience: analyze the source and materialize the canonical adjusted input.

    Returns (out_path, Manifest). Raises SectorCoverageError if the sector is blocked.
    """
    manifest = analyze(source_path, sector)
    write_adjusted_input(manifest, source_path, out_path)
    return out_path, manifest


def _report(manifest: Manifest) -> str:
    mapped = sum(1 for line in manifest.lines if line.disposition == MAPPED)
    out = [
        f"Fase 1 intake — {manifest.company} [{manifest.sector or 'sector?'}]",
        f"  lines: {len(manifest.lines)} ({mapped} mapped, {len(manifest.residual)} residual)",
        f"  content sufficient: {manifest.content_sufficient}",
    ]
    if manifest.missing_required:
        out.append("  MISSING required slots:")
        for cov in manifest.missing_required:
            out.append(f"    - [{cov.sheet}] {cov.slot}")
    if manifest.residual:
        out.append("  residual (for the AI semantic pass / your review):")
        for raw in manifest.residual[:20]:
            out.append(f"    - {raw}")
    return "\n".join(out)


if __name__ == "__main__":
    import sys

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass

    if len(sys.argv) < 2:
        print("usage: python -m engine.fase1_intake <input.xlsx> [sector]")
        sys.exit(2)
    sec = sys.argv[2] if len(sys.argv) > 2 else None
    try:
        manifest = analyze(sys.argv[1], sec)
    except scov.SectorCoverageError as exc:
        print("BLOCKED:", exc)
        sys.exit(3)
    print(_report(manifest))
