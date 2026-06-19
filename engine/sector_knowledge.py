"""Sector method-card grounding for the assumption session (Stage 3).

The session should not propose premises from model memory — it should read the sector's method
card in knowledge/sector_modeling_rules/sectors/<sector>.md and ground each proposal in it (the revenue
identity, the drivers, the accounting treatments, the pitfalls), citing the card. This module
locates the card for an input, loads it (cached), and extracts the slice most relevant to the
premise being discussed, so the prompt stays small but on-point.

Public surface:
    detect_sector_from_input(input_path) -> slug | None
    card_excerpt(sector, line_item)      -> grounding text (or "" if no card)
"""
from __future__ import annotations

import re
from functools import lru_cache

from .template_loader import CARDS_DIR, identify_sector


def detect_sector_from_input(input_path: str) -> str | None:
    """Identify the sector from the operational input tab (signal heuristic; None if unsure)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(input_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if "Input Operational" not in wb.sheetnames:
            return None
        labels: dict[str, bool] = {}
        for row in wb["Input Operational"].iter_rows(min_col=2, max_col=2):
            v = row[0].value
            if isinstance(v, str) and v.strip():
                labels[v.strip()] = True
        return identify_sector(labels)
    finally:
        wb.close()


@lru_cache(maxsize=16)
def load_card(sector: str | None) -> str:
    if not sector:
        return ""
    path = CARDS_DIR / f"{sector}.md"
    return path.read_text(encoding="utf-8") if path.exists() else ""


def _blocks(text: str) -> tuple[str, list[tuple[str, str]]]:
    """Split markdown into (intro, [(heading, body), ...])."""
    parts = re.split(r"(?m)^(#{1,6} .*)$", text)
    intro = parts[0].strip()
    blocks: list[tuple[str, str]] = []
    rest = parts[1:]
    for i in range(0, len(rest) - 1, 2):
        blocks.append((rest[i].strip(), rest[i + 1].strip()))
    return intro, blocks


def card_excerpt(sector: str | None, line_item: str, max_chars: int = 2600) -> str:
    """The card's thesis head plus the section(s) most relevant to `line_item`, char-budgeted."""
    text = load_card(sector)
    if not text:
        return ""
    intro, blocks = _blocks(text)
    keywords = [w.lower() for w in re.findall(r"[A-Za-zÀ-ÿ]{4,}", line_item)]

    def score(heading: str, body: str) -> int:
        hay = (heading + " " + body).lower()
        return sum(hay.count(k) for k in keywords)

    ranked = sorted(blocks, key=lambda hb: score(*hb), reverse=True)
    out: list[str] = []
    budget = max_chars
    head = intro[:900].strip()
    if head:
        out.append(head); budget -= len(head)
    for heading, body in ranked:
        if score(heading, body) == 0:
            break
        chunk = f"{heading}\n{body}"[:900]
        if budget - len(chunk) < 0:
            break
        out.append(chunk); budget -= len(chunk)
    return "\n\n".join(out)[:max_chars]
