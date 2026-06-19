"""Engine v0: reads the filled input workbook and builds the integrated model.

The model is written as a copy of the input + new tabs (Cover, Premises,
Operational Model, IS, BS, CF, Schedules), so the input tabs stay inside the
final file and all history on the model tabs is a formula linking the input
(green). Projection: quarterly through the end of the year after the last
historical quarter, then annual through year 10. Assumptions live in the
Premises tab (blue, auto-seeded from the last historical value) and are the
model's only editable cells. Dynamic schedules: working capital (DSO/DPO/days),
debt (interest on the opening balance), IFRS-16 leases (RoU + lease liability
roll-forward), a simple revolver (keeps cash >= a minimum via MIN/MAX), and ARO.
No circularity: interest is always charged on the opening balance.

Usage: python engine/build_model.py <filled_input.xlsx> <output_model.xlsx>
"""

import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from template_loader import analyze_input, BuildDiagnostics

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10, color="000000")
GREEN = Font(name=FONT, size=10, color="006100")
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True)
SUB = Font(name=FONT, size=9, italic=True, color="595959")
WHITE_BOLD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
NAVY = PatternFill("solid", start_color="1F3864")
GRAY_Y = PatternFill("solid", start_color="EDEDED")
PROJ_FILL = PatternFill("solid", start_color="FFF7E6")   # marks projection columns
THIN_BOT = Border(bottom=Side(style="thin", color="9DC3E6"))
NUM_MN = '#,##0;(#,##0);"-"'
NUM_1D = '#,##0.0;(#,##0.0);"-"'
NUM_2D = '#,##0.00;(#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'
FIRST_COL = 4
HDR = 5
PROJ_YEARS_ANNUAL = 8        # annual years beyond the quarterly block (through year 10)

L = get_column_letter


# --------------------------------------------------------------------- input parsing
def header_map(ws):
    return {ws.cell(row=HDR, column=c).value: c
            for c in range(FIRST_COL, ws.max_column + 1)
            if ws.cell(row=HDR, column=c).value}


def label_rows(ws):
    out = {}
    for r in range(HDR + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.strip() and v not in out:
            out[v.strip()] = r
    return out


def q_seq(label):  # '3Q24' -> (2024, 3)
    return 2000 + int(label[2:]), int(label[0])


def detect_history(fin):
    hdr = header_map(fin)
    rev_row = label_rows(fin)["(=) Net revenue"]
    filled = [p for p, c in hdr.items()
              if "Q" in p and fin.cell(row=rev_row, column=c).value is not None]
    filled.sort(key=q_seq)
    if not filled:
        raise SystemExit("input has no filled quarter on the Net revenue line")
    return filled


# --------------------------------------------------------------------- timeline
def build_timeline(hist_quarters):
    """List of model periods: dicts {label, kind, year, q}.

    kind: QH historical quarter | QP projected quarter | YB year closing a block
          of 4 quarters | YS pure annual year (no quarters)
    """
    first_y = q_seq(hist_quarters[0])[0]
    last_y, last_q = q_seq(hist_quarters[-1])
    proj_end_y = last_y + 1
    tl = []
    for y in range(first_y, proj_end_y + 1):
        for q in range(1, 5):
            lbl = f"{q}Q{str(y)[2:]}"
            if (y, q) <= (last_y, last_q):
                kind = "QH" if lbl in hist_quarters else "QP"
                tl.append(dict(label=lbl, kind=kind, year=y, q=q))
            else:
                tl.append(dict(label=lbl, kind="QP", year=y, q=q))
        tl.append(dict(label=str(y), kind="YB", year=y, q=None))
    for y in range(proj_end_y + 1, proj_end_y + 1 + PROJ_YEARS_ANNUAL):
        tl.append(dict(label=str(y), kind="YS", year=y, q=None))
    for i, p in enumerate(tl):
        p["col"] = FIRST_COL + i
    return tl


def block_cols(tl, yb):       # columns of the 4 quarters of year YB
    return [p["col"] for p in tl if p["q"] and p["year"] == yb["year"]]


def prev_period(tl, i):
    """Previous period in the real sequence (quarters chain skipping YB; YS chains to YB/YS)."""
    p = tl[i]
    if p["kind"] in ("QH", "QP"):
        for j in range(i - 1, -1, -1):
            if tl[j]["q"]:
                return tl[j]
        return None
    for j in range(i - 1, -1, -1):
        if tl[j]["kind"] in ("YB", "YS"):
            return tl[j]
    return None


def yoy_col(tl, i):
    p = tl[i]
    for j in range(len(tl)):
        t = tl[j]
        if p["q"] and t["q"] == p["q"] and t["year"] == p["year"] - 1:
            return t["col"]
        if not p["q"] and not t["q"] and t["year"] == p["year"] - 1:
            return t["col"]
    return None


# --------------------------------------------------------------------- writing
def sheet_shell(wb, name, purpose, tl, tab_color="305496"):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    ws["B1"] = wb["Input Financials"]["B1"].value or "Model"
    ws["B1"].font = TITLE
    ws["B2"], ws["B2"].font = purpose, SUB
    ws["B3"] = "History = link to the input (green). Projection = formulas driven by the Premises tab."
    ws["B3"].font = SUB
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = f"{L(FIRST_COL)}{HDR + 1}"
    for p in tl:
        c = ws.cell(row=HDR, column=p["col"], value=p["label"])
        c.font, c.border = BOLD, THIN_BOT
        c.alignment = Alignment(horizontal="center")
        if not p["q"]:
            c.fill = GRAY_Y
        elif p["kind"] == "QP":
            c.fill = PROJ_FILL
        ws.column_dimensions[L(p["col"])].width = 9.5
    ws.cell(row=HDR, column=2, value="Line item").font = BOLD
    ws.cell(row=HDR, column=3, value="Unit").font = BOLD
    return ws


def section(ws, row, text, tl):
    ws.cell(row=row, column=2, value=text).font = WHITE_BOLD
    for col in range(2, FIRST_COL + len(tl)):
        ws.cell(row=row, column=col).fill = NAVY
    return row + 1


class Line:
    """One model line: history links the input; projection uses per-column-type functions."""

    def __init__(self, ws, row, label, unit, fmt=NUM_MN):
        self.ws, self.row = ws, row
        ws.cell(row=row, column=2, value=label).font = BLACK
        ws.cell(row=row, column=3, value=unit).font = SUB

    def fill(self, tl, hist=None, qp=None, yb="sum", ys=None):
        for i, p in enumerate(tl):
            cell = self.ws.cell(row=self.row, column=p["col"])
            cell.number_format = self.fmt if hasattr(self, "fmt") else NUM_MN
            if not p["q"]:
                cell.fill = GRAY_Y
            elif p["kind"] == "QP":
                cell.fill = PROJ_FILL
            r = self.row
            if p["kind"] == "QH":
                if hist:
                    cell.value, cell.font = hist(p), GREEN
            elif p["kind"] == "QP":
                if qp:
                    cell.value, cell.font = qp(p, i), BLACK
            elif p["kind"] == "YB":
                bc = block_cols(tl, p)
                if yb == "sum":
                    cell.value = f"=SUM({L(bc[0])}{r}:{L(bc[-1])}{r})"
                elif yb == "eop":
                    cell.value = f"={L(bc[-1])}{r}"
                elif yb == "avg":
                    cell.value = f'=IFERROR(AVERAGE({L(bc[0])}{r}:{L(bc[-1])}{r}),"-")'
                elif callable(yb):
                    cell.value = yb(p, i)
                cell.font = BLACK
            elif p["kind"] == "YS":
                if ys:
                    cell.value, cell.font = ys(p, i), BLACK
        return self


def add_line(ws, row, label, unit, fmt, tl, **kw):
    ln = Line(ws, row, label, unit)
    ln.fmt = fmt
    ln.fill(tl, **kw)
    return row + 1


# --------------------------------------------------------------------- assumptions
def seed_premises(wb, tl, fin, op, assets, seeds):
    proj = [p for p in tl if p["kind"] in ("QP", "YS")]
    ws = wb.create_sheet("Premises")
    ws.sheet_properties.tabColor = "BF8F00"
    ws["B1"], ws["B1"].font = "Premises — projection assumptions (blue cells)", TITLE
    ws["B2"] = ("Auto-seed = last historical value ('seed' method, to be replaced in the "
                "line-by-line session). Annual years use the annualized quarterly seed where it makes sense.")
    ws["B2"].font = SUB
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 10
    ws.freeze_panes = f"D{HDR + 1}"
    cols = {}
    for k, p in enumerate(proj):
        c = ws.cell(row=HDR, column=FIRST_COL + k, value=p["label"])
        c.font, c.fill = BOLD, PROJ_FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[L(FIRST_COL + k)].width = 9.5
        cols[p["label"]] = FIRST_COL + k
    ws.cell(row=HDR, column=2, value="Premise").font = BOLD
    ws.cell(row=HDR, column=3, value="Unit").font = BOLD

    rows = {}
    r = HDR + 2
    for key, label, unit, fmt, qval, mult_y in seeds:
        ws.cell(row=r, column=2, value=label).font = BLACK
        ws.cell(row=r, column=3, value=unit).font = SUB
        for p in proj:
            c = ws.cell(row=r, column=cols[p["label"]])
            c.value = round(qval * (mult_y if p["kind"] == "YS" else 1), 6)
            c.font, c.number_format = BLUE, fmt
        ws.cell(row=r, column=2).comment = Comment(
            "Auto-seeded from the last historical value. Replace in the assumption session "
            "(method, source and rationale in the Assumptions tab).", "engine")
        rows[key] = r
        r += 1
    return ws, rows, cols


def prem(ws_name, rows, cols):
    def ref(key, p):
        return f"='{ws_name}'!{L(cols[p['label']])}{rows[key]}"
    def raw(key, p):
        return f"'{ws_name}'!{L(cols[p['label']])}{rows[key]}"
    return ref, raw


# --------------------------------------------------------------------- valuation
def add_valuation_tab(wb, tl, hist_q, rows_is, rows_bs, rows_sc, rows_cf, shares: float | None = None):
    """DCF (FCFF) valuation tab. Reads the projected statements and builds: WACC (with Kd
    linked to the debt schedule by formula), explicit-period PV, Gordon terminal value,
    EV→equity bridge (leases treated as debt, IFRS-16), implied value per share, multiples,
    bull/base/bear scenarios, and a WACC×g sensitivity grid. It only *reads* the statements,
    so the balance check is unaffected. Analytical output — never a recommendation."""
    last_hist_year = q_seq(hist_q[-1])[0]
    last_hist_col = next(p["col"] for p in tl if p["label"] == hist_q[-1])
    fc = sorted((p["year"], p["col"]) for p in tl
                if p["q"] is None and p["year"] > last_hist_year)   # forecast annual columns
    N = len(fc)
    fc1 = fc[0][1]
    first_c, last_c = L(FIRST_COL), L(FIRST_COL + N - 1)
    DV = L(FIRST_COL)

    vs = wb.create_sheet("Valuation")
    vs.sheet_properties.tabColor = "548235"
    vs.column_dimensions["A"].width = 2
    vs.column_dimensions["B"].width = 42
    vs.column_dimensions["C"].width = 12
    for k in range(max(N, 5)):
        vs.column_dimensions[L(FIRST_COL + k)].width = 10
    vs["B1"] = (wb["Input Financials"]["B1"].value or "Model") + " — Valuation (DCF / FCFF)"
    vs["B1"].font = TITLE
    vs["B2"] = ("Analytical output, not an investment recommendation or a regulatory price target. "
                "Every blue cell is an assumption to be set with a source; method over numbers.")
    vs["B2"].font = SUB

    def lbl(r, text, font=BLACK, col=2):
        c = vs.cell(row=r, column=col, value=text); c.font = font; return c

    def inp(r, value, fmt=NUM_2D, col=FIRST_COL):
        c = vs.cell(row=r, column=col, value=value); c.font = BLUE; c.number_format = fmt; return c

    def fx(r, formula, fmt=NUM_2D, col=FIRST_COL, font=BLACK):
        c = vs.cell(row=r, column=col, value=formula); c.font = font; c.number_format = fmt; return c

    def navy(r, ncols):
        for cc in range(2, FIRST_COL + ncols):
            vs.cell(row=r, column=cc).fill = NAVY

    def D(r):
        return f"{DV}{r}"

    # --- WACC & assumptions -------------------------------------------------
    r = 4
    lbl(r, "WACC & VALUATION ASSUMPTIONS", WHITE_BOLD); navy(r, 1); r += 1
    rf = r;  lbl(r, "Risk-free rate (Rf)"); inp(r, 0.045, PCT)
    lbl(r, "seed — set in session (e.g. 10Y govt yield)", SUB, col=FIRST_COL + 1); r += 1
    erp = r; lbl(r, "Equity risk premium (ERP)"); inp(r, 0.055, PCT)
    lbl(r, "seed — Damodaran country ERP", SUB, col=FIRST_COL + 1); r += 1
    beta = r; lbl(r, "Beta (levered)"); inp(r, 1.00, NUM_2D)
    lbl(r, "seed — peer median, re-levered", SUB, col=FIRST_COL + 1); r += 1
    ke = r;  lbl(r, "(=) Cost of equity (Ke = Rf + β·ERP)"); fx(r, f"={D(rf)}+{D(beta)}*{D(erp)}", PCT); r += 1
    di, db = rows_sc["debt"]["interest"], rows_sc["debt"]["bop"]
    kd_pre = r; lbl(r, "Cost of debt pre-tax (Kd) — from debt schedule")
    fx(r, f"=IFERROR(-Schedules!{L(fc1)}{di}/Schedules!{L(fc1)}{db},0)", PCT, font=GREEN)
    lbl(r, "Kd = annual interest / opening debt", SUB, col=FIRST_COL + 1); r += 1
    tax = r; lbl(r, "Marginal tax rate"); inp(r, 0.25, PCT); r += 1
    kd_at = r; lbl(r, "(=) Cost of debt after tax"); fx(r, f"={D(kd_pre)}*(1-{D(tax)})", PCT); r += 1
    wd = r; lbl(r, "Weight of debt (Wd)"); inp(r, 0.30, PCT)
    lbl(r, "seed — net debt ÷ (net debt + equity)", SUB, col=FIRST_COL + 1); r += 1
    we = r; lbl(r, "(=) Weight of equity (We)"); fx(r, f"=1-{D(wd)}", PCT); r += 1
    wacc = r; lbl(r, "(=) WACC", BOLD); fx(r, f"={D(ke)}*{D(we)}+{D(kd_at)}*{D(wd)}", PCT, font=BOLD); r += 1
    g = r; lbl(r, "Terminal growth (g)"); inp(r, 0.02, PCT)
    lbl(r, "seed — ≤ long-run nominal GDP", SUB, col=FIRST_COL + 1); r += 1
    troic = r; lbl(r, "Terminal ROIC (stable phase)"); inp(r, 0.10, PCT)
    lbl(r, "seed — = WACC ⇒ no excess return; > WACC only if the moat persists", SUB, col=FIRST_COL + 1); r += 1

    # --- FCFF projection ----------------------------------------------------
    r += 1
    lbl(r, "DISCOUNTED CASH FLOW — FCFF (annual, currency mn)", WHITE_BOLD); navy(r, N); r += 1
    lbl(r, "Forecast year", BOLD)
    for k, (yy, _c) in enumerate(fc):
        c = vs.cell(row=r, column=FIRST_COL + k, value=yy); c.font = BOLD
        c.alignment = Alignment(horizontal="center")
    r += 1
    t_row = r; lbl(r, "Period (t)")
    for k in range(N):
        fx(r, k + 1, NUM_MN, col=FIRST_COL + k)
    r += 1

    def comp_row(label, cell_fn, font=GREEN, fmt=NUM_MN):
        nonlocal r
        rr = r; lbl(rr, label)
        for k, (yy, mc) in enumerate(fc):
            fx(rr, cell_fn(k, mc), fmt, col=FIRST_COL + k, font=font)
        r += 1
        return rr

    C = lambda k: L(FIRST_COL + k)
    ebit_r = comp_row("EBIT", lambda k, mc: f"=IS!{L(mc)}{rows_is['(=) EBIT']}")
    nopat_r = comp_row("NOPAT = EBIT·(1−t)", lambda k, mc: f"={C(k)}{ebit_r}*(1-${DV}${tax})", font=BLACK)
    dda_r = comp_row("(+) D&A", lambda k, mc: f"=IS!{L(mc)}{rows_is['memo: (+) DD&A']}")
    capex_r = comp_row("(−) Capex", lambda k, mc: f"=CF!{L(mc)}{rows_cf['(-) Capex']}")
    wc_r = comp_row("(+/−) Δ net working capital", lambda k, mc: f"=Schedules!{L(mc)}{rows_sc['wc']['chg']}")
    lease_r = comp_row("(−) Lease payments (principal + interest)",
                       lambda k, mc: (f"=CF!{L(mc)}{rows_cf['(-) Lease principal payments']}"
                                      f"+Schedules!{L(mc)}{rows_sc['lease']['interest']}"))
    fcff_r = comp_row("(=) FCFF",
                      lambda k, mc: (f"={C(k)}{nopat_r}+{C(k)}{dda_r}+{C(k)}{capex_r}"
                                     f"+{C(k)}{wc_r}+{C(k)}{lease_r}"), font=BOLD)
    disc_r = comp_row("Discount factor 1/(1+WACC)^t",
                      lambda k, mc: f"=1/(1+${DV}${wacc})^{C(k)}{t_row}", font=BLACK, fmt=NUM_2D)
    pv_r = comp_row("PV of FCFF", lambda k, mc: f"={C(k)}{fcff_r}*{C(k)}{disc_r}", font=BLACK)

    fcff_rng = f"${first_c}${fcff_r}:${last_c}${fcff_r}"
    t_rng = f"${first_c}${t_row}:${last_c}${t_row}"
    pv_rng = f"${first_c}${pv_r}:${last_c}${pv_r}"

    # --- value-creation check: ROIC vs. cost of capital ---------------------
    # Growth only creates value when ROIC > WACC; below it, growth destroys value (Damodaran,
    # and ~80% of firms earn below their cost of capital). Especially the headline test for
    # commodity/"bad business" sectors like O&G. Invested capital is read off the BS.
    r += 1
    lbl(r, "VALUE CREATION — ROIC vs. cost of capital (annual)", WHITE_BOLD); navy(r, N); r += 1
    begin_cols = [last_hist_col] + [c for (_y, c) in fc[:-1]]   # beginning-of-year capital base

    def ic_at(cc):
        return (f"(BS!{cc}{rows_bs['(=) Total equity']}"
                f"+BS!{cc}{rows_bs['Loans and financing (current)']}"
                f"+BS!{cc}{rows_bs['Loans and financing (non-current)']}"
                f"+BS!{cc}{rows_bs['Lease liabilities (current)']}"
                f"+BS!{cc}{rows_bs['Lease liabilities (non-current)']}"
                f"-BS!{cc}{rows_bs['Cash and equivalents']}"
                f"-BS!{cc}{rows_bs['Short-term investments']})")

    ic_r = comp_row("Invested capital — beginning of year",
                    lambda k, mc: f"={ic_at(L(begin_cols[k]))}", font=GREEN)
    roic_r = comp_row("ROIC = NOPAT ÷ invested capital",
                      lambda k, mc: f"=IFERROR({C(k)}{nopat_r}/{C(k)}{ic_r},0)", font=BLACK, fmt=PCT)
    comp_row("ROIC − WACC (value-creation spread)",
             lambda k, mc: f"={C(k)}{roic_r}-${DV}${wacc}", font=BOLD, fmt=PCT)
    lbl(r, "Invested capital = equity + debt + leases − cash. Positive spread ⇒ growth adds value.", SUB)
    r += 1

    # --- EV → equity bridge -------------------------------------------------
    r += 1
    lbl(r, "ENTERPRISE & EQUITY VALUE", WHITE_BOLD); navy(r, 1); r += 1
    sumpv = r; lbl(r, "Σ PV of explicit FCFF"); fx(r, f"=SUM({pv_rng})", NUM_MN); r += 1
    # Terminal value tied to a CONSISTENT reinvestment: in stable growth the firm must reinvest
    # g ÷ ROIC of its NOPAT to sustain g (Damodaran), so terminal FCFF is rebuilt from NOPAT
    # rather than carrying the last explicit year's (possibly abnormal) capex forward.
    reinv = r; lbl(r, "Terminal reinvestment rate (g ÷ ROIC)")
    fx(r, f"=IFERROR(${DV}${g}/${DV}${troic},0)", PCT); r += 1
    tnopat = r; lbl(r, "Terminal NOPAT (= year-N NOPAT × (1+g))")
    fx(r, f"=${last_c}${nopat_r}*(1+${DV}${g})", NUM_MN); r += 1
    tfcff = r; lbl(r, "Terminal FCFF (= NOPAT × (1 − reinv. rate))")
    fx(r, f"={D(tnopat)}*(1-{D(reinv)})", NUM_MN); r += 1
    tv = r; lbl(r, "Terminal value (Gordon, reinvestment-consistent)")
    fx(r, f"=IFERROR({D(tfcff)}/(${DV}${wacc}-${DV}${g}),0)", NUM_MN); r += 1
    pvtv = r; lbl(r, "PV of terminal value"); fx(r, f"={D(tv)}/(1+${DV}${wacc})^{last_c}{t_row}", NUM_MN); r += 1
    ev = r; lbl(r, "(=) Enterprise value (EV)", BOLD); fx(r, f"={D(sumpv)}+{D(pvtv)}", NUM_MN, font=BOLD); r += 1
    lhc = L(last_hist_col)
    nd = r; lbl(r, "(−) Net debt (last actual BS)")
    fx(r, (f"=-(BS!{lhc}{rows_bs['Loans and financing (current)']}"
           f"+BS!{lhc}{rows_bs['Loans and financing (non-current)']}"
           f"-BS!{lhc}{rows_bs['Cash and equivalents']}-BS!{lhc}{rows_bs['Short-term investments']})"),
       NUM_MN, font=GREEN); r += 1
    lz = r; lbl(r, "(−) Lease liabilities (debt-like)")
    fx(r, (f"=-(BS!{lhc}{rows_bs['Lease liabilities (current)']}"
           f"+BS!{lhc}{rows_bs['Lease liabilities (non-current)']})"), NUM_MN, font=GREEN); r += 1
    mi = r; lbl(r, "(−) Minority interest")
    fx(r, f"=-BS!{lhc}{rows_bs['Minority interest (equity)']}", NUM_MN, font=GREEN); r += 1
    eqv = r; lbl(r, "(=) Equity value", BOLD); fx(r, f"={D(ev)}+{D(nd)}+{D(lz)}+{D(mi)}", NUM_MN, font=BOLD); r += 1
    sh = r; lbl(r, "Shares outstanding (mn)")
    if shares and shares > 0:
        inp(r, float(shares), NUM_MN)
        lbl(r, "from input — memo: Shares outstanding (EOP), last actual", SUB, col=FIRST_COL + 1)
    else:
        inp(r, 1000.0, NUM_MN)
        lbl(r, "seed — shares line absent/empty in input; set from the filing", SUB, col=FIRST_COL + 1)
    r += 1
    tp = r; lbl(r, "(=) Implied value per share", BOLD); fx(r, f"=IFERROR({D(eqv)}/{D(sh)},0)", NUM_2D, font=BOLD); r += 1
    lbl(r, "Analytical output — model-implied value, NOT a price target or investment "
           "recommendation; an estimate (see method & sources).", SUB); r += 1

    # --- multiples ----------------------------------------------------------
    r += 1
    lbl(r, "IMPLIED MULTIPLES (year 1)", WHITE_BOLD); navy(r, 1); r += 1
    lbl(r, "EV / EBITDA"); fx(r, f"=IFERROR({D(ev)}/IS!{L(fc1)}{rows_is['memo: (=) EBITDA']},0)", NUM_2D); r += 1
    lbl(r, "Implied P/E"); fx(r, f"=IFERROR({D(eqv)}/IS!{L(fc1)}{rows_is['(=) Net income']},0)", NUM_2D); r += 1

    def tp_formula(wref, gref):
        # Terminal FCFF rebuilt from NOPAT with the consistent reinvestment rate (gref ÷ ROIC),
        # matching the base-case terminal value above as WACC/g vary across the grid.
        tfcff_expr = f"${last_c}${nopat_r}*(1+{gref})*(1-IFERROR({gref}/${DV}${troic},0))"
        ev_expr = (f"SUMPRODUCT({fcff_rng},1/(1+{wref})^{t_rng})"
                   f"+IFERROR(({tfcff_expr})/({wref}-{gref}),0)/(1+{wref})^{last_c}{t_row}")
        return f"=IFERROR(({ev_expr}+{D(nd)}+{D(lz)}+{D(mi)})/{D(sh)},0)"

    # --- scenarios ----------------------------------------------------------
    r += 1
    lbl(r, "SCENARIOS (bull / base / bear)", WHITE_BOLD); navy(r, 3); r += 1
    for k, name in enumerate(["Bull", "Base", "Bear"]):
        c = vs.cell(row=r, column=FIRST_COL + k, value=name); c.font = BOLD
        c.alignment = Alignment(horizontal="center")
    r += 1
    sc_wacc = r; lbl(r, "WACC")
    fx(r, f"=${DV}${wacc}-0.01", PCT, col=FIRST_COL + 0)
    fx(r, f"=${DV}${wacc}", PCT, col=FIRST_COL + 1)
    fx(r, f"=${DV}${wacc}+0.01", PCT, col=FIRST_COL + 2)
    r += 1
    sc_g = r; lbl(r, "Terminal growth (g)")
    fx(r, f"=${DV}${g}+0.005", PCT, col=FIRST_COL + 0)
    fx(r, f"=${DV}${g}", PCT, col=FIRST_COL + 1)
    fx(r, f"=${DV}${g}-0.005", PCT, col=FIRST_COL + 2)
    r += 1
    lbl(r, "(=) Implied value per share", BOLD)
    for k in range(3):
        fx(r, tp_formula(f"{C(k)}{sc_wacc}", f"{C(k)}{sc_g}"), NUM_2D, col=FIRST_COL + k, font=BOLD)
    r += 1

    # --- sensitivity grid ---------------------------------------------------
    r += 1
    lbl(r, "SENSITIVITY — value/share: WACC (rows) × g (cols)", WHITE_BOLD); navy(r, 6); r += 1
    g_hdr = r; lbl(r, "WACC \\ g", BOLD)
    for j, gv in enumerate([0.010, 0.015, 0.020, 0.025, 0.030]):
        c = vs.cell(row=r, column=FIRST_COL + j, value=gv); c.font = BLUE; c.number_format = PCT
        c.alignment = Alignment(horizontal="center")
    r += 1
    for wo in (-0.02, -0.01, 0.0, 0.01, 0.02):
        c = vs.cell(row=r, column=2, value=f"=${DV}${wacc}+{wo}"); c.font = BLACK; c.number_format = PCT
        for j in range(5):
            fx(r, tp_formula(f"$B{r}", f"{C(j)}${g_hdr}"), NUM_2D, col=FIRST_COL + j, font=BLACK)
        r += 1

    r += 1
    lbl(r, ("FCFF = EBIT·(1−tax) + D&A − Capex − ΔNWC − lease payments. Leases are treated as debt in the "
            "bridge (IFRS-16) and Kd is read from the debt schedule by formula. All blue cells are "
            "assumptions to set with sources in the assumption session."), SUB)
    return vs


# --------------------------------------------------------------------- main
def main(input_path, out_path, sector: str | None = None):
    diag = BuildDiagnostics()
    diag.log_header(input_path, out_path)

    shutil.copy(input_path, out_path)
    wb = load_workbook(out_path)
    fin, op = wb["Input Financials"], wb["Input Operational"]
    fhdr, ohdr = header_map(fin), header_map(op)
    frow, orow = label_rows(fin), label_rows(op)

    hist_q = detect_history(fin)
    tl = build_timeline(hist_q)
    last = hist_q[-1]
    lc_f, lc_o = fhdr[last], ohdr[last]

    # --- Read & understand the input (flow stages 2-4) -----------------------
    # No per-sector template: identify company/sector, introspect the operational
    # tab, locate the method card. The operational layer is assembled from this.
    an = analyze_input(fin, op, orow, ohdr, hist_q, sector=sector)
    diag.log_analysis(an)

    if not an.sector:
        print("WARNING: could not identify the sector — building a generic top-down operational layer")
    elif not an.card_exists:
        print(f"WARNING: no method card found for sector '{an.sector}' in knowledge/sector_modeling_rules/sectors/")

    # The O&G engine is the only bottom-up (volume × price) operational build today.
    # Every other sector — or O&G without per-asset disclosure — gets the generic
    # top-down operational layer assembled from the introspected drivers (Stage 2).
    # Stage 3 will let the method card configure a sector-specific build here.
    og = (an.sector == "oil_and_gas") and bool(an.assets)

    # Convert AssetInfo dataclasses to the dict format the rest of the code expects
    assets = [
        dict(idx=a.idx, name=a.name,
             prod_row=a.prod_row,
             lift_row=a.lift_row,
             capex_row=a.capex_row)
        for a in an.assets
    ]

    # Helper: safely get a value from the operational sheet by label (0 if absent)
    def safe_o(label: str):
        row = an.op_row(label)
        return g(op, row, lc_o) if row else 0

    g = lambda ws, rr, cc: ws.cell(row=rr, column=cc).value or 0

    def gf(label, default=0.0):
        """Read a financial line by label (0/default if the line is absent)."""
        row = frow.get(label)
        return g(fin, row, lc_f) if row else default

    if og:
        # --- Oil & Gas bottom-up operational seeds (volume × price build) -------
        total_prod = sum(g(op, a["prod_row"], lc_o) for a in assets)
        lift_w = (sum(g(op, a["prod_row"], lc_o) * g(op, a["lift_row"], lc_o) for a in assets
                      if a["lift_row"])
                  / total_prod if total_prod else 0)
        rev_l = g(fin, frow["(=) Net revenue"], lc_f)
        vol_l = safe_o("Sales volume")
        seeds = [("brent", "Brent average", "USD/bbl", NUM_1D, safe_o("Brent average"), 1),
                 ("disc", "Realized premium (discount) to Brent", "USD/bbl", NUM_1D,
                  safe_o("Realized price") - safe_o("Brent average"), 1),
                 ("lift", "Lifting cost (weighted avg)", "USD/boe", NUM_1D, lift_w, 1),
                 ("roy", "Royalties (% of revenue)", "%", PCT,
                  (safe_o("Royalties and government take") / rev_l if rev_l else 0.1), 1),
                 ("depl", "Depletion rate", "USD/boe", NUM_1D,
                  (g(fin, frow["memo: (+) Depreciation, depletion & amortization"], lc_f) / vol_l * 1000)
                  if vol_l else 8.0, 1),
                 ("ga", "G&A (% of revenue)", "%", PCT,
                  -g(fin, frow["(-) General & administrative expenses"], lc_f) / rev_l if rev_l else 0.04, 1),
                 ("expl", "Exploration expenses (per period)", "USD mn", NUM_MN,
                  g(fin, frow["(-) Exploration expenses"], lc_f), 4),
                 ("tax", "Effective tax rate (% of EBT)", "%", PCT,
                  -g(fin, frow["(-) Income taxes"], lc_f) / g(fin, frow["(=) EBT"], lc_f)
                  if g(fin, frow["(=) EBT"], lc_f) else 0.25, 1),
                 ("payout", "Dividend payout (% of NI)", "%", PCT, 0.0, 1),
                 ("days", "Days in period", "days", NUM_1D, 91.25, 4)]
        for a in assets:
            seeds.insert(0, (f"prod{a['idx']}", f"Production — {a['name']}", "kboe/d", NUM_1D,
                             g(op, a["prod_row"], lc_o), 1))
            seeds.append((f"capex{a['idx']}", f"Capex — {a['name']} (per period)", "USD mn", NUM_MN,
                          g(op, a["capex_row"], lc_o), 4))
    else:
        # --- Generic top-down operational seeds (revenue growth × margin) -------
        # Revenue is projected from its own history with a growth premise; cost is a
        # margin premise; D&A and capex are levels seeded from the financials. The
        # disclosed operational drivers are surfaced (Stage 2) but do not yet drive
        # the build — Stage 3 wires driver → financial via the sector method card.
        rev_l = gf("(=) Net revenue")
        gp_l = gf("(=) Gross profit")
        dda_l = gf("memo: (+) Depreciation, depletion & amortization")
        ebt_l = gf("(=) EBT")
        # Margin is measured EX-D&A so that (COGS ex-D&A) + D&A reconstructs the
        # reported COGS exactly: COGS_total = -(rev × (1 - margin)) - D&A.
        seeds = [
            ("rev_growth", "Revenue growth (per period)", "%", PCT, 0.0, 1),
            ("gross_margin", "Gross margin ex-D&A (% of revenue)", "%", PCT,
             ((gp_l + dda_l) / rev_l if rev_l else 0.30), 1),
            ("depr", "Depreciation & amortization (per period)", "USD mn", NUM_MN,
             gf("memo: (+) Depreciation, depletion & amortization"), 4),
            # Damodaran sales-to-capital: capex is sized off revenue growth, not guessed as a
            # flat level. Reinvestment = Δrevenue ÷ this ratio (Capex across the Life Cycle).
            ("sales_to_capital", "Sales-to-capital ratio (Δrevenue ÷ net capex)", "x", NUM_2D, 2.0, 1),
            ("ga", "G&A (% of revenue)", "%", PCT,
             (-gf("(-) General & administrative expenses") / rev_l if rev_l else 0.04), 1),
            ("expl", "Exploration expenses (per period)", "USD mn", NUM_MN, 0.0, 4),
            ("tax", "Effective tax rate (% of EBT)", "%", PCT,
             (-gf("(-) Income taxes") / ebt_l if ebt_l else 0.25), 1),
            ("payout", "Dividend payout (% of NI)", "%", PCT, 0.0, 1),
            ("days", "Days in period", "days", NUM_1D, 91.25, 4),
        ]

    # working-capital days, back-solved from the last historical balances
    arl = g(fin, frow["Trade receivables"], lc_f)
    invl = g(fin, frow["Inventories"], lc_f)
    apl = g(fin, frow["Trade payables"], lc_f)
    ocal = g(fin, frow["Other current assets"], lc_f)
    ocll = g(fin, frow["Other current liabilities"], lc_f)
    cogs_base = -g(fin, frow["(-) Cost of goods sold"], lc_f)
    dseed = lambda bal, base: round(bal / base * 91.25, 1) if base else 30.0
    seeds += [
        ("dso", "Days sales outstanding (DSO)", "days", NUM_1D, dseed(arl, rev_l), 1),
        ("invdays", "Inventory days", "days", NUM_1D, dseed(invl, cogs_base), 1),
        ("ocadays", "Other current assets (days of revenue)", "days", NUM_1D, dseed(ocal, rev_l), 1),
        ("dpo", "Days payable outstanding (DPO)", "days", NUM_1D, dseed(apl, cogs_base), 1),
        ("ocldays", "Other current liabilities (days of revenue)", "days", NUM_1D, dseed(ocll, rev_l), 1),
    ]

    # debt schedule seeds (interest on opening balance; flat debt by default)
    loans_c = g(fin, frow["Loans and financing (current)"], lc_f)
    loans_nc = g(fin, frow["Loans and financing (non-current)"], lc_f)
    debt_total = loans_c + loans_nc
    finres_q = g(fin, frow["(+/-) Financial result"], lc_f)
    debt_rate_seed = round(max(0.0, -finres_q / debt_total), 6) if debt_total else 0.0
    other_finres_seed = round(finres_q + debt_rate_seed * debt_total, 6)
    curr_pct_seed = round(loans_c / debt_total, 4) if debt_total else 0.2
    seeds += [
        ("debt_draw", "Debt drawdowns (per period)", "USD mn", NUM_MN, 0.0, 4),
        ("debt_repay", "Debt repayments (per period)", "USD mn", NUM_MN, 0.0, 4),
        ("debt_rate", "Cost of debt (per period)", "%", PCT, debt_rate_seed, 4),
        ("debt_curr_pct", "Debt classified as current (%)", "%", PCT, curr_pct_seed, 1),
        ("other_finres", "Other financial result (per period)", "USD mn", NUM_MN, other_finres_seed, 4),
        ("min_cash", "Minimum cash (revolver floor)", "USD mn", NUM_MN, 0.0, 1),
    ]

    # lease (IFRS-16) seeds — neutral by default; activated in the assumption session
    lease_c = g(fin, frow["Lease liabilities (current)"], lc_f)
    lease_nc = g(fin, frow["Lease liabilities (non-current)"], lc_f)
    lease_total = lease_c + lease_nc
    lease_curr_seed = round(lease_c / lease_total, 4) if lease_total else 0.2
    seeds += [
        ("lease_add", "New leases / RoU additions (per period)", "USD mn", NUM_MN, 0.0, 4),
        ("lease_depr_amt", "Lease depreciation (per period)", "USD mn", NUM_MN, 0.0, 4),
        ("lease_princ_amt", "Lease principal repayment (per period)", "USD mn", NUM_MN, 0.0, 4),
        ("lease_rate", "Lease discount rate (per period)", "%", PCT, 0.0, 4),
        ("lease_curr_pct", "Lease classified as current (%)", "%", PCT, lease_curr_seed, 1),
    ]

    # ARO / decommissioning seeds (O&G) — neutral by default
    seeds += [
        ("aro_accr", "ARO accretion (discount unwind, per period)", "USD mn", NUM_MN, 0.0, 4),
        ("aro_settle", "ARO settlements / decommissioning spend (per period)", "USD mn", NUM_MN, 0.0, 4),
    ]

    pws, prows, pcols = seed_premises(wb, tl, fin, op, assets, seeds)
    P, Praw = prem("Premises", prows, pcols)

    # ---------------------------------------------------------- Operational Model
    om = sheet_shell(wb, "Operational Model", "Operational drivers → financial lines", tl)
    OM = "Operational Model"
    if og:
        # ===== Oil & Gas bottom-up build (volume × price) =====================
        r = HDR + 2
        r = section(om, r, "PRODUCTION (kboe/d)", tl)
        om_prod = {}
        for a in assets:
            om_prod[a["idx"]] = r
            r = add_line(om, r, a["name"], "kboe/d", NUM_1D, tl,
                         hist=lambda p, a=a: f"='Input Operational'!{L(ohdr[p['label']])}{a['prod_row']}",
                         qp=lambda p, i, a=a: P(f"prod{a['idx']}", p),
                         yb="avg", ys=lambda p, i, a=a: P(f"prod{a['idx']}", p))
        row_tprod = r
        r = add_line(om, r, "(=) Total production", "kboe/d", NUM_1D, tl,
                     hist=lambda p: f"=SUM({L(p['col'])}{min(om_prod.values())}:{L(p['col'])}{max(om_prod.values())})",
                     qp=lambda p, i: f"=SUM({L(p['col'])}{min(om_prod.values())}:{L(p['col'])}{max(om_prod.values())})",
                     yb="avg",
                     ys=lambda p, i: f"=SUM({L(p['col'])}{min(om_prod.values())}:{L(p['col'])}{max(om_prod.values())})")
        r += 1
        # Row references from input introspection (None → no history link for that input)
        _svol_row  = an.op_row("Sales volume")
        _brent_row = an.op_row("Brent average")
        _real_row  = an.op_row("Realized price")

        r = section(om, r, "VOLUMES AND PRICES", tl)
        row_vol = r
        r = add_line(om, r, "Sales volume", "kboe", NUM_MN, tl,
                     hist=(lambda p: f"='Input Operational'!{L(ohdr[p['label']])}{_svol_row}")
                          if _svol_row else None,
                     qp=lambda p, i: f"={L(p['col'])}{row_tprod}*{Praw('days', p)}",
                     yb="sum",
                     ys=lambda p, i: f"={L(p['col'])}{row_tprod}*{Praw('days', p)}")
        row_brent = r
        r = add_line(om, r, "Brent average", "USD/bbl", NUM_1D, tl,
                     hist=(lambda p: f"='Input Operational'!{L(ohdr[p['label']])}{_brent_row}")
                          if _brent_row else None,
                     qp=lambda p, i: P("brent", p), yb="avg", ys=lambda p, i: P("brent", p))
        row_real = r
        r = add_line(om, r, "Realized price", "USD/boe", NUM_1D, tl,
                     hist=(lambda p: f"='Input Operational'!{L(ohdr[p['label']])}{_real_row}")
                          if _real_row else None,
                     qp=lambda p, i: f"={L(p['col'])}{row_brent}+{Praw('disc', p)}",
                     yb="avg",
                     ys=lambda p, i: f"={L(p['col'])}{row_brent}+{Praw('disc', p)}")
        r += 1
        r = section(om, r, "REVENUE AND COST BUILD (USD mn)", tl)
        row_rev = r
        rev_f = lambda p, i: f"={L(p['col'])}{row_vol}*{L(p['col'])}{row_real}/1000"
        r = add_line(om, r, "(=) Revenue build (volume × price)", "USD mn", NUM_MN, tl,
                     hist=lambda p: f"='Input Financials'!{L(fhdr[p['label']])}{frow['(=) Net revenue']}",
                     qp=rev_f, yb="sum", ys=rev_f)
        row_liftc = r
        lift_f = lambda p, i: f"=-{L(p['col'])}{row_vol}*{Praw('lift', p)}/1000"
        r = add_line(om, r, "(-) Lifting cost", "USD mn", NUM_MN, tl, qp=lift_f, yb="sum", ys=lift_f)
        row_royc = r
        roy_f = lambda p, i: f"=-{L(p['col'])}{row_rev}*{Praw('roy', p)}"
        r = add_line(om, r, "(-) Royalties", "USD mn", NUM_MN, tl, qp=roy_f, yb="sum", ys=roy_f)
        row_depl = r
        depl_f = lambda p, i: f"=-{L(p['col'])}{row_vol}*{Praw('depl', p)}/1000"
        r = add_line(om, r, "(-) Depletion (DD&A)", "USD mn", NUM_MN, tl,
                     hist=lambda p: f"=-'Input Financials'!{L(fhdr[p['label']])}{frow['memo: (+) Depreciation, depletion & amortization']}",
                     qp=depl_f, yb="sum", ys=depl_f)
        row_capext = r
        cap_f = lambda p, i: "=" + "+".join(Praw(f"capex{a['idx']}", p) for a in assets)
        # Total capex row in Input Operational is max_slots+1 rows below the section header.
        _capex_hdr  = orow.get("CAPEX BY ASSET")
        _capex_tot  = (_capex_hdr + an.max_slots + 1) if _capex_hdr else None
        r = add_line(om, r, "(=) Total capex", "USD mn", NUM_MN, tl,
                     hist=(lambda p: f"='Input Operational'!{L(ohdr[p['label']])}{_capex_tot}")
                          if _capex_tot else None,
                     qp=cap_f, yb="sum", ys=cap_f)
        row_cogs = None   # O&G COGS = lifting + royalties + depletion (three rows)
    else:
        # ===== Generic top-down build (revenue growth × margin) ===============
        # Stage 2: assemble the operational layer from whatever the company
        # discloses. Disclosed drivers are surfaced as linked, flat-projected
        # memo lines; revenue/cost are projected top-down from premises. They are
        # not yet wired to revenue — Stage 3 lets the method card do that.
        r = HDR + 2
        r = section(om, r, "OPERATIONAL DRIVERS (as disclosed in the input)", tl)
        disclosed = [(lbl, rr) for lbl, rr in sorted(an.op_labels.items(), key=lambda kv: kv[1])
                     if lbl not in an.op_sections]
        if not disclosed:
            r = add_line(om, r, "(no operational drivers disclosed)", "", NUM_1D, tl)
        for lbl, _rr in disclosed:
            this = r
            r = add_line(om, r, lbl, "", NUM_1D, tl,
                         hist=(lambda p, lbl=lbl: f"='Input Operational'!{L(ohdr[p['label']])}{an.op_row(lbl)}")
                              if an.op_row(lbl) else None,
                         qp=lambda p, i, this=this: f"={L(prev_period(tl, i)['col'])}{this}",
                         yb="avg",
                         ys=lambda p, i, this=this: f"={L(prev_period(tl, i)['col'])}{this}")
        r += 1
        r = section(om, r, "REVENUE AND COST BUILD (top-down, USD mn)", tl)
        row_rev = r
        rev_f = lambda p, i: f"={L(prev_period(tl, i)['col'])}{row_rev}*(1+{Praw('rev_growth', p)})"
        r = add_line(om, r, "(=) Revenue", "USD mn", NUM_MN, tl,
                     hist=lambda p: f"='Input Financials'!{L(fhdr[p['label']])}{frow['(=) Net revenue']}",
                     qp=rev_f, yb="sum", ys=rev_f)
        row_cogs = r
        cogs_g = lambda p, i: f"=-{L(p['col'])}{row_rev}*(1-{Praw('gross_margin', p)})"
        r = add_line(om, r, "(-) Cost of goods sold (ex-D&A)", "USD mn", NUM_MN, tl,
                     qp=cogs_g, yb="sum", ys=cogs_g)
        row_depl = r
        depl_g = lambda p, i: f"=-{Praw('depr', p)}"
        r = add_line(om, r, "(-) Depreciation & amortization", "USD mn", NUM_MN, tl,
                     hist=lambda p: f"=-'Input Financials'!{L(fhdr[p['label']])}{frow['memo: (+) Depreciation, depletion & amortization']}",
                     qp=depl_g, yb="sum", ys=depl_g)
        row_capext = r
        # Capex = maintenance (≈ D&A) + growth (Δrevenue ÷ sales-to-capital). PP&E (and so
        # invested capital) then grows in step with revenue, and the valuation's FCFF
        # reinvestment is internally consistent with the revenue path (Damodaran).
        cap_g = lambda p, i: (f"={Praw('depr', p)}+({L(p['col'])}{row_rev}"
                              f"-{L(prev_period(tl, i)['col'])}{row_rev})/{Praw('sales_to_capital', p)}")
        r = add_line(om, r, "(=) Total capex", "USD mn", NUM_MN, tl,
                     hist=(lambda p: f"=-'Input Financials'!{L(fhdr[p['label']])}{frow['(-) Capex']}")
                          if "(-) Capex" in frow else None,
                     qp=cap_g, yb="sum", ys=cap_g)
        row_liftc = row_royc = None

    # ---------------------------------------------------------- Schedules
    sc = sheet_shell(wb, "Schedules", "Roll-forwards: PP&E, retained earnings, cash", tl)
    r = HDR + 2
    rows_sc = {}
    def corkscrew(r, name, bop_seed, plus, minus, fmt=NUM_MN):
        r = section(sc, r, name, tl)
        rb, rp, rm, re_ = r, r + 1, r + 2, r + 3
        def bop(p, i):
            pv = prev_period(tl, i)
            if pv is None or (pv["kind"] == "QH"):
                return bop_seed(p)
            if p["kind"] == "YB":
                bc = block_cols(tl, p)
                return f"={L(bc[0])}{rb}"
            return f"={L(pv['col'])}{re_}"
        add_line(sc, rb, "BOP", "USD mn", fmt, tl,
                 hist=lambda p: None, qp=bop, yb=bop, ys=bop)
        add_line(sc, rp, "(+) additions", "USD mn", fmt, tl, qp=plus, yb="sum", ys=plus)
        add_line(sc, rm, "(-) reductions", "USD mn", fmt, tl, qp=minus, yb="sum", ys=minus)
        eop = lambda p, i: f"=SUM({L(p['col'])}{rb}:{L(p['col'])}{rm})"
        def eop_yb(p, i):
            bc = block_cols(tl, p)
            return f"={L(bc[-1])}{re_}"
        add_line(sc, re_, "(=) EOP", "USD mn", fmt, tl, qp=eop, yb=eop_yb, ys=eop)
        return re_ + 2, dict(bop=rb, plus=rp, minus=rm, eop=re_)

    seed_ppe = lambda p: f"='Input Financials'!{L(fhdr[last])}{frow['PP&E']}"
    r, rows_sc["ppe"] = corkscrew(
        r, "PP&E ROLL-FORWARD", seed_ppe,
        plus=lambda p, i: f"='{OM}'!{L(p['col'])}{row_capext}",
        minus=lambda p, i: f"='{OM}'!{L(p['col'])}{row_depl}")

    # debt roll-forward + interest on the opening balance (no circularity)
    seed_debt = lambda p: (f"='Input Financials'!{L(fhdr[last])}{frow['Loans and financing (current)']}"
                           f"+'Input Financials'!{L(fhdr[last])}{frow['Loans and financing (non-current)']}")
    r, rows_sc["debt"] = corkscrew(
        r, "DEBT ROLL-FORWARD", seed_debt,
        plus=lambda p, i: P("debt_draw", p),
        minus=lambda p, i: f"=-{Praw('debt_repay', p)}")
    debt_bop = rows_sc["debt"]["bop"]
    rows_sc["debt"]["interest"] = r
    int_f = lambda p, i: f"=-{Praw('debt_rate', p)}*{L(p['col'])}{debt_bop}"
    r = add_line(sc, r, "(-) Debt interest (on opening balance)", "USD mn", NUM_1D, tl,
                 qp=int_f, yb="sum", ys=int_f)
    r += 1

    # lease IFRS-16: twin roll-forwards (RoU asset + lease liability) + interest on opening balance
    seed_rou = lambda p: f"='Input Financials'!{L(fhdr[last])}{frow['Right-of-use assets']}"
    r, rows_sc["rou"] = corkscrew(
        r, "RIGHT-OF-USE ASSET (IFRS-16)", seed_rou,
        plus=lambda p, i: P("lease_add", p),
        minus=lambda p, i: f"=-{Praw('lease_depr_amt', p)}")
    seed_ll = lambda p: (f"='Input Financials'!{L(fhdr[last])}{frow['Lease liabilities (current)']}"
                         f"+'Input Financials'!{L(fhdr[last])}{frow['Lease liabilities (non-current)']}")
    r, rows_sc["lease"] = corkscrew(
        r, "LEASE LIABILITY (IFRS-16)", seed_ll,
        plus=lambda p, i: P("lease_add", p),
        minus=lambda p, i: f"=-{Praw('lease_princ_amt', p)}")
    lease_bop = rows_sc["lease"]["bop"]
    rows_sc["lease"]["interest"] = r
    lint_f = lambda p, i: f"=-{Praw('lease_rate', p)}*{L(p['col'])}{lease_bop}"
    r = add_line(sc, r, "(-) Lease interest (on opening balance)", "USD mn", NUM_1D, tl,
                 qp=lint_f, yb="sum", ys=lint_f)
    r += 1

    # ARO / decommissioning roll-forward (accretion is a premise; non-cash, added back in CFO)
    seed_aro = lambda p: f"='Input Financials'!{L(fhdr[last])}{frow['Provisions (incl. asset retirement obligation)']}"
    r, rows_sc["aro"] = corkscrew(
        r, "ASSET RETIREMENT OBLIGATION (decommissioning)", seed_aro,
        plus=lambda p, i: P("aro_accr", p),
        minus=lambda p, i: f"=-{Praw('aro_settle', p)}")

    seed_re = lambda p: f"='Input Financials'!{L(fhdr[last])}{frow['Retained earnings (accumulated)']}"
    # IS is built before the RE/cash corkscrews because RE needs NI (IS only needs OM).

    # ---------------------------------------------------------- IS
    is_ = sheet_shell(wb, "IS", "Integrated income statement", tl)
    ri = HDR + 2
    ri = section(is_, ri, "INCOME STATEMENT", tl)
    rows_is = {}
    def is_line(label, unit, fmt, **kw):
        nonlocal ri
        rows_is[label] = ri
        ri = add_line(is_, ri, label, unit, fmt, tl, **kw)
    hist_fin = lambda lbl: (lambda p: f"='Input Financials'!{L(fhdr[p['label']])}{frow[lbl]}")
    is_line("(=) Net revenue", "USD mn", NUM_MN, hist=hist_fin("(=) Net revenue"),
            qp=lambda p, i: f"='{OM}'!{L(p['col'])}{row_rev}", yb="sum",
            ys=lambda p, i: f"='{OM}'!{L(p['col'])}{row_rev}")
    ryoy = ri
    def yoy_f(target_row):
        def f(p, i):
            pc = yoy_col(tl, i)
            return (f'=IFERROR({L(p["col"])}{target_row}/{L(pc)}{target_row}-1,"-")' if pc else None)
        return f
    ri = add_line(is_, ri, "% YoY", "%", PCT, tl,
                  hist=lambda p: yoy_f(rows_is["(=) Net revenue"])(p, [t["label"] for t in tl].index(p["label"])),
                  qp=yoy_f(rows_is["(=) Net revenue"]), yb=yoy_f(rows_is["(=) Net revenue"]),
                  ys=yoy_f(rows_is["(=) Net revenue"]))
    if og:
        cogs_f = lambda p, i: (f"='{OM}'!{L(p['col'])}{row_liftc}+'{OM}'!{L(p['col'])}{row_royc}"
                               f"+'{OM}'!{L(p['col'])}{row_depl}")
    else:
        # Generic COGS = cost-of-goods (ex-D&A) + D&A, both off the Operational Model.
        cogs_f = lambda p, i: (f"='{OM}'!{L(p['col'])}{row_cogs}+'{OM}'!{L(p['col'])}{row_depl}")
    is_line("(-) Cost of goods sold", "USD mn", NUM_MN, hist=hist_fin("(-) Cost of goods sold"),
            qp=cogs_f, yb="sum", ys=cogs_f)
    gp_f = lambda p, i: f"={L(p['col'])}{rows_is['(=) Net revenue']}+{L(p['col'])}{rows_is['(-) Cost of goods sold']}"
    is_line("(=) Gross profit", "USD mn", NUM_MN, hist=hist_fin("(=) Gross profit"),
            qp=gp_f, yb="sum", ys=gp_f)
    ga_f = lambda p, i: f"=-{L(p['col'])}{rows_is['(=) Net revenue']}*{Praw('ga', p)}"
    is_line("(-) General & administrative expenses", "USD mn", NUM_MN,
            hist=hist_fin("(-) General & administrative expenses"), qp=ga_f, yb="sum", ys=ga_f)
    is_line("(-) Exploration expenses", "USD mn", NUM_MN, hist=hist_fin("(-) Exploration expenses"),
            qp=lambda p, i: P("expl", p), yb="sum", ys=lambda p, i: P("expl", p))
    other_op_f = lambda p, i: f"=-{Praw('lease_depr_amt', p)}"
    is_line("(+/-) Other operating income (expenses)", "USD mn", NUM_MN,
            hist=hist_fin("(+/-) Other operating income (expenses)"),
            qp=other_op_f, yb="sum", ys=other_op_f)
    ebit_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_is['(=) Gross profit']},"
                           f"{L(p['col'])}{rows_is['(-) General & administrative expenses']},"
                           f"{L(p['col'])}{rows_is['(-) Exploration expenses']},"
                           f"{L(p['col'])}{rows_is['(+/-) Other operating income (expenses)']})")
    is_line("(=) EBIT", "USD mn", NUM_MN, hist=hist_fin("(=) EBIT"), qp=ebit_f, yb="sum", ys=ebit_f)
    finres_f = lambda p, i: (f"=Schedules!{L(p['col'])}{rows_sc['debt']['interest']}"
                             f"+Schedules!{L(p['col'])}{rows_sc['lease']['interest']}"
                             f"-{Praw('aro_accr', p)}+{Praw('other_finres', p)}")
    is_line("(+/-) Financial result", "USD mn", NUM_MN, hist=hist_fin("(+/-) Financial result"),
            qp=finres_f, yb="sum", ys=finres_f)
    ebt_f = lambda p, i: (f"={L(p['col'])}{rows_is['(=) EBIT']}+{L(p['col'])}{rows_is['(+/-) Financial result']}")
    is_line("(=) EBT", "USD mn", NUM_MN, hist=hist_fin("(=) EBT"), qp=ebt_f, yb="sum", ys=ebt_f)
    tax_f = lambda p, i: f"=-MAX({L(p['col'])}{rows_is['(=) EBT']},0)*{Praw('tax', p)}"
    is_line("(-) Income taxes", "USD mn", NUM_MN, hist=hist_fin("(-) Income taxes"),
            qp=tax_f, yb="sum", ys=tax_f)
    ni_f = lambda p, i: f"={L(p['col'])}{rows_is['(=) EBT']}+{L(p['col'])}{rows_is['(-) Income taxes']}"
    is_line("(=) Net income", "USD mn", NUM_MN, hist=hist_fin("(=) Net income"),
            qp=ni_f, yb="sum", ys=ni_f)
    da_f = lambda p, i: f"=-'{OM}'!{L(p['col'])}{row_depl}+{Praw('lease_depr_amt', p)}"
    is_line("memo: (+) DD&A", "USD mn", NUM_MN,
            hist=hist_fin("memo: (+) Depreciation, depletion & amortization"),
            qp=da_f, yb="sum", ys=da_f)
    ebitda_f = lambda p, i: f"={L(p['col'])}{rows_is['(=) EBIT']}+{L(p['col'])}{rows_is['memo: (+) DD&A']}"
    is_line("memo: (=) EBITDA", "USD mn", NUM_MN,
            hist=lambda p: f"={L(p['col'])}{rows_is['(=) EBIT']}+{L(p['col'])}{rows_is['memo: (+) DD&A']}",
            qp=ebitda_f, yb="sum", ys=ebitda_f)
    ebitdaal_f = lambda p, i: (f"={L(p['col'])}{rows_is['memo: (=) EBITDA']}-{Praw('lease_depr_amt', p)}"
                               f"+Schedules!{L(p['col'])}{rows_sc['lease']['interest']}")
    is_line("memo: (=) EBITDA-AL (after leases)", "USD mn", NUM_MN, qp=ebitdaal_f, yb="sum", ys=ebitdaal_f)
    mg_f = lambda p, i: (f'=IFERROR({L(p["col"])}{rows_is["memo: (=) EBITDA"]}'
                         f'/{L(p["col"])}{rows_is["(=) Net revenue"]},"-")')
    ri = add_line(is_, ri, "EBITDA margin", "%", PCT, tl,
                  hist=lambda p: mg_f(p, 0), qp=mg_f, yb=mg_f, ys=mg_f)
    NI_ROW = rows_is["(=) Net income"]

    # ------------------------------------------------ Working capital schedule (drives BS + CF)
    rev_is, cogs_is = rows_is["(=) Net revenue"], rows_is["(-) Cost of goods sold"]
    r = section(sc, r, "WORKING CAPITAL (days-driven)", tl)
    wc = {}

    def wc_line(label, key, base):
        nonlocal r
        wc[label] = r
        base_ref = ((lambda p: f"IS!{L(p['col'])}{rev_is}") if base == "rev"
                    else (lambda p: f"(-IS!{L(p['col'])}{cogs_is})"))
        f = lambda p, i: f"={Praw(key, p)}*{base_ref(p)}/{Praw('days', p)}"
        r = add_line(sc, r, label, "USD mn", NUM_MN, tl, hist=hist_fin(label),
                     qp=f, yb="eop", ys=f)

    wc_line("Trade receivables", "dso", "rev")
    wc_line("Inventories", "invdays", "cost")
    wc_line("Other current assets", "ocadays", "rev")
    wc_line("Trade payables", "dpo", "cost")
    wc_line("Other current liabilities", "ocldays", "rev")
    netwc_row = r
    netwc_f = lambda p, i: (
        f"={L(p['col'])}{wc['Trade receivables']}+{L(p['col'])}{wc['Inventories']}"
        f"+{L(p['col'])}{wc['Other current assets']}-{L(p['col'])}{wc['Trade payables']}"
        f"-{L(p['col'])}{wc['Other current liabilities']}")
    r = add_line(sc, r, "(=) Net working capital", "USD mn", NUM_MN, tl,
                 hist=lambda p: netwc_f(p, 0), qp=netwc_f, yb=netwc_f, ys=netwc_f)
    chg_row = r

    def wc_chg(p, i):
        pv = prev_period(tl, i)
        return f"={L(pv['col'])}{netwc_row}-{L(p['col'])}{netwc_row}" if pv else "=0"

    r = add_line(sc, r, "(+/-) Change in WC (cash impact)", "USD mn", NUM_MN, tl,
                 qp=wc_chg, yb="sum", ys=wc_chg)
    r += 1
    rows_sc["wc"] = dict(ar=wc["Trade receivables"], inv=wc["Inventories"],
                         oca=wc["Other current assets"], ap=wc["Trade payables"],
                         ocl=wc["Other current liabilities"], netwc=netwc_row, chg=chg_row)

    # ---------------------------------------------------------- Schedules (RE and cash)
    r, rows_sc["re"] = corkscrew(
        r, "RETAINED EARNINGS ROLL-FORWARD", seed_re,
        plus=lambda p, i: f"=IS!{L(p['col'])}{NI_ROW}",
        minus=lambda p, i: f"=-IS!{L(p['col'])}{NI_ROW}*{Praw('payout', p)}")

    # ---------------------------------------------------------- CF
    cf = sheet_shell(wb, "CF", "Cash flow: reconciliation", tl)
    rc = HDR + 2
    rc = section(cf, rc, "CASH FLOW STATEMENT", tl)
    rows_cf = {}
    def cf_line(label, src_label, qp=None, ys=None, yb="sum"):
        nonlocal rc
        rows_cf[label] = rc
        rc = add_line(cf, rc, label, "USD mn", NUM_MN, tl,
                      hist=hist_fin(src_label) if src_label else None,
                      qp=qp, yb=yb, ys=ys if ys else qp)
    cf_line("(=) Net income", "(=) Net income (CF basis)",
            qp=lambda p, i: f"=IS!{L(p['col'])}{NI_ROW}")
    cf_line("(+) D&A", "(+) D&A and non-cash items",
            qp=lambda p, i: f"=IS!{L(p['col'])}{rows_is['memo: (+) DD&A']}+{Praw('aro_accr', p)}")
    cf_line("(+/-) Working capital changes", "(+/-) Working capital changes",
            qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['wc']['chg']}")
    cfo_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_cf['(=) Net income']}:"
                          f"{L(p['col'])}{rows_cf['(+/-) Working capital changes']})")
    cf_line("(=) Cash from operations", "(=) Cash from operations", qp=cfo_f)
    cf_line("(-) Capex", "(-) Capex", qp=lambda p, i: f"=-'{OM}'!{L(p['col'])}{row_capext}")
    cf_line("(-) ARO settlements", None, qp=lambda p, i: f"=-{Praw('aro_settle', p)}")
    cfi_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_cf['(-) Capex']}:"
                          f"{L(p['col'])}{rows_cf['(-) ARO settlements']})")
    cf_line("(=) Cash from investing", "(=) Cash from investing", qp=cfi_f)
    cf_line("(-) Dividends", "(-) Dividends and buybacks",
            qp=lambda p, i: f"=-IS!{L(p['col'])}{NI_ROW}*{Praw('payout', p)}")
    cf_line("(+/-) Net debt drawdowns / (repayments)", None,
            qp=lambda p, i: (f"=Schedules!{L(p['col'])}{rows_sc['debt']['plus']}"
                             f"+Schedules!{L(p['col'])}{rows_sc['debt']['minus']}"))
    cf_line("(-) Lease principal payments", None,
            qp=lambda p, i: f"=-{Praw('lease_princ_amt', p)}")
    cff_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_cf['(-) Dividends']}:"
                          f"{L(p['col'])}{rows_cf['(-) Lease principal payments']})")
    cf_line("(=) Cash from financing", "(=) Cash from financing", qp=cff_f)
    dch_f = lambda p, i: (f"={L(p['col'])}{rows_cf['(=) Cash from operations']}"
                          f"+{L(p['col'])}{rows_cf['(=) Cash from investing']}"
                          f"+{L(p['col'])}{rows_cf['(=) Cash from financing']}")
    cf_line("(=) Net change in cash", "(=) Net change in cash", qp=dch_f)

    # cash + revolver block (revolver keeps cash >= minimum; interest on the drawn revolver
    # is a follow-up refinement — omitted here so the balance check stays exact and circular-free)
    r = section(sc, r, "CASH & REVOLVER", tl)
    seed_cash = f"='Input Financials'!{L(fhdr[last])}{frow['Cash and equivalents']}"
    cb_row, nc_row, cbf_row = r, r + 1, r + 2
    rb_row, dr_row, sw_row, re_row, ce_row = r + 3, r + 4, r + 5, r + 6, r + 7

    def roll_bop(bop_row, eop_row, seed):
        def f(p, i):
            pv = prev_period(tl, i)
            if pv is None or pv["kind"] == "QH":
                return seed
            if p["kind"] == "YB":
                return f"={L(block_cols(tl, p)[0])}{bop_row}"
            return f"={L(pv['col'])}{eop_row}"
        return f

    eop_yb = lambda row: (lambda p, i: f"={L(block_cols(tl, p)[-1])}{row}")
    add_line(sc, cb_row, "Cash BOP", "USD mn", NUM_MN, tl, qp=roll_bop(cb_row, ce_row, seed_cash),
             yb=roll_bop(cb_row, ce_row, seed_cash), ys=roll_bop(cb_row, ce_row, seed_cash))
    nc_f = lambda p, i: f"=CF!{L(p['col'])}{rows_cf['(=) Net change in cash']}"
    add_line(sc, nc_row, "(+) Net change in cash (pre-revolver)", "USD mn", NUM_MN, tl,
             qp=nc_f, yb="sum", ys=nc_f)
    cbf_f = lambda p, i: f"={L(p['col'])}{cb_row}+{L(p['col'])}{nc_row}"
    add_line(sc, cbf_row, "(=) Cash before revolver", "USD mn", NUM_MN, tl,
             qp=cbf_f, yb=cbf_f, ys=cbf_f)
    add_line(sc, rb_row, "Revolver BOP", "USD mn", NUM_MN, tl, qp=roll_bop(rb_row, re_row, "=0"),
             yb=roll_bop(rb_row, re_row, "=0"), ys=roll_bop(rb_row, re_row, "=0"))
    dr_f = lambda p, i: f"=MAX(0,{Praw('min_cash', p)}-{L(p['col'])}{cbf_row})"
    add_line(sc, dr_row, "(+) Revolver draw", "USD mn", NUM_MN, tl, qp=dr_f, yb="sum", ys=dr_f)
    sw_f = lambda p, i: f"=-MIN({L(p['col'])}{rb_row},MAX(0,{L(p['col'])}{cbf_row}-{Praw('min_cash', p)}))"
    add_line(sc, sw_row, "(-) Revolver sweep", "USD mn", NUM_MN, tl, qp=sw_f, yb="sum", ys=sw_f)
    re_f = lambda p, i: f"={L(p['col'])}{rb_row}+{L(p['col'])}{dr_row}+{L(p['col'])}{sw_row}"
    add_line(sc, re_row, "(=) Revolver EOP", "USD mn", NUM_MN, tl, qp=re_f, yb=eop_yb(re_row), ys=re_f)
    ce_f = lambda p, i: f"={L(p['col'])}{cbf_row}+{L(p['col'])}{dr_row}+{L(p['col'])}{sw_row}"
    add_line(sc, ce_row, "(=) Cash EOP", "USD mn", NUM_MN, tl, qp=ce_f, yb=eop_yb(ce_row), ys=ce_f)
    r = ce_row + 2
    rows_sc["cash"] = dict(bop=cb_row, eop=ce_row)
    rows_sc["rev"] = dict(bop=rb_row, eop=re_row)

    def cash_line(label, key):
        nonlocal rc
        rows_cf[label] = rc
        def yb_cash(p, i):
            bc = block_cols(tl, p)
            cidx = bc[0] if key == "bop" else bc[-1]
            return f"={L(cidx)}{rows_cf[label]}"
        rc = add_line(cf, rc, label, "USD mn", NUM_MN, tl,
                      hist=hist_fin("Cash — beginning of period" if key == "bop" else "Cash — end of period"),
                      qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['cash'][key]}",
                      yb=yb_cash,
                      ys=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['cash'][key]}")
    cash_line("Cash — beginning of period", "bop")
    cash_line("Cash — end of period", "eop")

    # ---------------------------------------------------------- BS
    bs = sheet_shell(wb, "BS", "Integrated balance sheet", tl)
    rb_ = HDR + 2
    rb_ = section(bs, rb_, "BALANCE SHEET", tl)
    rows_bs = {}
    def bs_line(label, qp=None, ys=None):
        nonlocal rb_
        rows_bs[label] = rb_
        flat_q = lambda p, i: f"={L(prev_period(tl, i)['col'])}{rows_bs[label]}"
        rb_ = add_line(bs, rb_, label, "USD mn", NUM_MN, tl,
                       hist=hist_fin(label), qp=qp or flat_q, yb="eop", ys=ys or qp or flat_q)
    bs_line("Cash and equivalents",
            qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['cash']['eop']}",
            ys=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['cash']['eop']}")
    bs_line("Short-term investments")
    wcref = lambda key: (lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['wc'][key]}")
    bs_line("Trade receivables", qp=wcref("ar"), ys=wcref("ar"))
    bs_line("Inventories", qp=wcref("inv"), ys=wcref("inv"))
    bs_line("Other current assets", qp=wcref("oca"), ys=wcref("oca"))
    tca_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_bs['Cash and equivalents']}:"
                          f"{L(p['col'])}{rows_bs['Other current assets']})")
    bs_line("(=) Total current assets", qp=tca_f, ys=tca_f)
    bs_line("PP&E", qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['ppe']['eop']}",
            ys=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['ppe']['eop']}")
    bs_line("Intangible assets and goodwill")
    bs_line("Right-of-use assets",
            qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['rou']['eop']}",
            ys=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['rou']['eop']}")
    bs_line("Deferred tax assets")
    bs_line("Other non-current assets")
    tnca_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_bs['PP&E']}:"
                           f"{L(p['col'])}{rows_bs['Other non-current assets']})")
    bs_line("(=) Total non-current assets", qp=tnca_f, ys=tnca_f)
    ta_f = lambda p, i: (f"={L(p['col'])}{rows_bs['(=) Total current assets']}"
                         f"+{L(p['col'])}{rows_bs['(=) Total non-current assets']}")
    bs_line("(=) Total assets", qp=ta_f, ys=ta_f)
    debt_cur = lambda p, i: (f"=Schedules!{L(p['col'])}{rows_sc['debt']['eop']}*{Praw('debt_curr_pct', p)}"
                             f"+Schedules!{L(p['col'])}{rows_sc['rev']['eop']}")
    debt_ncur = lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['debt']['eop']}*(1-{Praw('debt_curr_pct', p)})"
    lease_cur = lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['lease']['eop']}*{Praw('lease_curr_pct', p)}"
    lease_ncur = lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['lease']['eop']}*(1-{Praw('lease_curr_pct', p)})"
    bs_line("Trade payables", qp=wcref("ap"), ys=wcref("ap"))
    bs_line("Loans and financing (current)", qp=debt_cur, ys=debt_cur)
    bs_line("Lease liabilities (current)", qp=lease_cur, ys=lease_cur)
    bs_line("Other current liabilities", qp=wcref("ocl"), ys=wcref("ocl"))
    tcl_f = lambda p, i: (f"=SUM({L(p['col'])}{rows_bs['Trade payables']}:"
                          f"{L(p['col'])}{rows_bs['Other current liabilities']})")
    bs_line("(=) Total current liabilities", qp=tcl_f, ys=tcl_f)
    bs_line("Loans and financing (non-current)", qp=debt_ncur, ys=debt_ncur)
    bs_line("Lease liabilities (non-current)", qp=lease_ncur, ys=lease_ncur)
    bs_line("Provisions (incl. asset retirement obligation)",
            qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['aro']['eop']}",
            ys=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['aro']['eop']}")
    for lbl in ["Deferred tax liabilities", "Other non-current liabilities"]:
        bs_line(lbl)
    tl_f = lambda p, i: (f"={L(p['col'])}{rows_bs['(=) Total current liabilities']}"
                         f"+SUM({L(p['col'])}{rows_bs['Loans and financing (non-current)']}:"
                         f"{L(p['col'])}{rows_bs['Other non-current liabilities']})")
    bs_line("(=) Total liabilities", qp=tl_f, ys=tl_f)
    bs_line("Share capital and reserves")
    bs_line("Retained earnings (accumulated)",
            qp=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['re']['eop']}",
            ys=lambda p, i: f"=Schedules!{L(p['col'])}{rows_sc['re']['eop']}")
    eq_f = lambda p, i: (f"={L(p['col'])}{rows_bs['Share capital and reserves']}"
                         f"+{L(p['col'])}{rows_bs['Retained earnings (accumulated)']}")
    bs_line("(=) Equity attributable to shareholders", qp=eq_f, ys=eq_f)
    bs_line("Minority interest (equity)")
    teq_f = lambda p, i: (f"={L(p['col'])}{rows_bs['(=) Equity attributable to shareholders']}"
                          f"+{L(p['col'])}{rows_bs['Minority interest (equity)']}")
    bs_line("(=) Total equity", qp=teq_f, ys=teq_f)
    chk_f = lambda p, i: (f"={L(p['col'])}{rows_bs['(=) Total assets']}"
                          f"-{L(p['col'])}{rows_bs['(=) Total liabilities']}"
                          f"-{L(p['col'])}{rows_bs['(=) Total equity']}")
    rows_bs["check"] = rb_
    rb_ = add_line(bs, rb_, "Balance check (= 0)", "USD mn", NUM_2D, tl,
                   hist=lambda p: chk_f(p, 0), qp=chk_f, yb=chk_f, ys=chk_f)

    # ---------------------------------------------------------- Valuation (DCF)
    add_valuation_tab(wb, tl, hist_q, rows_is, rows_bs, rows_sc, rows_cf,
                      shares=gf("memo: Shares outstanding (EOP)", 0.0))

    # ---------------------------------------------------------- Cover
    cov = wb.create_sheet("Cover", 0)
    cov.sheet_properties.tabColor = "1F3864"
    cov.column_dimensions["B"].width = 100
    notes = [
        (fin["B1"].value, TITLE),
        ("Model generated by engine/build_model.py (engine v0). History linked to the input tabs; "
         "projections driven by the Premises tab (auto-seeds to be replaced in the assumption session).", SUB),
        ("", None),
        (f"History: {hist_q[0]} to {hist_q[-1]}. Projection: quarterly through 4Q{str(q_seq(hist_q[-1])[0] + 1)[2:]}, "
         f"annual through {q_seq(hist_q[-1])[0] + 1 + PROJ_YEARS_ANNUAL}.", None),
        ("Dynamic schedules: working capital (DSO/DPO days), debt (interest on the opening balance), "
         "lease IFRS-16 (RoU + liability roll-forward, EBITDA-AL memo), ARO accretion, and a simple "
         "revolver/cash-sweep keeping cash above a minimum. Lease and ARO seed neutral and are "
         "activated in the assumption session.", None),
        ("Valuation tab: FCFF DCF with WACC (Kd linked to the debt schedule), Gordon terminal value, "
         "EV→equity bridge (leases as debt), implied value per share, multiples, bull/base/bear "
         "scenarios and a WACC×g sensitivity grid. Analytical output only — never a recommendation.", None),
        ("Blue = input/assumption | Black = formula | Green = link between tabs | Cream fill = projection.", None),
        ("This file may contain synthetic test data; check the title of the input tabs.", None),
    ]
    for k, (txt, f) in enumerate(notes, start=2):
        c = cov.cell(row=k, column=2, value=txt)
        if f:
            c.font = f
        c.alignment = Alignment(wrap_text=True, vertical="top")

    order = ["Cover", "IS", "BS", "CF", "Valuation", "Operational Model", "Schedules", "Premises",
             "Assumptions", "Input Financials", "Input Operational", "README"]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames] + \
                 [ws for ws in wb._sheets if ws.title not in order]
    wb.save(out_path)

    # Write diagnostic log alongside the model file
    log_path = str(Path(out_path).parent / (Path(out_path).stem + "_build_log.txt"))
    diag.log(f"History  : {hist_q[0]}..{hist_q[-1]}")
    diag.log(f"Assets   : {len(assets)}")
    diag.log(f"Periods  : {len(tl)}")
    diag.write(log_path)

    print(f"OK: {out_path} | hist {hist_q[0]}..{hist_q[-1]} | {len(assets)} assets | "
          f"{len(tl)} period columns | log: {log_path}")


if __name__ == "__main__":
    # Usage: python build_model.py <input.xlsx> <output.xlsx> [sector]
    # Sector omitted → auto-detect from the input (oil_and_gas falls out of its signals).
    main(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
