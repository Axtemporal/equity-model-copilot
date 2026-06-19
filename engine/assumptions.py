"""Assumptions plumbing: the premise log as the source of truth.

The YAML log is authoritative; the model's Premises and Assumptions tabs are renderings of
it. On (re)build the engine seeds every projected premise from the last historical value;
this layer then OVERRIDES each cell that has an Approved assumption (merge by line + period)
— seeds survive only where no decision exists. Implements the "approved assumptions are the
source of truth on rebuild" decision, and is the data layer the Phase-B web UI will drive
(propose -> discuss -> approve -> next).
"""
from __future__ import annotations

import datetime
import os
import re
from dataclasses import asdict, dataclass, field
from enum import Enum

import openpyxl
import yaml

from .harness.invariants import HDR, LABEL_COL, find_label_row, period_columns

ASSUMPTIONS_HEADER_ROW = 4  # data rows start just below, in the Assumptions tab


class AssumptionStatus(str, Enum):
    PROPOSED = "Proposta"
    APPROVED = "Aprovada"
    REJECTED = "Rejeitada"


# Compliance fields every approved estimate must carry (method + source + date).
PROVENANCE_FIELDS = ("method", "source", "source_date")


class AssumptionProvenanceError(ValueError):
    """Raised when approving a premise without method + source + date.

    The method+source gate as an ENGINE INVARIANT (not a conversation convention): an estimate
    cannot be approved into the model unlabeled. See CLAUDE.md compliance + flow §Fase 1.
    """


def _missing_provenance(entry) -> list[str]:
    return [f for f in PROVENANCE_FIELDS if not str(getattr(entry, f, "") or "").strip()]


@dataclass
class Assumption:
    """One premise decision. `values` maps a period label (e.g. '2029') to its number."""
    line_item: str
    values: dict = field(default_factory=dict)
    method: str = ""
    source: str = ""
    source_date: str = ""
    rationale: str = ""
    status: str = AssumptionStatus.PROPOSED.value
    decided_by: str = ""
    decided_date: str = ""
    tab: str = "Premises"
    id: str = ""


class AssumptionsLog:
    """A versionable YAML log of premise decisions, keyed by line item."""

    def __init__(self, path: str | None = None, assumptions: list[Assumption] | None = None):
        self.path = path
        self.assumptions: list[Assumption] = assumptions or []

    @classmethod
    def load(cls, path: str | None) -> "AssumptionsLog":
        if path and os.path.exists(path):
            with open(path, encoding="utf-8") as handle:
                data = yaml.safe_load(handle) or {}
            items = [Assumption(**entry) for entry in data.get("assumptions", [])]
            return cls(path, items)
        return cls(path, [])

    def save(self, path: str | None = None) -> str:
        path = path or self.path
        if not path:
            raise ValueError("no path to save the assumptions log")
        payload = {"assumptions": [self._normalise(asdict(a)) for a in self.assumptions]}
        with open(path, "w", encoding="utf-8") as handle:
            yaml.safe_dump(payload, handle, allow_unicode=True, sort_keys=False)
        return path

    @staticmethod
    def _normalise(entry: dict) -> dict:
        entry["values"] = {str(k): v for k, v in (entry.get("values") or {}).items()}
        return entry

    def get(self, line_item: str) -> Assumption | None:
        return next((a for a in self.assumptions if a.line_item == line_item), None)

    def _next_id(self) -> str:
        return f"A{len(self.assumptions) + 1:03d}"

    def upsert(
        self,
        line_item: str,
        values: dict,
        *,
        method: str | None = None,
        source: str | None = None,
        source_date: str | None = None,
        rationale: str | None = None,
        status: str | None = None,
        decided_by: str | None = None,
        decided_date: str | None = None,
    ) -> Assumption:
        """Add a new premise or merge new per-period values into an existing one."""
        values = {str(k): v for k, v in values.items()}
        existing = self.get(line_item)
        if existing is None:
            existing = Assumption(line_item=line_item, values=values, id=self._next_id(),
                                  status=status or AssumptionStatus.PROPOSED.value)
            self.assumptions.append(existing)
        else:
            existing.values.update(values)
            if status is not None:
                existing.status = status
        for attr, val in (("method", method), ("source", source), ("source_date", source_date),
                          ("rationale", rationale), ("decided_by", decided_by),
                          ("decided_date", decided_date)):
            if val is not None:
                setattr(existing, attr, val)
        return existing

    def approve(self, line_item: str, *, decided_by: str = "analyst",
                decided_date: str | None = None) -> Assumption:
        entry = self.get(line_item)
        if entry is None:
            raise KeyError(f"no assumption for line item {line_item!r}")
        missing = _missing_provenance(entry)
        if missing:
            raise AssumptionProvenanceError(
                f"cannot approve {line_item!r} without {', '.join(missing)} — compliance "
                f"requires every estimate to carry method + source + date"
            )
        entry.status = AssumptionStatus.APPROVED.value
        entry.decided_by = decided_by
        entry.decided_date = decided_date or datetime.date.today().isoformat()
        return entry

    def reject(self, line_item: str, *, decided_by: str = "analyst",
               decided_date: str | None = None) -> Assumption:
        entry = self.get(line_item)
        if entry is None:
            raise KeyError(f"no assumption for line item {line_item!r}")
        entry.status = AssumptionStatus.REJECTED.value
        entry.decided_by = decided_by
        entry.decided_date = decided_date or datetime.date.today().isoformat()
        return entry

    def approved_overrides(self) -> dict[tuple[str, str], object]:
        """{(line_item, period): value} for every Approved assumption — the rebuild merge key."""
        out: dict[tuple[str, str], object] = {}
        for entry in self.assumptions:
            if entry.status == AssumptionStatus.APPROVED.value:
                for period, value in entry.values.items():
                    out[(entry.line_item, str(period))] = value
        return out


def provenance_violations(log: AssumptionsLog) -> list[str]:
    """Approved assumptions missing method/source/date — the data for a harness FAIL invariant.

    The approve() gate prevents these at write time; this is the read-side audit (e.g. for an
    older log or a hand-edited YAML) so the harness can FAIL a model whose log slipped through.
    """
    out: list[str] = []
    for entry in log.assumptions:
        if entry.status == AssumptionStatus.APPROVED.value:
            missing = _missing_provenance(entry)
            if missing:
                out.append(f"{entry.line_item}: missing {', '.join(missing)}")
    return out


def apply_to_model(model_path: str, log: AssumptionsLog) -> int:
    """Override the Premises tab with Approved assumptions and render the Assumptions tab.

    Returns the number of (line, period) cells overridden. The model is saved in place; run
    the harness afterwards to confirm the overrides did not break statement integration.
    """
    wb = openpyxl.load_workbook(model_path)
    applied = 0

    if "Premises" in wb.sheetnames:
        prem = wb["Premises"]
        period_col = {label: col for col, label in period_columns(prem)}
        for (line_item, period), value in log.approved_overrides().items():
            row = find_label_row(prem, line_item)
            col = period_col.get(period)
            if row is not None and col is not None:
                prem.cell(row, col).value = value
                applied += 1

    _render_assumptions_tab(wb, log)
    wb.save(model_path)
    return applied


def _render_assumptions_tab(wb, log: AssumptionsLog) -> None:
    if "Assumptions" not in wb.sheetnames:
        return
    ws = wb["Assumptions"]
    start = ASSUMPTIONS_HEADER_ROW + 1
    for row in range(start, ws.max_row + 1):  # clear any previous render
        for col in range(2, 14):
            ws.cell(row, col).value = None
    for i, entry in enumerate(log.assumptions):
        row = start + i
        periods = ", ".join(entry.values.keys())
        value_str = "; ".join(f"{p}={v}" for p, v in entry.values.items())
        for col, value in enumerate(
            [entry.id, entry.decided_date, entry.tab, entry.line_item, periods, value_str,
             entry.method, entry.source, entry.source_date, entry.rationale, entry.status,
             entry.decided_by],
            start=2,
        ):
            ws.cell(row, col).value = value


def read_premises(model_path: str) -> list[dict]:
    """Read the Premises tab into a list of {line_item, unit, periods: [(label, value)]}.

    The list of premise lines the assumption session walks through, with their current values.
    """
    wb = openpyxl.load_workbook(model_path)
    if "Premises" not in wb.sheetnames:
        return []
    prem = wb["Premises"]
    cols = period_columns(prem)
    lines = []
    for row in range(HDR + 1, prem.max_row + 1):
        label = prem.cell(row, LABEL_COL).value
        if not isinstance(label, str) or not label.strip():
            continue
        unit = prem.cell(row, 3).value
        periods = [(plabel, prem.cell(row, col).value) for col, plabel in cols]
        lines.append({"line_item": label.strip(), "unit": unit, "periods": periods})
    return lines


_NUMBER_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def parse_adjustment(message: str, valid_periods) -> dict[str, float]:
    """Interpret a chat message like '2029 para 70' or '2029 = 70' into {period: value}.

    The offline editor for the session chat: for every valid period named in the message, take
    the nearest number that is not itself another period. Lets the chat drive value changes
    without an LLM; a connected Claude can replace/augment this with free-form reasoning.
    """
    text = str(message)
    low = text.lower()
    valid = set(valid_periods)
    updates: dict[str, float] = {}
    for period in valid_periods:
        idx = low.find(period.lower())
        if idx < 0:
            continue
        for segment in (text[idx + len(period):], text[:idx]):
            picked = None
            for match in _NUMBER_RE.finditer(segment):
                token = match.group()
                if token in valid:  # don't read another period label as the value
                    continue
                picked = float(token.replace(",", "."))
                break
            if picked is not None:
                updates[period] = picked
                break
    return updates
