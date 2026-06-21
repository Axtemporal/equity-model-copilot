"""Fase 1 intake orchestrator: flatten + resolve + classify + content-sufficiency end to end."""
import openpyxl
import pytest

from engine import canonical_schema as cs
from engine import fase1_intake as intake
from engine import role_classifier as rc
from engine.fase1_manifest import AI_WEB_FETCH, MAPPED, LineRecord, Manifest
from engine.harness.invariants import find_label_row, period_columns


def _messy_input(rows, op_signals=("Brent average", "Sales volume", "PRODUCTION BY ASSET"),
                 periods=("1Q22", "2Q22", "", "2022")):
    """A non-canonical input: PT labels, a blank gap column, O&G signals on the op tab."""
    wb = openpyxl.Workbook()
    fin = wb.active
    fin.title = "Input Financials"
    fin.cell(1, 2, "ACME Co")
    fin.cell(5, 2, "Line item")
    fin.cell(5, 3, "Unit")
    for j, p in enumerate(periods):
        if p:
            fin.cell(5, 4 + j, p)
    for i, (label, unit, vals) in enumerate(rows):
        r = 6 + i
        fin.cell(r, 2, label)
        if unit:
            fin.cell(r, 3, unit)
        for j, v in enumerate(vals):
            if v is not None:
                fin.cell(r, 4 + j, v)
    op = wb.create_sheet("Input Operational")
    op.cell(5, 2, "Line item")
    op.cell(5, 4, "1Q22")
    for i, label in enumerate(op_signals):
        op.cell(6 + i, 2, label)
        op.cell(6 + i, 4, 80)
    return wb


def test_analyze_synth_detects_sector_and_is_sufficient(synth_inputs):
    m = intake.analyze(synth_inputs)
    assert m.sector == cs.OIL_AND_GAS
    # every required canonical slot is satisfied by a mapped line from the real input
    assert m.content_sufficient, [c.slot for c in m.missing_required]


def test_analyze_maps_core_lines(synth_inputs):
    m = intake.analyze(synth_inputs)
    by_canonical = {line.canonical: line for line in m.lines if line.disposition == MAPPED}
    assert "(=) Net revenue" in by_canonical
    assert by_canonical["(=) Net revenue"].role == rc.STATEMENT
    # an operational driver is tagged operational-raw (deferred to Fase 2/3)
    assert any(line.canonical == "Brent average" and line.role == rc.OPERATIONAL_RAW
               for line in m.lines)


def test_every_line_has_a_terminal_disposition(synth_inputs):
    m = intake.analyze(synth_inputs)
    assert all(line.disposition != "pending" for line in m.lines)
    # unresolved lines are carried as contextual, never dropped silently
    assert all((line.canonical is not None) or (line.raw_label in m.residual) for line in m.lines)


def test_manifest_roundtrips_to_yaml(synth_inputs, tmp_path):
    m = intake.analyze(synth_inputs)
    path = m.save(str(tmp_path / "fase1_manifest.yaml"))
    import yaml
    data = yaml.safe_load(open(path, encoding="utf-8"))
    assert data["sector"] == cs.OIL_AND_GAS
    assert data["content_sufficient"] is True


def test_write_adjusted_input_canonicalizes_a_messy_input(tmp_path):
    wb = _messy_input([
        ("Receita líquida", "R$ mn", [100, 110, None, 210]),
        ("Lucro líquido", None, [10, 11, None, 21]),
        ("Imobilizado", None, [500, 510, None, 510]),
        ("Caixa e equivalentes de caixa", None, [50, 55, None, 55]),
    ])
    src = str(tmp_path / "messy.xlsx")
    wb.save(src)
    out = str(tmp_path / "adjusted.xlsx")
    intake.build_adjusted_input(src, out)

    adj = openpyxl.load_workbook(out)
    fin = adj["Input Financials"]
    # PT labels are now canonical
    assert find_label_row(fin, "(=) Net revenue") is not None
    assert find_label_row(fin, "PP&E") is not None
    # the blank gap column is gone and periods are in canonical order
    assert [lbl for _c, lbl in period_columns(fin)] == ["1Q22", "2Q22", "2022"]
    # mapped values land on the right canonical line/period
    rev = find_label_row(fin, "(=) Net revenue")
    cols = {lbl: c for c, lbl in period_columns(fin)}
    assert fin.cell(rev, cols["1Q22"]).value == 100
    assert fin.cell(rev, cols["2022"]).value == 210
    # every required financial label exists as a row (so the engine cannot KeyError)
    present = {fin.cell(r, 2).value for r in range(1, fin.max_row + 1)
               if isinstance(fin.cell(r, 2).value, str)}
    for label in cs.required_labels(cs.OIL_AND_GAS)[cs.FINANCIALS]:
        assert label in present, label
    # the operational sheet is carried as-is
    assert find_label_row(adj["Input Operational"], "Brent average") is not None


def test_write_adjusted_input_rolls_up_sublines(tmp_path):
    # two raw lines that resolve to the SAME canonical are summed into one row
    wb = _messy_input([
        ("Receita líquida", "R$ mn", [100]),
        ("Receita operacional líquida", "R$ mn", [5]),
    ], periods=("1Q22",))
    src = str(tmp_path / "messy.xlsx")
    wb.save(src)
    out = str(tmp_path / "adjusted.xlsx")
    intake.build_adjusted_input(src, out)

    fin = openpyxl.load_workbook(out)["Input Financials"]
    rev = find_label_row(fin, "(=) Net revenue")
    cols = {lbl: c for c, lbl in period_columns(fin)}
    assert fin.cell(rev, cols["1Q22"]).value == 105


def test_fase1_gate_passes_on_a_sufficient_input(synth_inputs, tmp_path):
    out = str(tmp_path / "adjusted.xlsx")
    _path, manifest = intake.build_adjusted_input(synth_inputs, out)
    assert manifest.content_sufficient
    assert intake.assert_fase1_pass(out, manifest) is True


def test_fase1_gate_blocks_on_missing_slots(tmp_path):
    wb = _messy_input([("Receita líquida", "R$ mn", [100])], periods=("1Q22",))
    src = str(tmp_path / "messy.xlsx")
    wb.save(src)
    out = str(tmp_path / "adjusted.xlsx")
    _p, manifest = intake.build_adjusted_input(src, out)   # writes rows so the build won't KeyError
    assert not manifest.content_sufficient                 # ...but the data isn't there
    with pytest.raises(intake.Fase1GateError) as exc:
        intake.assert_fase1_pass(out, manifest)
    assert any("missing required slot" in p for p in exc.value.problems)


def test_assess_currency_flags_stale():
    m = Manifest(company="X", sector=cs.OIL_AND_GAS)
    cur = intake.assess_currency(m, "3Q25", "1Q26", source_url="http://ri.example/q1")
    assert cur["stale"] and cur["missing_periods"] == ["4Q25", "1Q26"]
    assert m.currency["source_url"] == "http://ri.example/q1"


def _fetched_shares(accepted, **prov):
    return LineRecord(raw_label="Shares", sheet="Input Financials",
                      canonical="memo: Shares outstanding (EOP)", method=AI_WEB_FETCH,
                      role=rc.CAPITAL_STRUCTURE, disposition=MAPPED, values={"1Q22": 1000.0},
                      user_accepted=accepted, **prov)


def test_writer_rejects_web_fetch_without_provenance(tmp_path):
    wb = _messy_input([("Receita líquida", "R$ mn", [100])], periods=("1Q22",))
    src = str(tmp_path / "m.xlsx")
    wb.save(src)
    m = intake.analyze(src)
    m.lines.append(_fetched_shares(accepted=False))        # fetched, not accepted, no source
    with pytest.raises(intake.ProvenanceError):
        intake.write_adjusted_input(m, src, str(tmp_path / "out.xlsx"))


def test_writer_accepts_web_fetch_with_full_provenance(tmp_path):
    wb = _messy_input([("Receita líquida", "R$ mn", [100])], periods=("1Q22",))
    src = str(tmp_path / "m.xlsx")
    wb.save(src)
    m = intake.analyze(src)
    m.lines.append(_fetched_shares(accepted=True, source_url="http://b3.example", source_date="2026-05"))
    out = str(tmp_path / "out.xlsx")
    intake.write_adjusted_input(m, src, out)               # no raise
    fin = openpyxl.load_workbook(out)["Input Financials"]
    assert find_label_row(fin, "memo: Shares outstanding (EOP)") is not None
