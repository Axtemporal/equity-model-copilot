"""Geometry / layout checks that need no recalc (always run)."""
import openpyxl

from engine.harness.invariants import CORE_SHEETS, find_label_row


def test_core_sheets_present(synth_model):
    wb = openpyxl.load_workbook(synth_model)
    for sheet in CORE_SHEETS:
        assert sheet in wb.sheetnames, f"missing sheet {sheet!r}"


def test_balance_check_row_exists(synth_model):
    wb = openpyxl.load_workbook(synth_model)
    assert find_label_row(wb["BS"], "balance check") is not None


def test_no_external_links(synth_model):
    wb = openpyxl.load_workbook(synth_model)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str):
                    assert "[" not in cell.value, f"external link in {ws.title}!{cell.coordinate}"
