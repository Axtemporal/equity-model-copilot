"""A non-O&G sector builds (and balances) through the generic top-down path.

Proves the engine's financial reads are sector-AGNOSTIC: an input carrying only the universal
base lines (no O&G-specific COGS / Exploration / ARM / DD&A-memo) builds instead of KeyError-ing.
The O&G path is unchanged (guarded by the golden + invariant suites).
"""
import openpyxl
import pytest

from engine import build_model
from engine.harness import verify_model


def _generic_input(path):
    wb = openpyxl.Workbook()
    fin = wb.active
    fin.title = "Input Financials"
    fin.cell(1, 2, "GenCo")
    fin.cell(5, 2, "Line item")
    fin.cell(5, 3, "Unit")
    periods = ["1Q22", "2Q22", "3Q22", "4Q22", "2022"]
    for j, p in enumerate(periods):
        fin.cell(5, 4 + j, p)
    # a footing BS (Assets 330 = Liab 130 + Equity 200) + a minimal IS — NO O&G-specific lines
    rows = {
        "(=) Net revenue": [50, 50, 50, 50, 200], "(=) EBIT": [10, 10, 10, 10, 40],
        "(=) EBT": [8, 8, 8, 8, 32], "(-) Income taxes": [-2, -2, -2, -2, -8],
        "(=) Net income": [6, 6, 6, 6, 24],
        "Cash and equivalents": [100] * 5, "Trade receivables": [20] * 5, "Inventories": [10] * 5,
        "Other current assets": [0] * 5, "PP&E": [200] * 5, "Intangible assets and goodwill": [0] * 5,
        "Right-of-use assets": [0] * 5, "Deferred tax assets": [0] * 5, "Other non-current assets": [0] * 5,
        "Short-term investments": [0] * 5, "Trade payables": [10] * 5, "Other current liabilities": [0] * 5,
        "Loans and financing (current)": [0] * 5, "Loans and financing (non-current)": [120] * 5,
        "Lease liabilities (current)": [0] * 5, "Lease liabilities (non-current)": [0] * 5,
        "Deferred tax liabilities": [0] * 5, "Other non-current liabilities": [0] * 5,
        "Share capital and reserves": [150] * 5, "Retained earnings (accumulated)": [50] * 5,
    }
    r = 6
    for label, vals in rows.items():
        fin.cell(r, 2, label)
        for j, v in enumerate(vals):
            fin.cell(r, 4 + j, v)
        r += 1
    op = wb.create_sheet("Input Operational")
    op.cell(5, 2, "Line item")
    op.cell(5, 3, "Unit")
    for j, p in enumerate(periods):
        op.cell(5, 4 + j, p)
    op.cell(6, 2, "Employees")
    for j in range(5):
        op.cell(6, 4 + j, 1000)
    wb.save(path)
    return path


def test_non_og_sector_builds_without_keyerror(tmp_path):
    src = _generic_input(str(tmp_path / "gen.xlsx"))
    out = str(tmp_path / "gen_model.xlsx")
    build_model.main(src, out, sector=None)        # would KeyError before the None-safe reads
    assert "BS" in openpyxl.load_workbook(out).sheetnames


@pytest.mark.recalc
def test_non_og_generic_build_balances(require_recalc, tmp_path):
    src = _generic_input(str(tmp_path / "gen.xlsx"))
    out = str(tmp_path / "gen_model.xlsx")
    build_model.main(src, out)
    report = verify_model(out)
    assert report.ok, report.summary()
