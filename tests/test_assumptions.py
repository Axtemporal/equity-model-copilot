"""Assumptions plumbing: YAML round-trip, approved override on rebuild, model still balances."""
import shutil

import openpyxl
import pytest

from engine.assumptions import (
    AssumptionsLog,
    AssumptionStatus,
    apply_to_model,
    parse_adjustment,
    read_premises,
)
from engine.harness import verify_model
from engine.harness.invariants import find_label_row, period_columns


def test_parse_adjustment():
    valid = ["2Q26", "2029", "2030"]
    assert parse_adjustment("muda o Brent de 2029 para 70", valid) == {"2029": 70.0}
    assert parse_adjustment("2030 = 75", valid) == {"2030": 75.0}
    assert parse_adjustment("coloca 72 em 2029", valid) == {"2029": 72.0}
    assert parse_adjustment("sem ajuste aqui", valid) == {}


def test_read_premises(synth_model):
    lines = read_premises(synth_model)
    items = [line["line_item"] for line in lines]
    assert "Brent average" in items
    brent = next(line for line in lines if line["line_item"] == "Brent average")
    assert any(period == "2029" for period, _value in brent["periods"])


def test_log_roundtrip(tmp_path):
    path = str(tmp_path / "log.yaml")
    log = AssumptionsLog(path)
    log.upsert("Brent average", {"2029": 70}, method="forward curve", source="EIA STEO",
               source_date="2026-06", rationale="mean reversion")
    log.approve("Brent average", decided_date="2026-06-15")
    log.save()

    reloaded = AssumptionsLog.load(path)
    assert reloaded.get("Brent average").status == AssumptionStatus.APPROVED.value
    assert reloaded.approved_overrides()[("Brent average", "2029")] == 70


def _premises_cell(model_path, line_item, period):
    wb = openpyxl.load_workbook(model_path)
    prem = wb["Premises"]
    row = find_label_row(prem, line_item)
    col = {label: c for c, label in period_columns(prem)}[period]
    return row, col, prem.cell(row, col).value


@pytest.mark.recalc
def test_approved_override_applies_and_balances(synth_model, require_recalc, tmp_path):
    model = str(tmp_path / "SYNTH_model.xlsx")
    shutil.copy(synth_model, model)

    _row, _col, seed = _premises_cell(model, "Brent average", "2029")
    assert seed != 70  # we are genuinely changing it

    log = AssumptionsLog()
    log.upsert("Brent average", {"2029": 70}, method="forward curve",
               source="EIA STEO", source_date="2026-06")
    log.approve("Brent average", decided_date="2026-06-15")
    applied = apply_to_model(model, log)
    assert applied == 1

    _row, _col, new_value = _premises_cell(model, "Brent average", "2029")
    assert new_value == 70

    assum = openpyxl.load_workbook(model)["Assumptions"]
    assert any(assum.cell(r, 5).value == "Brent average" for r in range(5, assum.max_row + 1))

    report = verify_model(model)
    assert report.ok, report.summary()
