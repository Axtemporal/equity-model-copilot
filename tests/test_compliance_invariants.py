"""Compliance invariants: no-target-price disclaimer + approved-assumption provenance.

Fast (no recalc) — exercises the invariant functions directly on tiny workbooks. The positive
path on the real synth model is covered by test_invariants.test_synth_passes_all_invariants.
"""
import openpyxl

from engine.harness.invariants import inv_assumptions_provenance, inv_valuation_disclaimer


def _sheet(title: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def test_per_share_without_disclaimer_fails():
    wb, ws = _sheet("Valuation")
    ws.cell(1, 2, "(=) Implied value per share")
    assert not inv_valuation_disclaimer(wb, wb).passed


def test_per_share_with_disclaimer_passes():
    wb, ws = _sheet("Valuation")
    ws.cell(1, 2, "(=) Implied value per share")
    ws.cell(2, 2, "Analytical output — NOT a price target or investment recommendation.")
    assert inv_valuation_disclaimer(wb, wb).passed


def test_prose_mentioning_per_share_is_not_flagged():
    wb, ws = _sheet("Cover")
    ws.cell(1, 2, "Valuation tab: FCFF DCF, implied value per share, multiples")  # prose, not a label
    assert inv_valuation_disclaimer(wb, wb).passed


def test_approved_without_provenance_fails():
    wb, ws = _sheet("Assumptions")
    ws.cell(5, 5, "Brent average")
    ws.cell(5, 12, "Aprovada")  # status approved, method/source blank
    assert not inv_assumptions_provenance(wb, wb).passed


def test_approved_with_provenance_passes():
    wb, ws = _sheet("Assumptions")
    ws.cell(5, 5, "Brent average")
    ws.cell(5, 8, "forward curve")
    ws.cell(5, 9, "EIA STEO")
    ws.cell(5, 12, "Aprovada")
    assert inv_assumptions_provenance(wb, wb).passed


def test_proposed_row_needs_no_provenance():
    wb, ws = _sheet("Assumptions")
    ws.cell(5, 5, "Brent average")
    ws.cell(5, 12, "Proposta")
    assert inv_assumptions_provenance(wb, wb).passed
