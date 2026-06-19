"""Pre-build input validator — positive and negative cases (no recalc, always runs)."""
import openpyxl
import pytest

from engine.harness.invariants import find_label_row
from engine.harness.validator import (
    InputValidationError,
    assert_valid_input,
    validate_input,
)


def test_valid_synth_input_passes(synth_inputs):
    report = validate_input(synth_inputs)
    assert report.ok, [r.message for r in report.failures()]


def test_typo_in_required_label_is_caught(synth_inputs):
    wb = openpyxl.load_workbook(synth_inputs)
    fin = wb["Input Financials"]
    row = find_label_row(fin, "(=) Net revenue")
    fin.cell(row, 2).value = "(=) Net rev"  # typo the engine would choke on
    report = validate_input(wb)
    assert not report.ok
    assert any("Net revenue" in m for m in [r.message for r in report.failures()])


def test_missing_sheet_raises(synth_inputs):
    wb = openpyxl.load_workbook(synth_inputs)
    wb.remove(wb["Input Operational"])
    with pytest.raises(InputValidationError):
        assert_valid_input(wb)
