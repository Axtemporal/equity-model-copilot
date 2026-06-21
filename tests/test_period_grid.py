"""Period-grid normalization: tolerant parsing + canonical interleaved reorder."""
import openpyxl

from engine import period_grid as pg


def test_parse_quarter_spellings():
    for raw in ("1Q23", "1Q2023", "1T23", "1 Q 23", "1Q'23", "Q1 2023", "Q1'23", "T1 2023"):
        p = pg.parse_period(raw)
        assert p is not None and p.token == "1Q23", raw
        assert p.year == 2023 and p.quarter == 1


def test_parse_annual_spellings():
    for raw in ("2023", "FY2023", "FY 2023", "2023A", "12M23", "FY23"):
        p = pg.parse_period(raw)
        assert p is not None and p.token == "2023" and p.is_annual, raw


def test_non_periods_return_none():
    for raw in ("Line item", "Unit", "", None, "TTM", "1H23", "23", "Receita"):
        assert pg.parse_period(raw) is None, raw


def test_reorder_quarters_blank_then_annuals():
    # the messy layout: all quarters, a blank separator column, then all annuals
    headers = [
        (1, "Line item", True), (2, "Unit", True),
        (3, "1Q22", True), (4, "2Q22", True), (5, "3Q22", True), (6, "4Q22", True),
        (7, "1Q23", True), (8, "2Q23", True), (9, "3Q23", True), (10, "4Q23", True),
        (11, "", False),                                   # blank gap
        (12, "2022", True), (13, "2023", True),
    ]
    g = pg.normalize_grid(headers)
    assert g.order == ["1Q22", "2Q22", "3Q22", "4Q22", "2022",
                       "1Q23", "2Q23", "3Q23", "4Q23", "2023"]
    assert g.col_to_token[3] == "1Q22" and g.col_to_token[12] == "2022"
    assert g.dropped == [""]          # blank separator removed
    assert g.flagged == []            # leading label/unit columns ignored, not flagged


def test_unparseable_header_with_data_is_flagged_not_dropped():
    headers = [(1, "Line item", True), (2, "1Q23", True), (3, "TTM", True), (4, "2023", True)]
    g = pg.normalize_grid(headers)
    assert g.flagged == ["TTM"]
    assert g.order == ["1Q23", "2023"]


def test_flatten_reorders_a_messy_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Input Financials"
    cols = ["1Q22", "2Q22", "3Q22", "4Q22", "1T23", "2Q23", "3Q23", "4Q23", "", "2022", "2023"]
    ws.cell(5, 2, "Line item")
    ws.cell(5, 3, "Unit")
    for i, label in enumerate(cols):
        ws.cell(5, 4 + i, label)
    ws.cell(6, 2, "(=) Net revenue")
    ws.cell(6, 3, "R$ mn")
    for i, label in enumerate(cols):
        if label:                                          # leave the blank gap empty
            ws.cell(6, 4 + i, 100 + i)

    from engine.fase1_intake import flatten
    rec = next(r for r in flatten(wb) if r.raw_label == "(=) Net revenue")
    assert list(rec.values.keys()) == ["1Q22", "2Q22", "3Q22", "4Q22", "2022",
                                       "1Q23", "2Q23", "3Q23", "4Q23", "2023"]
    assert rec.unit == "R$ mn"
