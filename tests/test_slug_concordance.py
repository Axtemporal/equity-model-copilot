"""Sector-slug concordance — the three detectors must agree on ONE canonical slug.

Regression guard for the 'oil_and_gas' vs 'oil_gas' divergence: validator.detect_sector,
sector_knowledge.detect_sector_from_input, and template_loader.identify_sector must all
return the same canonical slug, and it must be the one the build gate / pipeline branch on.
"""
import openpyxl

from engine import canonical_schema as cs
from engine import sector_knowledge
from engine.harness import validator
from engine.template_loader import identify_sector


def _op_labels(path):
    wb = openpyxl.load_workbook(path)
    op = wb["Input Operational"]
    return {
        op.cell(r, 2).value.strip(): r
        for r in range(1, op.max_row + 1)
        if isinstance(op.cell(r, 2).value, str) and op.cell(r, 2).value.strip()
    }


def test_all_detectors_agree_on_canonical_slug(synth_inputs):
    wb = openpyxl.load_workbook(synth_inputs)
    by_validator = validator.detect_sector(wb)
    by_knowledge = sector_knowledge.detect_sector_from_input(synth_inputs)
    by_loader = identify_sector(_op_labels(synth_inputs))

    assert by_validator == by_knowledge == by_loader == cs.OIL_AND_GAS


def test_canonical_slug_is_the_validator_required_key(synth_inputs):
    # the detected slug must be a key the validator can look up (no KeyError downstream)
    wb = openpyxl.load_workbook(synth_inputs)
    assert validator.detect_sector(wb) in validator.REQUIRED_LABELS
