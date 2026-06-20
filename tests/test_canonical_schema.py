"""Canonical schema — the universal base + per-sector delta is the single source of truth,
and what the validator requires must equal what the engine actually hard-reads from the input.
"""
import openpyxl

from engine import canonical_schema as cs


def test_known_sectors_are_disk_driven():
    # the set of sectors is whatever has a delta on disk — nothing hardcoded
    sectors = cs.known_sectors()
    assert sectors, "expected at least one sector with a delta"
    assert all(cs.has_delta(s) for s in sectors)
    assert sectors == tuple(sorted(sectors))


def test_required_labels_is_base_plus_delta():
    base = set(cs.UNIVERSAL_BASE_FINANCIAL)
    for sector in cs.known_sectors():
        req = cs.required_labels(sector)
        fin = set(req[cs.FINANCIALS])
        assert base <= fin, f"{sector} financial req must include the universal base"
        # the delta adds sector-specific lines on top of the base
        assert fin - base == set(cs.load_delta(sector).get(cs.FINANCIALS, []))
        assert req[cs.OPERATIONAL]  # both pilots disclose operational drivers


def test_oil_and_gas_delta_has_sector_specifics():
    fin = cs.required_labels(cs.OIL_AND_GAS)[cs.FINANCIALS]
    assert "(-) Exploration expenses" in fin
    assert "(+/-) Financial result" in fin            # O&G spelling
    assert "(+/-) Net financial result" not in fin    # telecom spelling stays out


def test_telecom_delta_has_sector_specifics():
    fin = cs.required_labels(cs.TELECOM)[cs.FINANCIALS]
    assert "(=) EBITDA (as disclosed)" in fin
    assert "(+/-) Net financial result" in fin        # telecom spelling
    assert "(-) Exploration expenses" not in fin      # O&G line stays out


def test_synth_input_satisfies_oil_and_gas_required(synth_inputs):
    """The real guarantee: every label the schema marks required for O&G is present in a real
    input — so the stricter validator does not false-positive on a valid workbook."""
    wb = openpyxl.load_workbook(synth_inputs)
    req = cs.required_labels(cs.OIL_AND_GAS)
    for sheet, labels in req.items():
        present = {
            wb[sheet].cell(r, 2).value.strip()
            for r in range(1, wb[sheet].max_row + 1)
            if isinstance(wb[sheet].cell(r, 2).value, str) and wb[sheet].cell(r, 2).value.strip()
        }
        missing = [lbl for lbl in labels if lbl not in present]
        assert not missing, f"SYNTH missing required {sheet} labels: {missing}"
